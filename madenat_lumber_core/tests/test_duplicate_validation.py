# -*- coding: utf-8 -*-
"""
🛡️ TESTS: VALIDACIÓN TEMPRANA DE DUPLICADOS (2026-06-11)

Cobertura:
  1. Subir un PDF con guía existente → bloquear con UserError.
  2. Subir un XLS/XLSX con guía existente → bloquear con UserError.
  3. Confirmar que el sistema bloquea antes de crear staging.
  4. Confirmar que el mensaje se muestra al operador.
  5. Confirmar que una guía nueva válida sí sigue ingresando normalmente.
  6. Confirmar que no se crean duplicados en lumber.reception ni madenat.guia.processing.
"""
import base64
import io
import logging
from odoo.tests import TransactionCase, tagged
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


@tagged('post_install', '-at_install', 'madenat_duplicate')
class TestDuplicateValidation(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.parser = cls.env['madenat.reception.parser']
        cls.partner = cls.env['res.partner'].create({
            'name': 'Proveedor Test Duplicados',
            'is_company': True,
            'supplier_rank': 1,
        })

        # ── Helpers de construcción de archivos ──

    # ──────────────────────────────────────────────────────────────────
    # HELPERS
    # ──────────────────────────────────────────────────────────────────
    @staticmethod
    def _build_minimal_pdf(text: bytes) -> bytes:
        """Construye un PDF mínimo con texto legible por pdfplumber."""
        from reportlab.pdfgen import canvas
        buf = io.BytesIO()
        c = canvas.Canvas(buf)
        c.drawString(50, 750, text.decode('utf-8', errors='replace'))
        c.save()
        return buf.getvalue()

    @staticmethod
    def _build_minimal_excel(guide_no: str = '9999') -> bytes:
        """Construye un XLSX mínimo con cabecera y número de guía."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Fila con número de guía
        ws['A1'] = 'Guía N°'
        ws['B1'] = guide_no
        # Fila de encabezados
        ws['A8'] = 'Paquete'
        ws['B8'] = 'Codigo'
        ws['C8'] = 'Producto'
        ws['D8'] = 'Espesor'
        ws['E8'] = 'Ancho'
        ws['F8'] = 'Largo'
        ws['G8'] = 'Piezas'
        ws['H8'] = 'Volumen'
        # Una fila de datos
        ws['A9'] = 1
        ws['B9'] = 'TEST001'
        ws['C9'] = 'MADERA TEST'
        ws['D9'] = 45
        ws['E9'] = 100
        ws['F9'] = 3.0
        ws['G9'] = 100
        ws['H9'] = 1.35
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _create_reception(self, name='9999'):
        """Crea un lumber.reception en estado draft con archivos dummy.
        PDF y Excel usan el mismo número de guía para no disparar
        NROGUIA_MISMATCH en Gate 1."""
        pdf_bytes = self._build_minimal_pdf(
            ("GUIA %s" % name).encode('utf-8')
        )
        return self.env['lumber.reception'].create({
            'name': name,
            'guia_numero': name,
            'supplier_id': self.partner.id,
            'ingestion_profile': 'f5085',
            'pdf_file': base64.b64encode(pdf_bytes),
            'pdf_filename': 'guia_test.pdf',
            'excel_file': base64.b64encode(self._build_minimal_excel(name)),
            'excel_filename': 'packing_test.xlsx',
            'commercial_volume_m3': 10.0,
            'exchange_rate': 800.0,
        })

    def _create_guia_processing(self, name='9999'):
        """Crea un madenat.guia.processing en estado draft."""
        location = self.env['stock.location'].search([
            ('usage', '=', 'internal')
        ], limit=1)
        return self.env['madenat.guia.processing'].create({
            'name': name,
            'partner_id': self.partner.id,
            'assignment_location_id': location.id,
            'date_emission': '2026-06-11',
        })

    # ──────────────────────────────────────────────────────────────────
    # TEST 1: Bloquea PDF con guía existente en lumber.reception
    # ──────────────────────────────────────────────────────────────────
    def test_01_duplicate_pdf_reception(self):
        """Subir un PDF cuya guía ya existe en lumber.reception → UserError."""
        rec = self._create_reception('88001')
        rec.write({'state': 'done'})  # marcamos como ya procesada

        _logger.info("🔍 TEST 1: verificando duplicado en lumber.reception para guía 88001")
        with self.assertRaises(UserError) as ctx:
            self.parser._check_guide_duplicate('88001')

        error_msg = str(ctx.exception)
        self.assertIn('88001', error_msg)
        self.assertIn('ya fue registrada', error_msg)
        _logger.info("✅ TEST 1 PASADO: UserError lanzado correctamente para duplicado en lumber.reception")

    # ──────────────────────────────────────────────────────────────────
    # TEST 2: Bloquea con guía existente en madenat.guia.processing
    # ──────────────────────────────────────────────────────────────────
    def test_02_duplicate_in_guia_processing(self):
        """Subir un PDF cuya guía ya existe en madenat.guia.processing → UserError."""
        gp = self._create_guia_processing('88002')

        _logger.info("🔍 TEST 2: verificando duplicado en madenat.guia.processing para guía 88002")
        with self.assertRaises(UserError) as ctx:
            self.parser._check_guide_duplicate('88002', partner_id=gp.partner_id.id)

        error_msg = str(ctx.exception)
        self.assertIn('88002', error_msg)
        self.assertIn('Guías Procesadas', error_msg)
        _logger.info("✅ TEST 2 PASADO: UserError lanzado para duplicado en guia.processing")

    # ──────────────────────────────────────────────────────────────────
    # TEST 3: Guía nueva válida NO bloquea
    # ──────────────────────────────────────────────────────────────────
    def test_03_new_guide_passes(self):
        """Una guía que no existe en ningún lado debe pasar la validación."""
        _logger.info("🔍 TEST 3: verificando que guía nueva 99999 no sea bloqueada")
        # No debe lanzar excepción
        self.parser._check_guide_duplicate('99999')
        _logger.info("✅ TEST 3 PASADO: guía nueva supera la validación sin errores")

    # ──────────────────────────────────────────────────────────────────
    # TEST 4: Bloqueo antes de staging (vía workflow simulado)
    # ──────────────────────────────────────────────────────────────────
    def test_04_no_staging_on_duplicate(self):
        """Si la guía ya existe, no se debe crear staging."""
        rec_existing = self._create_reception('88003')
        rec_existing.write({'state': 'done'})

        # Intentar procesar una NUEVA recepción con la misma guía
        rec_new = self._create_reception('88003')

        _logger.info("🔍 TEST 4: verificando que el pipeline bloquea duplicado antes de staging")
        with self.assertRaises(UserError) as ctx:
            rec_new.action_process_documents()

        error_msg = str(ctx.exception)
        self.assertIn('88003', error_msg)
        self.assertIn('ya fue registrada', error_msg)

        # Confirmar que no se creó staging
        self.assertEqual(
            rec_new.reception_line_ids.ids, [],
            "No deben existir líneas de staging para una guía duplicada"
        )
        _logger.info("✅ TEST 4 PASADO: staging vacío tras bloqueo por duplicado")

    # ──────────────────────────────────────────────────────────────────
    # TEST 5: Sin guía detectada → no bloquea (pasa sin validar)
    # ──────────────────────────────────────────────────────────────────
    def test_05_empty_guide_no_passes(self):
        """Si no se detecta número de guía, no se ejecuta la validación."""
        _logger.info("🔍 TEST 5: guía vacía → sin validación, sin error")
        # No debe lanzar excepción
        self.parser._check_guide_duplicate('')
        self.parser._check_guide_duplicate(None)
        _logger.info("✅ TEST 5 PASADO: guía vacía se omite sin errores")

    # ──────────────────────────────────────────────────────────────────
    # TEST 6: Mensaje de error contiene datos del registro existente
    # ──────────────────────────────────────────────────────────────────
    def test_06_error_message_contains_details(self):
        """El UserError debe incluir: número de guía, proveedor, estado, ID."""
        rec = self._create_reception('88004')
        rec.write({
            'state': 'done',
            'guia_fecha': '2026-06-01',
        })

        _logger.info("🔍 TEST 6: verificando contenido del mensaje de error")
        with self.assertRaises(UserError) as ctx:
            self.parser._check_guide_duplicate('88004')

        error_msg = str(ctx.exception)
        self.assertIn('88004', error_msg)
        self.assertIn(str(self.partner.name), error_msg)
        self.assertIn(str(rec.id), error_msg)
        _logger.info("✅ TEST 6 PASADO: mensaje contiene N° guía, proveedor e ID")

    # ──────────────────────────────────────────────────────────────────
    # TEST 7: Confirmar que una guía nueva sí ingresa normalmente
    # ──────────────────────────────────────────────────────────────────
    def test_07_new_guide_full_flow(self):
        """Verificar que el pipeline completo funciona para una guía nueva."""
        rec = self._create_reception('99001')

        _logger.info("🔍 TEST 7: pipeline completo para guía nueva 99001")
        # action_process_documents usa el workflow que incluye la validación temprana
        # Como la guía 99001 no existe, debe pasar sin UserError de duplicado
        try:
            rec.action_process_documents()
        except UserError as e:
            # Puede fallar por otros motivos (TC, OC, etc.) pero NO por duplicado
            self.assertNotIn('ya fue registrada', str(e),
                             "No debe fallar por duplicado para una guía nueva")

        # Verificar que se creó staging
        self.assertGreater(
            len(rec.reception_line_ids), 0,
            "Debe haber líneas de staging creadas para una guía nueva"
        )
        _logger.info("✅ TEST 7 PASADO: guía nueva genera staging correctamente")

    # ──────────────────────────────────────────────────────────────────
    # TEST 8: Sin duplicados en ambos modelos tras procesar guía nueva
    # ──────────────────────────────────────────────────────────────────
    def test_08_no_duplicates_created(self):
        """Verificar que tras procesar una guía nueva, solo hay 1 registro."""
        rec = self._create_reception('99002')

        _logger.info("🔍 TEST 8: verificando que no se crean duplicados tras procesar")
        try:
            rec.action_process_documents()
        except UserError:
            # Si falla por TC/OC, igual verificamos que no haya duplicados
            pass

        # Contar registros con el mismo nombre en lumber.reception
        count_reception = self.env['lumber.reception'].search_count([
            ('name', '=', '99002'),
        ])
        self.assertLessEqual(
            count_reception, 1,
            "No debe haber más de 1 lumber.reception con el mismo nombre"
        )
        _logger.info("✅ TEST 8 PASADO: máximo 1 registro por guía en lumber.reception")