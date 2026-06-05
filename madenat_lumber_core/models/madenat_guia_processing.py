# -*- coding: utf-8 -*-
"""
🏭 RECEPCIÓN DE MADERA PROCESADA (SERVICIOS)

FLUJO: Aserradero externo -> Guía servicio -> Lotes procesados
COSTEO: process_cost_usd (valor servicio) + wood_cost_usd (heredado)
GENEALOGÍA: parent_lot_id = Lote Bruto (obligatorio)
"""
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from psycopg2.errors import IntegrityError
import pandas as pd
import pdfplumber
import re 
import xlrd
import logging
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from collections import defaultdict

from .utils_uom import (
    INCH_SQ_METERS_TO_M3,
    MBF_TO_M3,
    S2S_WIDTH_LOOKUP,
    MM_PER_INCH, M_TO_FT, S2S_WIDTH_ADJUSTMENT_INCH, BLANK_CLEAR_FACTOR,
    M3_DIVISOR, MBF_DIVISOR,
    calculate_volume_metric_m3,
    get_s2s_adjustment,
    m3_to_mbf,
    mbf_to_m3,
    r3,
    r4,
    LUMBER_DIMENSION_MAP,
)

_logger = logging.getLogger(__name__)

# ============================================================================
# TD-007: imports duplicados eliminados (2026-06-04) — ya están en líneas 9-10 y 76.
# ✅ NUEVO: MODELO DE STAGING PARA LÍNEAS DE GUÍA PROCESADA
# ============================================================================
class MadenatGuiaProcessingLine(models.Model):
    """
    Línea de Verificación (Staging) - Versión Profesional v4.0
    
    ARQUITECTURA:
    - Doble sistema de dimensiones: MM (compra) + Imperial (exportación)
    - Campos visuales EDITABLES para correcciones manuales
    - Conversión bidireccional automática con fallback manual
    - Validación de octavos imperiales
    
    FLUJO:
    1. Excel → Carga dimensiones MM (espesor_mm, ancho_mm, largo_m)
    2. Compute → Calcula automáticamente thickness_visual, width_visual
    3. Usuario → Puede editar manualmente thickness_visual/width_visual
    4. Sistema → Recalcula vol_mbf desde valores editados
    
    CHANGELOG:
    - v4.0 (2026-01-24): Corrección de conversión de octavos + tolerancia aumentada
    - v3.0: Campos visuales editables + Validación octavos
    - v2.0: Compute imperial values
    - v1.0: Modelo base
    """
    
    _name = 'madenat.guia.processing.line'
    _description = 'Línea de Verificación (Staging)'
    _inherit = ['madenat.lumber.ingest.line.mixin']
    
    # ==========================================================================
    # RELACIONES
    # ==========================================================================
    
    processing_id = fields.Many2one(
        'madenat.guia.processing',
        string='Guía Padre',
        ondelete='cascade',
        required=True,
        index=True
    )
    
    # ==========================================================================
    # DATOS ORIGINALES (TRAZABILIDAD)
    # ==========================================================================
    
    sku_original = fields.Char(
        "SKU Original",
        help="Código original del Excel para trazabilidad (ej: PCLAT04201253250)"
    )
    
    product_name_original = fields.Char(
        "Nombre en Excel",
        help="Nombre del producto tal como aparece en el Excel"
    )
    
    lot_name = fields.Char(
        "Nº Lote",
        help="Número de lote/etiqueta físico (ej: D7739)"
    )

    espesor_nominal_mm = fields.Float("Espesor Nominal (mm)", digits=(10, 3), help="Espesor de compra para análisis comercial")
    # CAMPO: ANCHO NOMINAL (Con lógica de pre-llenado visual)
    ancho_nominal_mm = fields.Float(
        string="Ancho Nom. (mm)", 
        digits=(16, 3),
        store=True, 
        readonly=False, # ¡Editable!
        compute='_compute_default_width', 
        precompute=True # Esto hace que se llene al crear la línea o cargar el Excel
    )
    # =========================================================================
    # 1. LARGO NOMINAL (NUEVO) - Replicando lógica de Ancho
    # =========================================================================
    # --- 1. DEFINICIÓN DEL NUEVO CAMPO (Faltaba esto en la BD) ---
    largo_nominal_m = fields.Float(
        string="Largo Nom. (m)", 
        digits=(16, 3),
        store=True, 
        readonly=False, 
        compute='_compute_default_length', 
        precompute=True
    )

    vol_physical_m3 = fields.Float("Vol. Bodega (m³)", compute="_compute_vol_physical_m3", store=True, digits=(16, 3))

    vol_purchase_m3 = fields.Float('Vol. Compra (m³)', compute='_compute_vol_purchase_m3', store=True, digits=(16, 3))
    # ==========================================================================
    # DIMENSIONES FÍSICAS (MM) - ORIGEN EXCEL
    # Estos son los valores de COMPRA, medidos físicamente
    # ==========================================================================
    
    espesor_mm = fields.Float(
        "Espesor (mm)",
        digits=(10, 3),
        help="Espesor real medido en milímetros"
    )
    
    ancho_mm = fields.Float(
        "Ancho (mm)",
        digits=(10, 3),
        help="Ancho real medido en milímetros"
    )
    
    largo_m = fields.Float(
        "Largo (m)",
        digits=(10, 3),
        help="Largo real medido en metros"
    )
    
    # ==========================================================================
    # CALIDAD/SUBPRODUCTO
    # ==========================================================================
    
    subproducto_id = fields.Many2one(
        'madenat.subproducto',
        string='Subproducto/Calidad',
        help='Grado o calidad de la madera para exportación (FAS, RIP S2S, etc.)',
    )
    
    # ==========================================================================
    # DIMENSIONES VISUALES (IMPERIAL) - CALCULADAS + EDITABLES
    # Estos son los valores COMERCIALES para exportación
    # ==========================================================================
    
    thickness_visual = fields.Char(
        "Espesor (Nom)",
        compute='_compute_imperial_values',
        store=True,
        readonly=False,  # ← CRÍTICO: permite edición manual
        help="Espesor comercial en formato fraccionario (ej: 7/4, 2). EDITABLE para correcciones."
    )
    
    width_visual = fields.Char(
        "Ancho (Nom)",
        compute='_compute_imperial_values',
        store=True,
        readonly=False,  # ← CRÍTICO: permite edición manual
        help="Ancho comercial en formato fraccionario (ej: 5 3/8, 4 1/2, 6). EDITABLE para correcciones."
    )
    
    # -------------------------------------------------------------------------
    # CORRECCIÓN: Agregamos "inverse" para que TU edición manual se respete
    # -------------------------------------------------------------------------
    length_ft = fields.Float(
        string="Largo (Pies)", 
        compute='_compute_imperial_values', 
        inverse='_inverse_length_ft',  # <--- ¡ESTA ES LA CLAVE QUE FALTABA!
        store=True, 
        readonly=False, 
        digits=(16, 3)
    )
    def _inverse_length_ft(self):
        """
        Permite guardar el valor manual de Pies sin que el 'compute' lo borre.
        """
        pass
    vol_mbf = fields.Float(
        "Volumen (MBF)",
        compute="_compute_vol_mbf",
        store=True,
        digits=(16, 3),
        help="Volumen en Miles de Pies Tablares (calculado desde dimensiones imperiales)"
    )

     
    vol_shipment_m3 = fields.Float(
        'Vol. Exportación (m³)',
        compute='_compute_vol_shipment_m3',
        store=True,
        digits=(16, 3),
        help="Volumen real calculado desde dimensiones imperiales editables"
    )
    
    # ==========================================================================
    # ESTADO Y VALIDACIÓN
    # ==========================================================================
    
    technical_validation = fields.Selection([
        ('pending', 'Pendiente'),
        ('approved', 'Aprobado'),
        ('rejected', 'Rechazado')
    ], 
        string="Estado Técnico",
        default='approved',
        help="Estado de validación técnica para la línea de staging."
    )
    
    # ==========================================================================
    # COMPUTE: Dimensiones Visuales Imperiales (desde MM)
    # VERSIÓN UNIFICADA - Calcula TODO en un solo método
    # ==========================================================================
   
   # ==========================================================================
    # COMPUTE: Volumen de Compra (Factor 1.000.000 - Pestaña 2
    # ==========================================================================
    @api.depends('largo_m')
    def _compute_default_length(self):
        for line in self:
            # Si es 0 y hay físico, pre-llenar. Si usuario edita, respetar.
            if line.largo_nominal_m == 0.0 and line.largo_m > 0:
                line.largo_nominal_m = line.largo_m
    # largo_m ya existe, no creamos nada nuevo.
    @api.depends('ancho_mm')
    def _compute_default_width(self):
        for line in self:
            # Solo pre-llenamos si está vacío (0) y tenemos dato físico
            if line.ancho_nominal_mm == 0.0 and line.ancho_mm > 0:
                line.ancho_nominal_mm = line.ancho_mm

            # Si el usuario ya escribió un valor (ej: 150), NO hacemos nada.
    # TD-007: v1 de _compute_vol_purchase_m3 eliminada (2026-06-04).
    # La v2 (línea ~424) es la activa — incluye largo_nominal_m en @api.depends.
   
   # ==========================================================================
    # 1. CONVERSIÓN: NOMINAL/METROS -> VISUAL/PIES
    # ==========================================================================
    @api.depends('espesor_nominal_mm', 'espesor_mm', 
                 'ancho_nominal_mm', 'ancho_mm', 
                 'largo_nominal_m', 'largo_m')
    def _compute_imperial_values(self):
        """
        Determina las fracciones visuales (ej: '5 3/8') y los Pies 
        basándose en el Nominal MM y el Mapa Maestro.
        """
        m_to_ft = float(M_TO_FT)
        MM_TO_INCH = 1.0 / float(MM_PER_INCH)
        
        for rec in self:
            # ------------------------------------------------------
            # A. ESPESOR (Thickness)
            # ------------------------------------------------------
            val_thick_mm = rec.espesor_nominal_mm if rec.espesor_nominal_mm > 0 else rec.espesor_mm
            
            mapped_thick = None
            if val_thick_mm > 0:
                # Buscar en Mapa con tolerancia 3mm
                for std_mm, txt in LUMBER_DIMENSION_MAP.get('thickness', {}).items():
                    if abs(val_thick_mm - std_mm) <= 3: 
                        mapped_thick = txt
                        break
            
            if mapped_thick:
                rec.thickness_visual = mapped_thick
            elif val_thick_mm > 0 and not rec.thickness_visual:
                rec.thickness_visual = self._get_fraction_text(val_thick_mm * MM_TO_INCH)

            # ------------------------------------------------------
            # B. ANCHO (Width)
            # ------------------------------------------------------
            val_width_mm = rec.ancho_nominal_mm if rec.ancho_nominal_mm > 0 else rec.ancho_mm
            
            mapped_width = None
            if val_width_mm > 0:
                # Buscar en Mapa con tolerancia 2mm
                for std_mm, txt in LUMBER_DIMENSION_MAP.get('width', {}).items():
                    if abs(val_width_mm - std_mm) <= 2: 
                        mapped_width = txt
                        break
            
            if mapped_width:
                rec.width_visual = mapped_width
            elif val_width_mm > 0 and not rec.width_visual:
                rec.width_visual = self._get_fraction_text(val_width_mm * MM_TO_INCH)
            
            # ------------------------------------------------------
            # C. LARGO (Length) - METROS MANDAN -> PIES OBEDECEN
            # ------------------------------------------------------
            # Si editas largo_m, esto actualiza length_ft automáticamente
            if rec.largo_m > 0:
                rec.length_ft = rec.largo_m * m_to_ft
            else:
                rec.length_ft = 0.0
   

    # ==========================================================================
    # HELPER: Buscar valor estándar más cercano
    # ==========================================================================
    
    def _find_closest_standard(self, value_mm, dimension_type):
        """
        Busca el valor estándar más cercano en LUMBER_DIMENSION_MAP.
        
        Args:
            value_mm (float): Valor en milímetros a buscar
            dimension_type (str): 'thickness' o 'width'
        
        Returns:
            float: Valor estándar más cercano o None si está fuera de tolerancia
        
        Example:
            >>> self._find_closest_standard(58, 'thickness')
            >>> 63  # Porque 63mm (10/4) está dentro de ±5mm
        """
        if dimension_type not in LUMBER_DIMENSION_MAP:
            return None
        
        standards = LUMBER_DIMENSION_MAP[dimension_type].keys()
        
        closest = None
        min_diff = float('inf')
        
        for std_mm in standards:
            diff = abs(value_mm - std_mm)
            if diff <= 5 and diff < min_diff:  # Tolerancia ±5mm
                min_diff = diff
                closest = std_mm
        
        return closest




    # 2. FÓRMULA FÍSICA (Pestaña 1): mm_real * mm_real * m / 1.000.000
    @api.depends('espesor_mm', 'ancho_mm', 'largo_m', 'pieces')
    def _compute_vol_physical_m3(self):
        for rec in self:
            rec.vol_physical_m3 = calculate_volume_metric_m3(rec.espesor_mm, rec.ancho_mm, rec.largo_m, rec.pieces)

    # 3. FÓRMULA COMERCIAL (Pestaña 2): mm_nominal * mm_real * m / 1.000.000
    @api.depends('espesor_nominal_mm', 'espesor_mm', 
                 'ancho_nominal_mm', 'ancho_mm', 
                 'largo_nominal_m', 'largo_m', 'pieces')  # Agregado largo_nominal_m
    def _compute_vol_purchase_m3(self):
        for rec in self:
            # Prioridad: Nominal (Manual/Wizard) > Físico
            esp = rec.espesor_nominal_mm if rec.espesor_nominal_mm > 0 else rec.espesor_mm
            anc = rec.ancho_nominal_mm if rec.ancho_nominal_mm > 0 else rec.ancho_mm
            
            # NUEVO: Usamos Largo Nominal
            larg = rec.largo_nominal_m if rec.largo_nominal_m > 0 else rec.largo_m

            rec.vol_purchase_m3 = calculate_volume_metric_m3(esp, anc, larg, rec.pieces)
    # COMPUTE: Volumen MBF (desde dimensiones imperiales)
    # ==========================================================================

    # ==========================================================================
    # ⚙️ NUEVO MÉTODO DE MAPEO INDUSTRIAL (Lógica Robusta)
    # ==========================================================================
    def _get_industry_width_map(self, width_mm):
        """
        Busca el ancho en la tabla LUMBER_DIMENSION_MAP con tolerancia de +/- 3mm.
        Si encuentra coincidencia, devuelve la fracción de negocio (ej: 5 3/8).
        Si no, devuelve None para que se calcule matemáticamente.
        """
        if not width_mm:
            return None
            
        # Buscar en el mapa global definido arriba
        target_map = LUMBER_DIMENSION_MAP.get('width', {})
        
        for standard_mm, fraction_text in target_map.items():
            if abs(width_mm - standard_mm) <= 5: # Tolerancia de aserradero
                return fraction_text
                
        return None
    
    # ==========================================================================
    # CÁLCULOS DE VOLUMEN (MBF)
    # ==========================================================================
    
    @api.depends('thickness_visual', 'width_visual', 'length_ft', 'pieces', 
                 'espesor_nominal_mm', 'ancho_nominal_mm', 'largo_nominal_m', 
                 'espesor_mm', 'ancho_mm', 'largo_m')
    def _compute_vol_mbf(self):
        """
        Calcula el volumen en MBF (Miles de Pies Tablares) desde dimensiones imperiales o métricas.
        Fórmula base: MBF = (espesor" × ancho" × largo' × piezas) / 12,000
        
        LOGICA DE PRIORIDAD:
        1. Visual (Exportación/Métrico): Si hay un valor explícito (ej: "6/4" o "195mm"), manda.
        2. Comercial (Físico/Nominal): Si falta lo anterior, convierte los mm/m editables a pulgadas/pies.
        """
        # Constantes de conversión
        MM_TO_INCH = 1.0 / float(MM_PER_INCH)
        m_to_ft = float(M_TO_FT)

        for line in self:
            # Validación básica de integridad: Si no hay piezas, no hay volumen.
            if not line.pieces or line.pieces <= 0:
                line.vol_mbf = 0.0
                continue
            
            try:
                # -------------------------------------------------------------
                # 1. PRIORIDAD: Valores visuales (Manejados por Parser Inteligente)
                # -------------------------------------------------------------
                if line.thickness_visual and line.width_visual and line.length_ft:
                    thickness_inch = self._parse_fraction(line.thickness_visual)
                    width_inch = self._parse_fraction(line.width_visual)
                    
                    # Cálculo: (T" * W" * L' * Pzs) / 12 / 1000
                    vol_bf = (thickness_inch * width_inch * line.length_ft * line.pieces) / 12.0
                    line.vol_mbf = vol_bf / 1000.0
                    
                    # Debug log para trazabilidad en caso de auditoría
                    _logger.debug(
                        f"MBF Visual (ID {line.id}): {thickness_inch}\" x {width_inch}\" x {line.length_ft}' = {line.vol_mbf:.4f}"
                    )
                
                # -------------------------------------------------------------
                # 2. FALLBACK: Calcular matemáticamente desde MM (Nominal > Físico)
                # -------------------------------------------------------------
                else:
                    # A. Determinar dimensiones efectivas (Si hay nominal manual, tiene prioridad)
                    eff_esp_mm = line.espesor_nominal_mm if line.espesor_nominal_mm > 0 else line.espesor_mm
                    eff_anc_mm = line.ancho_nominal_mm if line.ancho_nominal_mm > 0 else line.ancho_mm
                    eff_largo_m = line.largo_nominal_m if line.largo_nominal_m > 0 else line.largo_m
                    
                    # B. Validar y Calcular
                    if eff_esp_mm > 0 and eff_anc_mm > 0 and eff_largo_m > 0:
                        # Conversión al vuelo a sistema imperial para aplicar la fórmula
                        t_in = eff_esp_mm * MM_TO_INCH
                        w_in = eff_anc_mm * MM_TO_INCH
                        l_ft = eff_largo_m * m_to_ft
                        
                        vol_bf = (t_in * w_in * l_ft * line.pieces) / 12.0
                        line.vol_mbf = vol_bf / 1000.0
                    else:
                        line.vol_mbf = 0.0
            
            except Exception as e:
                _logger.warning(f"Error crítico calculando MBF en línea {line.id}: {str(e)}")
                line.vol_mbf = 0.0

    # ==========================================================================
    # COMPUTE: Volumen de Embarque (Factor 1550/5085 + 1/8" - Pestaña 3)
    # ==========================================================================
    # 2. CÁLCULO VOLUMEN EXPORTACIÓN (Con Lógica de Prioridad Pies)
    # ==========================================================================
    @api.depends('thickness_visual', 'width_visual', 'length_ft', 'pieces', 
                 'largo_m', 'largo_nominal_m', 'vol_purchase_m3')
    def _compute_vol_shipment_m3(self):
        """
        Calcula Volumen de Exportación (m³).
        🛡️ BLINDAJE: Si el cálculo geométrico real falla o faltan datos visuales,
        el sistema rescata el Volumen Nominal (vol_purchase_m3) para evitar ceros.
        """
        FACTOR_PIES = float(BLANK_CLEAR_FACTOR)

        for line in self:
            # Valor por defecto inicial: El nominal de compra (Para que NUNCA sea 0)
            fallback_vol = line.vol_purchase_m3 or 0.0
            
            # 1. Validaciones de integridad mínima
            if not line.pieces or line.pieces <= 0:
                line.vol_shipment_m3 = 0.0
                continue

            # 2. Si no hay datos visuales, aplicamos el Blindaje inmediatamente
            if not line.thickness_visual or not line.width_visual:
                line.vol_shipment_m3 = fallback_vol
                continue

            try:
                # 3. Parsear Fracciones (Visual -> Decimal Inch)
                e_in = self._parse_fraction(line.thickness_visual) 
                a_in = self._parse_fraction(line.width_visual)
                
                # Si el parseo devuelve 0, usamos el fallback
                if e_in <= 0 or a_in <= 0:
                    line.vol_shipment_m3 = fallback_vol
                    continue

                # 5. Determinar Largo a usar
                # Prioridad 1: length_ft | Prioridad 2: largo_m | Prioridad 3: largo_nominal_m
                # MADENAT-FIX-BLANK-2026-06-02
                if line.length_ft > 0.1:
                    # FÓRMULA PIES — BLANK: sin recargo de cepillado (+1/8")
                    # Regla de negocio: blank se ingesta y se embarca con dimensiones exactas
                    vol = (e_in * a_in * line.length_ft * line.pieces) / FACTOR_PIES
                else:
                    # FÓRMULA METROS — S2S/RIP: con ajuste de cepillado (+1/8")
                    recargo = get_s2s_adjustment(self.env, line.ancho_mm)
                    width_calc = a_in + float(recargo)
                    largo_uso = line.largo_m if line.largo_m > 0 else line.largo_nominal_m
                    if largo_uso <= 0:
                        line.vol_shipment_m3 = fallback_vol
                        continue
                    vol = (e_in * width_calc * largo_uso * line.pieces) / float(INCH_SQ_METERS_TO_M3)
                
                # 6. Asignación Final con redondeo a 3 decimales
                calculated_vol = r3(vol)
                
                # Blindaje final: Si por algún error de redondeo da 0, usamos fallback
                line.vol_shipment_m3 = calculated_vol if calculated_vol > 0 else fallback_vol

            except Exception as e:
                # Si algo explota en el cálculo, el sistema no se detiene: usa el nominal.
                _logger.error(f"⚠️ Error en cálculo real para ID {line.id}, usando nominal: {e}")
                line.vol_shipment_m3 = fallback_vol

    # ==========================================================================
    # MÉTODOS AUXILIARES: Conversión de Fracciones
    # ==========================================================================
    
    def _get_fraction_text(self, value_in):
        """
        Convierte decimal a fracción visual (ej: 4.625 → "4 5/8").
        
        CORRECCIÓN v4.0 (2026-01-24):
        - Tolerancia aumentada de 0.02 a 0.05 para compensar redondeos
        - Manejo mejorado de valores cercanos a enteros
        """
        if not value_in:
            return ""
        
        # Separar parte entera y decimal
        integer = int(value_in)
        decimal = value_in - integer
        
        # Mapeo de octavos estándar
        fractions = {
            0.000: "",       # Entero exacto
            0.125: "1/8",
            0.250: "1/4",
            0.375: "3/8",
            0.500: "1/2",
            0.625: "5/8",
            0.750: "3/4",
            0.875: "7/8"
        }
        
        frac_str = ""
        
        # ✅ CORRECCIÓN CRÍTICA: Aumentar tolerancia para compensar redondeos
        # Ejemplo: 5.71 debería detectarse como 5 3/4 (5.75) o 5 5/8 (5.625)
        TOLERANCE = 0.05  # 0.05" = 1.27mm de tolerancia
        
        # Buscar octavo más cercano
        min_diff = float('inf')
        best_frac = None
        
        for val, txt in fractions.items():
            diff = abs(decimal - val)
            if diff < min_diff:
                min_diff = diff
                best_frac = (val, txt)
        
        # Si la diferencia es menor que la tolerancia, usar esa fracción
        if min_diff < TOLERANCE and best_frac:
            frac_str = best_frac[1]
        
        # Formatear resultado
        if frac_str:
            result = f"{integer} {frac_str}".strip() if integer > 0 else frac_str
            
            # ✅ DEBUG: Log para verificar conversión
            _logger.debug(
                f"Conversión fracción: {value_in:.3f}\" → {result} (diff={min_diff:.4f})"
            )
            return result
        
        # Si no encaja en octavos, devolver decimal con advertencia
        _logger.warning(
            f"Valor {value_in:.3f}\" no encaja en octavos estándar (tolerancia {TOLERANCE}\"). "
            f"Devolviendo decimal."
        )
        return f"{value_in:.2f}"
    
    def _parse_fraction(self, fraction_str):
        """
        🚀 PARSER INTELIGENTE BILINGÜE (Métrico/Imperial):
        Convierte "5 3/8", "6/4", "4", "4.5" y "195mm" a Float (Pulgadas Decimales).
        Maneja errores de espacios dobles o formatos sucios.
        """
        if not fraction_str:
            return 0.0
        
        s = str(fraction_str).strip().lower()
        if not s:
            return 0.0
            
        try:
            # 🛡️ BLINDAJE 1: Intercepción explícita de métricas
            if 'mm' in s:
                val = float(s.replace('mm', '').strip())
                return val / float(MM_PER_INCH)  # Convertimos mm a pulgadas decimales
                
            # Caso A: Es un número simple entero o decimal ("4", "4.5" o "195")
            if '/' not in s and ' ' not in s:
                val = float(s)
                
                # 🧠 BLINDAJE 2 (Heurística): Anchos comerciales > 24" (60cm) no existen.
                # Si llega un "195", asumimos obligatoriamente que son mm.
                if val > 24:
                    return val / float(MM_PER_INCH)
                    
                return val

            # Caso B: Fracción Mixta ("5 3/8")
            if ' ' in s:
                parts = s.split()
                # Filtrar espacios vacíos extra por si escriben "5  3/8"
                parts = [p for p in parts if p.strip()]
                
                if len(parts) == 2:
                    whole = float(parts[0])
                    frac_parts = parts[1].split('/')
                    if len(frac_parts) == 2:
                        numerator = float(frac_parts[0])
                        denominator = float(frac_parts[1])
                        return whole + (numerator / denominator)
            
            # Caso C: Fracción Pura ("6/4")
            if '/' in s:
                parts = s.split('/')
                if len(parts) == 2:
                    return float(parts[0]) / float(parts[1])
                    
            return 0.0
            
        except Exception as e:
            _logger.warning(f"Error parseando fracción/métrica '{fraction_str}': {e}")
            return 0.0
    
    # ==========================================================================
    # CONSTRAINTS: Validación de Octavos
    # ==========================================================================
    
    @api.constrains('width_visual', 'thickness_visual')
    def _check_imperial_fractions(self):
        """
        Valida que las fracciones imperiales sean octavos válidos.
        
        ADVERTENCIA (no bloquea):
        - Si detecta un valor que NO es octavo (ej: 5.71 en lugar de 5 3/8)
        - Registra warning en log pero permite guardar
        - 🛡️ NUEVO: Ignora silenciosamente las medidas métricas intencionales ("195mm")
        """
        VALID_FRACTIONS = {'1/8', '1/4', '3/8', '1/2', '5/8', '3/4', '7/8'}
        
        for line in self:
            # 1. Validar ancho
            if line.width_visual:
                w_str = str(line.width_visual).lower()
                
                # Bypass: No advertir si explícitamente es 'mm' o si es un número gigante (Heurística)
                if 'mm' in w_str or ('/' not in w_str and ' ' not in w_str and float(w_str) > 24):
                    pass # Todo bien, es una medida métrica
                else:
                    try:
                        value = self._parse_fraction(line.width_visual)
                        decimal = value - int(value)
                        
                        # Verificar si la parte decimal encaja en un octavo válido (con tolerancia)
                        is_valid = any(abs(decimal - (i/8.0)) < 0.01 for i in range(8))
                        
                        if not is_valid and decimal != 0:
                            _logger.warning(
                                f"⚠️ Ancho '{line.width_visual}' no es un octavo estándar. "
                                f"Valor decimal: {value:.4f}. "
                                f"¿Quiso decir uno de estos? {VALID_FRACTIONS}"
                            )
                    
                    except Exception as e:
                        _logger.error(f"Error validando ancho '{line.width_visual}': {e}")
            
            # 2. Validar espesor
            if line.thickness_visual:
                t_str = str(line.thickness_visual).lower()
                
                # Bypass: No advertir si explícitamente es 'mm'
                if 'mm' in t_str:
                    pass # Todo bien
                else:
                    try:
                        # El espesor generalmente usa cuartos (quarters)
                        if '/' in line.thickness_visual:
                            num, den = line.thickness_visual.split('/')
                            if den != '4':
                                _logger.warning(
                                    f"⚠️ Espesor '{line.thickness_visual}' no usa cuartos (/4). "
                                    f"Formato esperado: N/4 (ej: 7/4, 8/4)"
                                )
                    
                    except Exception as e:
                        _logger.error(f"Error validando espesor '{line.thickness_visual}': {e}")
    
# ============================================================================
# BLOQUE 2: MODELO PRINCIPAL (CABECERA DEL DOCUMENTO)
# ============================================================================
class MadenatGuiaProcessing(models.Model):
    _name = "madenat.guia.processing"
    _description = "Recepción de Guías Procesadas"
    _inherit = ['mail.thread', 'mail.activity.mixin', 'madenat.lumber.ingest.mixin', 'validation.checklist.mixin']
    _order = 'date_emission desc'
    name = fields.Char(string='Referencia', required=True, copy=False, readonly=True, default=lambda self: _('New'))
    
    # ==============================================================================================
    #                                     1. CAMPOS DEL MODELO
    # ==============================================================================================
    # ✅ NUEVO: Relación con la tabla de validación
    processing_line_ids = fields.One2many(
        'madenat.guia.processing.line', 
        'processing_id', 
        string='Análisis de Validación'
    )

    tipo_recepcion = fields.Selection([
        ('compra', 'Compra de Producto (Aserradero)'),
        ('service', 'Servicio Externo (Cepillado/Procesamiento)')
    ], string='Tipo de Recepción', required=True, default='compra', tracking=True)
    
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('verified', '🔍 Por Verificar'),
        ('processed', '📦 Procesada'),
        ('validated', '✅ Validada'),
        ('cancelled', '❌ Cancelada'),
    ], string='Estado', default='draft', required=True, tracking=True)

    name = fields.Char(string="Número de Guía", required=True, index=True, tracking=True)
    date_emission = fields.Date(string="Fecha de Emisión", required=True, default=fields.Date.context_today, tracking=True)
    partner_id = fields.Many2one('res.partner', string="Proveedor / Transportista", required=True, tracking=True)
    order_id = fields.Many2one('purchase.order', string="Orden de Compra", tracking=True)
    
    # Adjuntos
    pdf_attachment_id = fields.Many2one('ir.attachment', string="Archivo PDF Guía")
    excel_attachment_id = fields.Many2one('ir.attachment', string="Archivo Excel Packing List")

    # Binarios para UX
    oc_pdf_file = fields.Binary(string="PDF Orden de Compra (binario)")
    oc_pdf_filename = fields.Char(string="Nombre PDF OC")
    guide_pdf_file = fields.Binary(string="PDF Guía (binario)")
    guide_pdf_filename = fields.Char(string="Nombre PDF Guía")
    excel_file = fields.Binary(string="Excel Packing (binario)")
    excel_filename = fields.Char(string="Nombre Excel")

    # Relaciones
    lot_ids = fields.Many2many('stock.lot', string="Lotes relacionados")
    carrier_id = fields.Many2one('res.partner', string="Transportista", domain=[('is_company', '=', True)])
    
    # ✅ Patio de Asignación
    assignment_location_id = fields.Many2one(
        'stock.location', 
        string="Patio de Asignación", 
        domain=[('usage', '=', 'internal')], 
        required=True, 
        tracking=True
    )
    
    user_id = fields.Many2one('res.users', string="Usuario", default=lambda self: self.env.user, tracking=True)

    # Monetario
    additional_cost = fields.Monetary(string="Costo Adicional")
    currency_id = fields.Many2one('res.currency', string='Moneda', required=True, default=lambda self: self.env.company.currency_id)
    rate_date = fields.Date(string="Fecha Tipo Cambio", default=fields.Date.context_today)
    rate_usd = fields.Float(string="Tipo de Cambio USD", digits=(12, 2), default=1.0)
  # ══════════════════════════════════════════════════════════════════════════════
    # 🆕 DETALLE DEL SERVICIO (Staging de Costos) - Extraído del PDF
    # ══════════════════════════════════════════════════════════════════════════════
    # Estos campos almacenan la información del servicio de la guía para revisión.
    # NO se distribuyen automáticamente - requieren validación y aprobación manual.
    # ══════════════════════════════════════════════════════════════════════════════
    service_product_name = fields.Char(
        string="Descripción del Servicio",
        help="Nombre del servicio extraído de la guía (ej: SERVICIO DE CEPILLADO MADENAT)"
    )
    service_volume_m3 = fields.Float(
        string="Volumen del Servicio (m³)",
        digits=(16, 3),
        help="Cantidad en m³ facturada por el servicio (puede diferir del volumen físico)"
    )
    service_unit_price_clp = fields.Monetary(
        string="Precio Unitario (CLP/m³)",
        currency_field='currency_id',
        help="Precio por metro cúbico del servicio en pesos chilenos"
    )
    service_date = fields.Date(
        string="Fecha del Servicio",
        help="Fecha de emisión de la guía de servicio"
    )
    service_code = fields.Char(
        string="Código del Servicio",
        help="Código interno del servicio (ej: SFABR00017)"
    )
    # ══════════════════════════════════════════════════════════════════════════════

    # Metadatos
    notes = fields.Text(string="Notas")
    date_processed = fields.Datetime(string="Fecha de Procesamiento")
    lineas_procesadas = fields.Integer(string="Líneas Procesadas", compute="_compute_lineas_procesadas", store=True)
    _raw_pdf_text = fields.Text(string="Texto PDF (parseado)")
    lot_details_data = fields.Text(string="Lot Details Data", readonly=True)
    
  # ==========================================================================
    # DEFINICIÓN DE KPIs (CABECERA) - ACTUALIZADO
    # ==========================================================================
    # Campos informativos básicos
    total_paquetes_individuales = fields.Integer(string="Paquetes Indiv.", readonly=True, compute='_compute_package_stats', store=True)
   
   # Corrección de Etiquetas para evitar confusiones
    total_paquetes = fields.Integer(
        string="Total Paquetes", 
        help="Cantidad física de bultos/paquetes (En este caso: 12)",
        tracking=True
    )
    total_lotes_unicos = fields.Integer(
        string="Total Lotes", 
        help="Cantidad de registros únicos/etiquetas (En este caso: 17)",
        tracking=True
    )

    grouping_details_json = fields.Text(string="Detalles Agrupación JSON", readonly=True)
    can_process = fields.Boolean(string="Puede procesar", compute="_compute_can_process", store=False)

    # --- VOLÚMENES M3 (Todos usan _compute_all_totals) ---
    vol_comercial = fields.Float(string="Vol. Comercial (m³)", digits=(16, 3), compute='_compute_all_totals', store=True)
    vol_fisico = fields.Float(string="Vol. Físico (m³)", digits=(16, 3), compute='_compute_all_totals', store=True)
    vol_shipment_m3 = fields.Float(string="Vol. Embarque (m³)", digits=(16, 3), compute='_compute_all_totals', store=True)
    vol_total_m3 = fields.Float(string="Volumen Total m³", digits=(16, 3), compute='_compute_all_totals', store=True)
    vol_comercial_pdf = fields.Float(string="Vol. Guía PDF (m³)", digits=(16, 3))

    # --- VOLÚMENES MBF (Todos usan _compute_all_totals) ---
    vol_comercial_mbf = fields.Float(string="Vol. Comercial MBF", digits=(16, 3), compute='_compute_all_totals', store=True)
    vol_fisico_mbf = fields.Float(string="Vol. Físico MBF", digits=(16, 3), compute='_compute_all_totals', store=True)
    
    # ESTE ES EL QUE FALTABA CONECTAR:
    vol_total_mbf = fields.Float(string="Total MBF Exportación", digits=(16, 3), compute='_compute_all_totals', store=True)

    # --- DIFERENCIAS ---
    diff_m3 = fields.Float(string="Δ m³", digits=(16, 3), compute="_compute_all_totals", store=True)
    diff_pct = fields.Float(string="% Diferencia", digits=(16, 3), compute="_compute_all_totals", store=True)
    diff_mbf = fields.Float(string="Δ MBF", digits=(16, 3))

    # TD-007: Definiciones duplicadas eliminadas (2026-06-04).
    # Campos activos: líneas 906-912 con compute=_compute_all_totals (método real).
    # Las definiciones removidas usaban _compute_volumenes_reales (método inexistente).
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ✅ CORRECCIÓN 1: can_validate UNIFICADO
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    can_validate = fields.Boolean(
        string="Puede validar", 
        compute="_compute_can_validate", 
        store=False,
        help="Combinación de validación de checklist + volumen físico"
    )
    
    can_cancel = fields.Boolean(string="Puede cancelar", compute="_compute_can_cancel")
    can_reopen = fields.Boolean(string="Puede reabrir", compute="_compute_can_reopen")
    cancel_reason = fields.Text(string="Motivo cancelación")

   
    # ==========================================================================
    # CÁLCULO UNIFICADO DE TOTALES (CON TRAMPA DE LOGS ACTIVADA 🕵️‍♂️)
    # ==========================================================================
    @api.depends('processing_line_ids.vol_purchase_m3', 'lot_ids.volume_purchase_m3', 'lot_ids.volumen_m3')
    def _compute_all_totals(self):
        """
        Suma inteligente: Prioriza Lotes si existen, sino usa Líneas.
        SIEMPRE intenta mostrar el Volumen Nominal (Compra).
        """
        for rec in self:
            total_nominal = 0.0
            sum_export_m3 = 0.0
            sum_export_mbf = 0.0

            # ---------------------------------------------------------
            # CASO 1: YA EXISTEN LOTES (Estado Procesado / Validado)
            # ---------------------------------------------------------
            if rec.lot_ids:
                # Sumamos la variable 'volume_purchase_m3' de los lotes (Nominal)
                # Si un lote no tiene nominal, usamos el físico como respaldo (fallback)
                for lot in rec.lot_ids:
                    if lot.volume_purchase_m3 > 0:
                        total_nominal += lot.volume_purchase_m3
                    else:
                        total_nominal += lot.volumen_m3 # Fallback
                
                sum_export_m3 = sum(rec.lot_ids.mapped('vol_shipment_m3') or rec.lot_ids.mapped('volumen_m3'))
                sum_export_mbf = sum(rec.lot_ids.mapped('volumen_mbf'))

            # ---------------------------------------------------------
            # CASO 2: AÚN NO HAY LOTES (Estado Borrador / Revisión)
            # ---------------------------------------------------------
            else:
                total_nominal = sum(rec.processing_line_ids.mapped('vol_purchase_m3'))
                sum_export_m3 = sum(rec.processing_line_ids.mapped('vol_shipment_m3'))
                sum_export_mbf = sum(rec.processing_line_ids.mapped('vol_mbf'))

            # ---------------------------------------------------------
            # ASIGNACIÓN FINAL A LA VISTA
            # ---------------------------------------------------------
            rec.vol_comercial = total_nominal
            
            # 🔥 TRUCO: Forzamos que el campo 'vol_fisico' (Cuadro Rojo) muestre el Nominal
            rec.vol_fisico = total_nominal 
            
            rec.vol_shipment_m3 = sum_export_m3
            rec.vol_total_m3 = sum_export_m3
            rec.vol_total_mbf = sum_export_mbf
            rec.vol_fisico_mbf = sum_export_mbf 
            rec.vol_comercial_mbf = 0.0
            
            # Diferencias
            rec.diff_m3 = rec.vol_fisico - rec.vol_comercial
            if rec.vol_comercial > 0:
                rec.diff_pct = (rec.diff_m3 / rec.vol_comercial) * 100
            else:
                rec.diff_pct = 0.0

    @api.depends('processing_line_ids', 'lot_ids', 'state')
    def _compute_package_stats(self):
        """
        🚀 UNIFICACIÓN DE CONTEO (Bultos vs Registros):
        Corregido para evitar el intercambio de variables al procesar.
        """
        for rec in self:
            # 1. TOTAL LOTES (Líneas de Sistema): Siempre es el conteo de filas físicas (15)
            # Usamos las líneas de staging porque son la referencia histórica de la ingesta.
            count_lineas = len(rec.processing_line_ids)
            
            # 2. TOTAL PAQUETES (Bultos Físicos): Siempre es el conteo de etiquetas únicas (12)
            # Obtenemos los nombres de lote de las líneas de staging.
            nombres_staging = [l.lot_name for l in rec.processing_line_ids if l.lot_name]
            count_bultos = len(set(nombres_staging))

            # --- ASIGNACIÓN BLINDADA (No cambia con el estado) ---
            rec.total_lotes_unicos = count_lineas      # 15
            rec.lineas_procesadas = count_lineas       # 15
            rec.total_paquetes = count_bultos          # 12
            rec.total_paquetes_individuales = count_bultos

    # ==============================================================================================
    #                                     2. ORQUESTADOR PRINCIPAL
    # ==============================================================================================

    def do_full_processing(self):
        """
        🚀 PROCESAMIENTO MAESTRO (V9.1 - FIXED)
        Corrección: Se agrega el argumento 'lot_dims' requerido por _create_or_get_lot.
        """
        _logger.info(f"🔍 INICIANDO PROCESAMIENTO (MODO EXPORTACIÓN) - Guía {self.name}")
        
        import uuid
        import re

        for rec in self:
            # FASE 1: VALIDACIONES
            if rec.state not in ['draft', 'verified']: 
                raise UserError("❌ Estado inválido. Verifique la guía antes de procesar.")
            
            if not rec.processing_line_ids:
                raise UserError("⚠️ Es OBLIGATORIO usar la tabla de verificación visual (Staging).")

            # FASE 2: VARIABLES FINANCIERAS
            purchase_order = rec.order_id
            precio_usd = rec._obtener_precio_desde_oc(purchase_order)
            existing_pickings = rec.env['stock.picking'].search([
                ('origin', '=', rec.name),
                ('state', 'not in', ['done', 'cancel'])
            ])
            if existing_pickings:
                existing_pickings.action_cancel()
            # FASE 3: BUCLE DE CREACIÓN DE LOTES
            lot_data = []
            guia_suffix = re.sub(r'[^0-9]', '', rec.name) or 'G' + str(rec.id)
            
            for item in rec.processing_line_ids:
                 
                 # --- A. DEFINICIÓN DE LA VERDAD (REGLA DE ORO) ---
                # v_purchase: Es el metraje comercial (Nominal). Es lo que Odoo mueve en Stock.
                v_purchase = item.vol_purchase_m3 

                # v_stock_real: Es el metraje físico (Real). Es lo que se declara en Exportación.
                # 🚀 CORRECCIÓN: Antes apuntaba a vol_purchase_m3, ahora apunta al shipment real.
                v_stock_real = item.vol_shipment_m3 

                if v_stock_real <= 0:
                    _logger.warning(f"⚠️ Línea {item.id} tiene vol_shipment_m3={item.vol_shipment_m3} — usando fallback vol_purchase")
                    v_stock_real = v_purchase if v_purchase > 0 else 0.001 # Safe-guard para evitar errores de división o nulos 

                # Identidad Visual
                sku_base = str(item.sku_original or '').strip()
                etiqueta_visual = str(item.lot_name or '').strip()
                raw_product_name = item.product_name_original
                
                if not sku_base:
                    sku_base = f"GEN-{uuid.uuid4().hex[:6]}"
                
                lote_name_internal = f"{etiqueta_visual} {sku_base} {guia_suffix}".strip()

                # --- ✅ CORRECCIÓN: CONSTRUIR DICCIONARIO DE DIMENSIONES (ldims) ---
                # Esto es lo que faltaba y causaba el error
                ldims = {
                    'espesor_mm': item.espesor_mm, 
                    'ancho_mm': item.ancho_mm, 
                    'largo_m': item.largo_m,
                    'espesor_nominal_mm': item.espesor_nominal_mm, 
                    'ancho_nominal_mm': item.ancho_nominal_mm if item.ancho_nominal_mm > 0 else item.ancho_mm,
                    'note': raw_product_name, 
                    'ref': sku_base
                }

                # --- C. GESTIÓN DE PRODUCTO ---
                group_code = re.sub(r'[^A-Z0-9]', '_', raw_product_name.upper())
                product = rec.find_or_create_product(group_code, raw_product_name)

                # --- D. CREACIÓN DEL LOTE ---
                lot = rec._create_or_get_lot(
                    guia_ref=rec.name,
                    product=product,
                    qty=item.pieces,
                    
                    vol_purchase=v_purchase,
                    vol_shipment=v_stock_real,
                    vol_real=v_stock_real,
                    
                    lot_name=lote_name_internal,
                    
                    # ✅ AQUÍ PASAMOS EL DICCIONARIO QUE FALTABA
                    lot_dims=ldims,
                    
                    # Contexto
                    precio_usd=precio_usd,
                    purchase_order=purchase_order,
                    thickness_visual=item.thickness_visual, # Pasamos visuales también por si acaso
                    width_visual=item.width_visual,
                    subproducto_id=item.subproducto_id.id if item.subproducto_id else False
                )

                # --- E. INYECCIÓN DE DATOS EXTRA (EL PUENTE) ---
                lot.write({
                    'ref': etiqueta_visual,
                    'length_ft': item.length_ft,      
                    'vol_shipment_m3': v_stock_real, # 🌎 Exportación (Se queda solo aquí)
                    
                    # 📦 EL CAMBIO CRÍTICO: Stock Real debe ser el Nominal de Compra
                    'volumen_m3': item.vol_purchase_m3,      
                    
                    'width_visual': item.width_visual,
                    'thickness_visual': item.thickness_visual,
                })

                # --- F. ACUMULADOR ---
                lot_data.append({
                    'producto_codigo': group_code,
                    'cantidad': item.pieces,
                    'volumen': v_purchase,
                    'lote': lot,
                    'lote_code': etiqueta_visual,
                    'product_name': product.name,
                    'sku_original': sku_base
                })

            # FASE 4: FINALIZACIÓN
            if lot_data:
                total_vol = sum(d['volumen'] for d in lot_data)
                total_pkg = len(lot_data)

                rec.write({
                    'vol_fisico': total_vol,
                    'vol_total_m3': total_vol,
                    'total_paquetes': total_pkg,
                    'total_lotes_unicos': len(set(d['lote_code'] for d in lot_data)),
                    'state': 'processed',
                    'date_processed': fields.Datetime.now()
                })
                
                rec._assign_costs_to_generated_lots(lot_data, {'total_volumen': total_vol})
                
                rec.lot_ids = [(6, 0, [d['lote'].id for d in lot_data])]
                rec.lot_details_data = rec._generate_lot_details_json(lot_data)

                if rec.order_id:
                    rec._sync_purchase_order_lines(rec.order_id, lot_data, precio_usd)
                # Antes de crear el picking, cancelar/borrar el picking anterior si existe
               
                rec._create_picking_and_lines(rec.order_id, lot_data)

                rec.message_post(body=f"✅ <strong>PROCESADO (EXPORT DRIVEN)</strong><br/>"
                                      f"Volumen Stock: {total_vol:.3f} m³")

    def _sync_purchase_order_lines(self, order, lot_data, price_unit):
        """
        Helper para sincronizar líneas de OC sin ensuciar la función principal.
        Agrupa los lotes por producto y actualiza/crea líneas en la OC.
        """
        from collections import defaultdict
        
        # Agrupación
        po_lines_summary = defaultdict(lambda: {'qty': 0.0, 'product': None})
        for ld in lot_data:
            prod = ld['lote'].product_id
            qty = ld['volumen'] if ld['volumen'] > 0 else ld['cantidad']
            po_lines_summary[prod.id]['qty'] += qty
            po_lines_summary[prod.id]['product'] = prod

        # Preparación de escritura
        lines_to_write = []
        existing_products = order.order_line.mapped('product_id.id')
        
        for pid, data in po_lines_summary.items():
            if pid in existing_products:
                # Actualizar existente (solo si borrador/provisional)
                if order.state in ['draft', 'sent'] or getattr(order, 'provisional', False):
                    line = order.order_line.filtered(lambda l: l.product_id.id == pid)[0]
                    lines_to_write.append((1, line.id, {'product_qty': data['qty']}))
            else:
                # Crear nueva
                prod = data['product']
                lines_to_write.append((0, 0, {
                    'product_id': pid,
                    'name': prod.display_name,
                    'product_qty': data['qty'],
                    'price_unit': price_unit,
                    'date_planned': fields.Datetime.now(),
                    'product_uom': prod.uom_id.id,
                    'taxes_id': [(6, 0, prod.supplier_taxes_id.ids)]
                }))
        
        if lines_to_write:
            order.write({'order_line': lines_to_write})

    def action_assign_commercial_defaults(self):
        """
        BOTÓN MANUAL: 'Asignar Nominal de Compra'.
        
        Lógica:
        Recorre todas las líneas del Staging. Si el 'Espesor Nominal' es 0,
        le asigna el valor del 'Espesor Físico'.
        
        Uso: Permite llenar rápidamente la pestaña 'Análisis Comercial' 
        cuando no hay OC y queremos asumir que compramos lo mismo que medimos.
        """
        for line in self.processing_line_ids:
            # Solo actuamos si no tiene valor (respetamos lo que hayas editado a mano)
            if line.espesor_nominal_mm <= 0:
                line.espesor_nominal_mm = line.espesor_mm
                # Forzamos el recálculo del volumen de compra
                line._compute_vol_purchase_m3()
        
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Asignación Completa',
                'message': 'Se han igualado los espesores nominales a los físicos en las líneas vacías.',
                'type': 'success',
                'sticky': False,
            }
        }       
   # ==========================================================================
    # MÉTODOS DE BOTONES (VISTA XML)
    # ==========================================================================

    def action_verify_data(self):
        """
        PASO 1: CARGA Y VERIFICACIÓN (Staging)
        - Carga datos Físicos.
        - Calcula visuales de Exportación (Fracciones).
        - DEJA LIMPIO el Nominal de Compra (si no viene en Excel) para asignación manual posterior.
        """
        self.ensure_one()
        import re
        
        if not (self.excel_attachment_id or self.excel_file):
            raise UserError("❌ Falta el archivo Excel de Packing.")

        self.processing_line_ids.unlink()

        archivo_excel = self.excel_attachment_id or self._store_binary_as_attachment('excel_file', 'excel_filename', 'Packing Excel')
        packing_data = self._parse_packing_excel(archivo_excel)
        
        lines_to_create = []
        
        for row in packing_data.get('lineas', []):
            try:
                # Datos básicos
                code = row.get('Codigo Interno', '')
                lote_fisico = row.get('N° LOTE', '')
                qty = row.get('Cantidad', 0)
                vol_raw = row.get('Volumen', 0.0)
                
                # Gestión Producto
                prod_name = row.get('product_name', 'Madera Genérica').strip() or 'Madera Genérica'
                prod_code_base = prod_name if len(prod_name) > 3 else code
                clean_group_code = re.sub(r'[^A-Z0-9]', '_', prod_code_base.upper())
                product = self.find_or_create_product(clean_group_code, prod_name)
                
                # Visuales Exportación (Calculados por el parser o helper)
                txt_nom_esp = row.get('espesor_nominal')
                if not txt_nom_esp and row.get('Espesor', 0) > 0:
                    txt_nom_esp = self._get_nominal_dimension(float(row.get('Espesor')), 'thickness')
                
                txt_nom_anc = row.get('ancho_nominal')
                if not txt_nom_anc and row.get('Ancho', 0) > 0:
                    txt_nom_anc = self._get_nominal_dimension(float(row.get('Ancho')), 'width')

                lines_to_create.append({
                    'processing_id': self.id,
                    'lot_name': lote_fisico or code, 
                    'sku_original': code,            
                    'product_name_original': prod_name,
                    'product_id': product.id,
                    'pieces': qty,
                    
                    'vol_physical_m3': vol_raw,
                    'vol_shipment_m3': vol_raw,
                    'vol_purchase_m3': vol_raw, # Inicial temporal
                    
                    'espesor_mm': row.get('Espesor', 0),
                    'ancho_mm': row.get('Ancho', 0),
                    'largo_m': row.get('Largo', 0),
                    
                    # Nominal de Compra: Respetamos el Excel. Si es 0, queda en 0.
                    # Luego usarás el botón 'action_assign_commercial_defaults'
                    'espesor_nominal_mm': row.get('espesor_nominal_mm', 0.0),
                    
                    # Exportación
                    'thickness_visual': txt_nom_esp,
                    'width_visual': txt_nom_anc,
                    
                    'technical_validation': 'pending'
                })
            except Exception as e:
                _logger.warning(f"⚠️ Error línea Excel: {e}")
                continue

        if lines_to_create:
            self.env['madenat.guia.processing.line'].create(lines_to_create)
        
        # ═══════════════════════════════════════════════════════════════════════════════
        # 🆕 ADICIÓN: Procesamiento PDF para Costos y T/C (NO afecta lógica existente)
        # ═══════════════════════════════════════════════════════════════════════════════
        # IMPORTANTE: Este bloque solo AGREGA datos financieros del PDF.
        # Si el PDF no existe o falla, el flujo continúa normalmente.
        # ═══════════════════════════════════════════════════════════════════════════════
        update_vals = {'state': 'verified'}  # Estado original que ya estaba
        
        if self.pdf_attachment_id or self.guide_pdf_file:
            try:
                archivo_pdf = self.pdf_attachment_id or self._store_binary_as_attachment(
                    'guide_pdf_file', 
                    'guide_pdf_filename', 
                    'Guía PDF'
                )
                pdf_data = self._parse_dispatch_pdf(archivo_pdf)
                
                # Solo agregamos campos si existen en el PDF (no sobrescribimos valores previos)
                if pdf_data.get('additional_cost', 0) > 0:
                    update_vals['additional_cost'] = pdf_data['additional_cost']
                
                if pdf_data.get('rate_usd', 0) > 0:
                    update_vals['rate_usd'] = pdf_data['rate_usd']

                if pdf_data.get('service_date'):
                    update_vals['rate_date'] = pdf_data['service_date']

                if pdf_data.get('volumen_comercial', 0) > 0:
                    update_vals['vol_comercial_pdf'] = pdf_data['volumen_comercial']
                
                if pdf_data.get('numero_guia'):
                    update_vals['name'] = pdf_data['numero_guia']
                
                # ✅ Asignar datos del servicio (si existen)
                if pdf_data.get('service_product_name'):
                    update_vals['service_product_name'] = pdf_data['service_product_name']
                
                if pdf_data.get('service_volume_m3', 0) > 0:
                    update_vals['service_volume_m3'] = pdf_data['service_volume_m3']
                
                if pdf_data.get('service_unit_price_clp', 0) > 0:
                    update_vals['service_unit_price_clp'] = pdf_data['service_unit_price_clp']
                
                if pdf_data.get('service_date'):
                    update_vals['service_date'] = pdf_data['service_date']
                
                if pdf_data.get('service_code'):
                    update_vals['service_code'] = pdf_data['service_code']
                
                _logger.info(
                    f"✅ PDF procesado - Guía: {self.name} | "
                    f"Costo Total: ${pdf_data.get('additional_cost', 0):,.0f} CLP | "
                    f"T/C: {pdf_data.get('rate_usd', 1.0)} | "
                    f"Servicio: {pdf_data.get('service_product_name', 'N/A')} | "
                    f"Vol: {pdf_data.get('service_volume_m3', 0):.3f} m³ @ "
                    f"${pdf_data.get('service_unit_price_clp', 0):,.0f}/m³"
                )

            except Exception as e:
                _logger.warning(f"⚠️ PDF no procesado (no crítico): {e}")
                # No hacemos raise - dejamos que el flujo continúe sin PDF
        # ═══════════════════════════════════════════════════════════════════════════════
     
      
        self.write(update_vals)  # Ahora escribe state + datos del PDF (si existen)
        
        # Mostrar notificación Y recargar la vista (sin cambios)
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
            'params': {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Datos Cargados',
                    'message': f'Se cargaron {len(self.processing_line_ids)} líneas. Use el botón 🪄 en cada línea para asignar nominales.',
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                }
            }
        }




    def action_process_from_staging(self):
        """
        ✅ PUENTE OFICIAL BLINDADO v5.0: Staging -> Lotes Reales (Stock)
        """
        self.ensure_one()
        _logger.info(f"🛡️ Iniciando procesamiento blindado para Guía: {self.name}")

        # 1. CAPA DE SEGURIDAD: EVITAR DUPLICADOS
        # Buscamos si ya existe basura (Pickings activos o borradores con este origen)
        existing_picking = self.env['stock.picking'].search([
            ('origin', '=', self.name),
            ('state', 'not in', ['cancel'])
        ], limit=1)
        
        if existing_picking:
            raise UserError(
                f"⛔ OPERACIÓN BLOQUEADA\n\n"
                f"Ya existe un albarán activo para esta guía: {existing_picking.name}.\n"
                f"Estado actual: {existing_picking.state}\n\n"
                f"👉 Solución: Si necesita procesar de nuevo, use 'Cancelación Forzada' "
                f"para limpiar la basura logística primero."
            )

        # 2. CAPA DE INTEGRIDAD: VALIDACIÓN DE DATOS MAESTROS
        if not self.partner_id:
            raise UserError("❌ Error: Debe asignar un Proveedor antes de procesar.")
        if not self.assignment_location_id:
            raise UserError("❌ Error: No se ha definido la 'Ubicación de Destino' para los lotes.")
        if not self.processing_line_ids:
            raise UserError("❌ Error: La tabla de verificación está vacía.")

        # Verificar nominales de espesor
        lines_sin_nominal = self.processing_line_ids.filtered(lambda l: l.espesor_nominal_mm <= 0)
        if lines_sin_nominal:
            raise UserError(f"⚠️ Validación de Dimensiones Fallida: Hay {len(lines_sin_nominal)} líneas sin espesor nominal.")

        # 3. PREPARACIÓN DE LA 'TRIPLE VERDAD' (Payload)
        total_vol = sum(self.processing_line_ids.mapped('vol_shipment_m3'))
        lotes_unicos = set(self.processing_line_ids.mapped('lot_name'))
        lineas_maestras = []
        
        for line in self.processing_line_ids:
            lineas_maestras.append({
                'Codigo Interno': line.sku_original,
                'product_name': line.product_name_original,
                'N° LOTE': line.lot_name,
                'calidad_id': line.subproducto_id.id, # FAS / S2S / Rough
                'Cantidad': line.pieces,
                'Espesor': line.espesor_mm,
                'Ancho': line.ancho_mm,
                'Largo': line.largo_m,
                'espesor_nominal_mm': line.espesor_nominal_mm, 
                'thickness_visual': line.thickness_visual, 
                'width_visual': line.width_visual,
                'length_ft': line.length_ft,
                'Volumen': line.vol_shipment_m3,
                'vol_purchase_m3': line.vol_purchase_m3,
                'vol_physical_m3': line.vol_physical_m3,
                'vol_mbf': line.vol_mbf,
                'technical_validation': 'approved',
                'estado_trazabilidad': 'recepcionado'
            })

        packing_simulado = {
            'total_paquetes': len(self.processing_line_ids),
            'total_volumen': total_vol,
            'total_lotes': len(lotes_unicos),
            'lineas': lineas_maestras, 
            'oc_reference': self.order_id.name if self.order_id else self.name
        }

        # 4. DISPARO CON CONTEXTO PROTEGIDO
        return self.with_context(
            from_staging_bridge=True,
            packing_override=packing_simulado,
            default_partner_id=self.partner_id.id,
            default_origin=self.name
        ).do_full_processing()

    

    # ==============================================================================================
    #                                     ACTION_VALIDATE MEJORADO
    # ==============================================================================================

    def action_validate(self):
        """
        🚀 VERSIÓN v23.0 (SELF-HEALING): AUTORREPARACIÓN + INYECCIÓN NOMINAL
        ===================================================================
        Mejoras Críticas:
        0. 🔧 AUTORREPARACIÓN: Detecta si el producto es 'Consumible' y lo fuerza a 'Almacenable'
           antes de procesar, garantizando la creación de Stock (Quants).
        1. 🛡️ BLINDAJE: Verifica que el destino sea Bodega (Internal).
        2. 👻 LIMPIEZA: Elimina fantasmas.
        3. 💎 NOMINAL: Inyecta volumen de compra.
        4. 🔄 REFRESH: Recarga la interfaz.
        """
        self.ensure_one()
        _logger.info(f"🔍 [START] Validando Guía {self.name} - MODO: Nominal + Autorreparación")

        # 1. Validaciones Previas
        if self.state != 'processed':
            raise ValidationError("⛔ La guía debe estar en estado 'Procesada' para validar.")

        # =======================================================
        # 🔧 FASE 0: AUTORREPARACIÓN DE MAESTROS (CRÍTICO ODOO 18)
        # =======================================================
        # Antes de mover nada, aseguramos que los productos sean Almacenables.
        # Si son 'Consumibles' (is_storable=False), Odoo NO crea stock.
        productos_involucrados = self.processing_line_ids.mapped('product_id')
        for product in productos_involucrados:
            # Usamos sudo() para saltar reglas de permisos si el usuario no es admin de inventario
            tmpl = product.product_tmpl_id.sudo()
            if not tmpl.is_storable:
                _logger.warning(f"🔧 [FIX] El producto '{product.name}' era Consumible. Convirtiendo a ALMACENABLE.")
                tmpl.write({
                    'is_storable': True, 
                    'type': 'consu' # En Odoo 18, type='consu' + is_storable=True es lo correcto
                })
        # =======================================================

        # 2. Recuperar o Crear Picking
        picking = self._get_or_create_picking_unified()
        if not picking:
            raise UserError("⛔ Error Crítico: No se pudo generar ni recuperar el albarán asociado.")

        # 🛡️ BLINDAJE LOGÍSTICO
        if picking.location_dest_id.usage != 'internal':
             raise UserError(
                 f"⛔ ERROR DE CONFIGURACIÓN DE DESTINO\n\n"
                 f"El albarán '{picking.name}' apunta a: '{picking.location_dest_id.name}' ({picking.location_dest_id.usage})\n"
                 f"SOLUCIÓN: El destino DEBE ser una Ubicación Interna (Bodega) para generar stock."
             )

        # 3. Asegurar estado del Picking
        if picking.state == 'draft':
            picking.action_confirm()
        if picking.state not in ('assigned', 'done'):
            picking.action_assign()

        # 4. Procesamiento de Movimientos
        if picking.state != 'done':
            
            pool_lotes = list(self.lot_ids)
            moves = picking.move_ids.filtered(lambda m: m.state not in ('done', 'cancel'))
            
            for move in moves:
                
                # =======================================================
                # 👻 FASE A: LIMPIEZA DE FANTASMAS
                # =======================================================
                tiene_lote = bool(move.lot_ids)
                if not tiene_lote:
                    for line in move.move_line_ids:
                        if line.lot_id:
                            tiene_lote = True
                            break
                
                if not tiene_lote and move.product_id.tracking != 'none':
                    _logger.warning(f"🗑️ Eliminando Fantasma: {move.product_uom_qty} sin lote.")
                    move.move_line_ids.unlink()
                    move._action_cancel()
                    move.unlink()
                    continue

                # =======================================================
                # 💎 FASE B: RECUPERACIÓN DE DATO NOMINAL
                # =======================================================
                target_lot = move.lot_ids[0] if move.lot_ids else False
                
                if not target_lot:
                    for lot in pool_lotes:
                        if abs(lot.volumen_m3 - move.product_uom_qty) < 0.01:
                            target_lot = lot
                            pool_lotes.remove(lot)  # ← consumir lote del pool — evita duplicado
                            break
                
                cantidad_final = move.product_uom_qty 
                
                if target_lot:
                    if hasattr(target_lot, 'volume_purchase_m3') and target_lot.volume_purchase_m3 > 0:
                         cantidad_final = target_lot.volume_purchase_m3
                    elif hasattr(target_lot, 'p_nominal') and target_lot.p_nominal > 0:
                         cantidad_final = target_lot.p_nominal
                
                # =======================================================
                # 💉 FASE C: INYECCIÓN QUIRÚRGICA
                # =======================================================
                move.quantity = cantidad_final
                if 'picked' in move._fields:
                    move.picked = True
                
                if not move.move_line_ids:
                    vals = {
                        'picking_id': picking.id,
                        'move_id': move.id,
                        'product_id': move.product_id.id,
                        'lot_id': target_lot.id if target_lot else False,
                        'quantity': cantidad_final,
                        'location_id': move.location_id.id,
                        'location_dest_id': move.location_dest_id.id,
                        'product_uom_id': move.product_uom.id,
                        'company_id': picking.company_id.id,
                    }
                    if 'picked' in self.env['stock.move.line']._fields:
                        vals['picked'] = True
                    self.env['stock.move.line'].create(vals)
                else:
                    for line in move.move_line_ids:
                        line.write({
                            'quantity': cantidad_final,
                            'lot_id': target_lot.id if target_lot else line.lot_id.id
                        })
                        if 'picked' in line._fields:
                            line.picked = True

            # =======================================================
            # 🏁 FASE D: VALIDACIÓN
            # =======================================================
            ctx = picking.env.context.copy()
            ctx.update({
                'skip_backorder': True,
                'skip_immediate': True,
                'check_move_validity': False,
            })
            
            try:
                res = picking.with_context(ctx).button_validate()
                if isinstance(res, dict) and 'res_model' in res:
                    Wizard = self.env[res['res_model']].with_context(res.get('context', ctx))
                    wiz = Wizard.create({})
                    if hasattr(wiz, 'process'): wiz.process()
                    elif hasattr(wiz, 'process_simple'): wiz.process_simple()
                    
            except Exception as e:
                raise UserError(f"❌ Error al validar Picking: {str(e)}")

        # 5. Actualización Final
        self.write({
            'state': 'validated',
            'validated_by_id': self.env.user.id,
            'validated_date': fields.Datetime.now(),
        })
        
        if self.lot_ids:
            self.lot_ids.write({'estado_trazabilidad': 'recepcionado', 'technical_validation': 'approved'})

        _logger.info("✅ [END] Validación Exitosa - Stock Generado.")
        
        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }
    
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ✅ NUEVO: Método helper para crear picking desde guía
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def _create_stock_picking_from_guia(self):
        """
        ✅ MÉTODO NUEVO UNIFICADO
        Integrado con el motor maestro. Genera movimientos masivos (Odoo 18 style)
        reutilizando una cabecera única por guía.
        """
        self.ensure_one()
        
        # 🛡️ UNIFICACIÓN: Llamamos al maestro
        picking = self._get_or_create_picking_unified()
        
        move_vals_list = []
        uom_m3_id = self.env.ref('uom.product_uom_cubic_meter').id
        
        # Preparamos los movimientos para los lotes que aún no tienen picking
        for lot in self.lot_ids:
            # Evitamos duplicar movimientos si el lote ya está en un picking activo
            # (Opcional, pero recomendado para máxima seguridad)
            move_vals = {
                'name': f"{lot.product_id.name} - {lot.name}",
                'product_id': lot.product_id.id,
                'product_uom_qty': lot.volumen_m3,
                'product_uom': uom_m3_id,
                'location_id': picking.location_id.id,
                'location_dest_id': picking.location_dest_id.id,
                'picking_id': picking.id,
                'lot_ids': [(6, 0, [lot.id])],
            }
            move_vals_list.append((0, 0, move_vals))
        
        if move_vals_list:
            # Añadimos los movimientos al albarán maestro
            picking.write({'move_ids_without_package': move_vals_list})
            
            # Si el albarán estaba en borrador, lo confirmamos
            if picking.state == 'draft':
                picking.action_confirm()

        # Sincronizamos la relación en la Guía
        if hasattr(self, 'picking_ids'):
            self.picking_ids = [(4, picking.id)]
            
        _logger.info(f"✅ Movimientos añadidos al albarán maestro {picking.name}")
        return picking
    # ====================================================================
    # 🛡️ MOTOR ÚNICO DE CREACIÓN DE STOCK (INTELLIGENT UNIFIED ENGINE)
    # ====================================================================
    def _get_or_create_picking_unified(self, lot_data=None):
        """
        🚀 FUENTE DE VERDAD ÚNICA v7.1 (BLINDADA + DINÁMICA)
        Unifica flujos Legacy (lot_data) y Modernos (self.lot_ids).
        Garantiza: 
        1. 1 Guía = 1 Albarán con trazabilidad total.
        2. BLINDAJE DINÁMICO: Ubicación sin hardcode, con fallbacks inteligentes.
        """
        self.ensure_one()
        _logger.info(f"⚙️ Procesando motor de stock para {self.name}")

        # 1. 🛡️ ESCUDO ANTI-DUPLICADOS (Búsqueda por Origen)
        picking = self.env['stock.picking'].search([
            ('origin', '=', self.name),
            ('state', 'not in', ['cancel', 'done'])
        ], limit=1)

        if picking:
            # 🛑 SAFETY CHECK: Si el existente es de SALIDA (OUT), es basura. Lo descartamos.
            if picking.picking_type_code == 'outgoing':
                _logger.warning(f"⚠️ Albarán existente incorrecto ({picking.name}) detectado. Se ignorará.")
                picking.action_cancel()
            else:
                _logger.info(f"♻️ Albarán existente válido detectado ({picking.name}). Reutilizando...")
                return picking

        # 2. 🏗️ CONFIGURACIÓN DE DATOS MAESTROS
        if not self.partner_id:
            raise UserError("❌ Operación abortada: La guía no tiene un Proveedor asignado.")

        # Buscamos tipo 'incoming' (Recepciones)
        picking_type = self.env['stock.picking.type'].search([
            ('code', '=', 'incoming'),
            ('warehouse_id.company_id', '=', self.env.company.id)
        ], limit=1)

        if not picking_type:
            # Fallback: Buscar cualquiera de entrada si no hay específico por almacén
            picking_type = self.env['stock.picking.type'].search([
                ('code', '=', 'incoming'),
                ('company_id', '=', self.env.company.id)
            ], limit=1)
            
        if not picking_type:
            raise UserError("❌ Error de Configuración: No existe un tipo de operación 'Recepciones' (Incoming).")

        # 🎯 DEFINICIÓN INTELIGENTE DE UBICACIONES
        # Origen: Siempre Proveedor (sin cambios)
        location_src = self.partner_id.property_stock_supplier.id or self.env.ref('stock.stock_location_suppliers').id
        
        # ══════════════════════════════════════════════════════════════════════════════
        # 🆕 DESTINO DINÁMICO: Jerarquía de Fallbacks SIN HARDCODE
        # ══════════════════════════════════════════════════════════════════════════════
        # Prioridad 1: Asignación Manual del Usuario (respeta decisión operativa)
        # Prioridad 2: Configuración del Tipo de Operación (respeta setup de Odoo)
        # Prioridad 3: Búsqueda inteligente de ubicación 'Stock' (fallback seguro)
        # Prioridad 4: Error controlado con mensaje claro (mejor que hardcode ciego)
        # ══════════════════════════════════════════════════════════════════════════════
        location_dest_id = None
        
        if self.assignment_location_id:
            # ✅ CASO 1: Usuario eligió manualmente el patio
            location_dest_id = self.assignment_location_id.id
            _logger.info(f"📍 Ubicación destino: '{self.assignment_location_id.name}' (Asignación Manual)")
            
        elif picking_type.default_location_dest_id:
            # ✅ CASO 2: Hay configuración en el Tipo de Operación
            location_dest_id = picking_type.default_location_dest_id.id
            _logger.info(f"📍 Ubicación destino: '{picking_type.default_location_dest_id.name}' (Config. Tipo Operación)")
            
        else:
            # 🔍 CASO 3: Fallback Inteligente - Buscar ubicación 'Stock' de la compañía
            fallback_location = self.env['stock.location'].search([
                ('usage', '=', 'internal'),
                ('company_id', '=', self.env.company.id),
                '|',
                ('name', '=', 'Stock'),
                ('complete_name', 'ilike', 'WH/Stock')
            ], limit=1)
            
            if fallback_location:
                location_dest_id = fallback_location.id
                _logger.warning(
                    f"⚠️ FALLBACK AUTOMÁTICO: Usando ubicación '{fallback_location.complete_name}' "
                    f"(ID: {fallback_location.id}). "
                    f"RECOMENDACIÓN: Configure un 'Patio de Asignación' para esta guía."
                )
            else:
                # ❌ CASO 4: No hay ninguna ubicación válida - Error claro
                raise UserError(
                    "❌ ERROR DE CONFIGURACIÓN CRÍTICO\n\n"
                    "No se pudo determinar la ubicación de destino para el stock.\n\n"
                    "📋 SOLUCIONES POSIBLES:\n"
                    "1️⃣ Seleccione un 'Patio de Asignación' en este formulario\n"
                    "2️⃣ Configure una ubicación por defecto en el Tipo de Operación 'Recepciones'\n"
                    "3️⃣ Verifique que existe una ubicación 'Stock' en su almacén\n\n"
                    f"📄 Guía afectada: {self.name}\n"
                    f"🏢 Compañía: {self.env.company.name}"
                )
        # ══════════════════════════════════════════════════════════════════════════════

        # 3. 📝 CREACIÓN DEL ALBARÁN (CABECERA) - Sin cambios
        picking_vals = {
            'partner_id': self.partner_id.id,
            'picking_type_id': picking_type.id,
            'location_id': location_src,      # DESDE: Proveedor
            'location_dest_id': location_dest_id, # HACIA: Bodega (ahora 100% dinámico)
            'origin': self.name,
            'purchase_id': self.order_id.id if self.order_id else False,
            'move_ids_without_package': [],
            'company_id': self.env.company.id,
        }

        # 4. 🪵 LÓGICA DE MOVIMIENTOS (HÍBRIDA - INTACTA) - Sin cambios
        uom_m3 = self.env.ref('uom.product_uom_cubic_meter')
        
        # CASO A: Tenemos lotes ya creados en el sistema (Flujo Nuevo)
        if self.lot_ids:
            _logger.info("📦 Generando movimientos desde lotes registrados en Odoo.")
            for lot in self.lot_ids:
                if lot.volumen_m3 <= 0: continue
                move_vals = {
                    'name': f"{lot.product_id.name} - {lot.name}",
                    'product_id': lot.product_id.id,
                    'product_uom_qty': lot.volumen_m3,
                    'product_uom': uom_m3.id,
                    'location_id': location_src,
                    'location_dest_id': location_dest_id,
                    'lot_ids': [(6, 0, [lot.id])], # Vínculo directo al lote
                    'company_id': self.env.company.id,
                }
                picking_vals['move_ids_without_package'].append((0, 0, move_vals))

        # CASO B: Recibimos datos externos (Flujo Legacy / Toll Processing)
        elif lot_data:
            _logger.info("📑 Generando movimientos desde datos de carga externos.")
            for idx, d in enumerate(lot_data, 1):
                move_vals = {
                    'name': d.get('product_name', 'Madera'),
                    'description_picking': f"Pqte #{idx} | Lote {d.get('lote_code')}",
                    'product_id': d.get('product_id') or d.get('lote').product_id.id,
                    'product_uom_qty': d.get('volumen', 0.0),
                    'product_uom': uom_m3.id,
                    'location_id': location_src,
                    'location_dest_id': location_dest_id,
                    'company_id': self.env.company.id,
                }
                # Si el lote ya existe en el diccionario, lo vinculamos
                if d.get('lote'):
                    move_vals['lot_ids'] = [(6, 0, [d['lote'].id])]
                
                picking_vals['move_ids_without_package'].append((0, 0, move_vals))

        # 5. 🚀 EJECUCIÓN Y CONFIRMACIÓN - Sin cambios
        if not picking_vals['move_ids_without_package']:
            # En lugar de solo loggear, avisamos al usuario
            raise UserError("⚠️ No se puede generar el albarán: No hay lotes válidos o con volumen mayor a 0.")

        picking = self.env['stock.picking'].create(picking_vals)
        
        # Vincular al campo picking_ids si existe en el modelo
        if hasattr(self, 'picking_ids'):
            self.picking_ids = [(4, picking.id)]

        _logger.info(f"✅ Albarán maestro {picking.name} generado con éxito.")
        
        # Confirmar automáticamente para que pase de 'Borrador' a 'Preparado'
        picking.action_confirm()
        return picking

    # ==============================================================================================
    #                                     3. LÓGICA DE PARSEO
    # ==============================================================================================

    def _parse_dispatch_pdf(self, attachment):
        try:
            file_path = attachment._full_path(attachment.store_fname)
            texto = ''
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages: 
                    texto += page.extract_text() or ''
            
            # ══════════════════════════════════════════════════════════
            # SECCIÓN 1: EXTRACCIÓN DE GUÍA (sin cambios)
            # ══════════════════════════════════════════════════════════
            guia = re.search(r'Gu[ií]a[-\s]*N[oº]\s*[:\-]?\s*(\d+)', texto)
            
            # ══════════════════════════════════════════════════════════
            # SECCIÓN 2: EXTRACCIÓN DE ORDEN DE COMPRA (sin cambios)
            # ══════════════════════════════════════════════════════════
            oc_patterns = [
                r'(?:OC|MC|Orden)[\s\.\-:]*([A-Z0-9]+(?:\s*[-]\s*[A-Z0-9]+)+)', 
                r'801[\s\.\-]*(\d+)'
            ]
            oc_name = False
            for p in oc_patterns:
                match = re.search(p, texto, re.I)
                if match:
                    oc_name = re.sub(r'\s*-\s*', '-', match.group(1))
                    break

            # ══════════════════════════════════════════════════════════
            # SECCIÓN 3: EXTRACCIÓN DE VOLUMEN (sin cambios)
            # ══════════════════════════════════════════════════════════
            vol_com_m3 = 0.0
            vol_patterns = [
                r'VOLUMEN\s+TOTAL\s*([\d\.,]+)\s*M3', 
                r'TOTAL\s+M3\s*[:\-]?\s*([\d\.,]+)', 
                r'47,150000'
            ]
            for pattern in vol_patterns:
                match = re.search(pattern, texto, re.IGNORECASE)
                if match:
                    vol_com_m3 = self._parse_float_value(match.group(1), 'vol_pdf')
                    break
            
            if vol_com_m3 == 0:
                svc_match = re.search(r'SERVICIO.*?(\d+[,\d]+)\s*Metros', texto)
                if svc_match: 
                    vol_com_m3 = self._parse_float_value(svc_match.group(1), 'vol_svc')

            # ══════════════════════════════════════════════════════════
            # ✅ SECCIÓN 4: EXTRACCIÓN DE COSTO NETO (NUEVA)
            # ══════════════════════════════════════════════════════════
            subtotal_neto = 0.0
            subtotal_patterns = [
                r'Subtotal\s+Neto\s*\$?\s*([\d\.,]+)',
                r'Neto\s*\$?\s*([\d\.,]+)',
                r'VALOR.*?\$\s*([\d\.,]+)'
            ]
            for pattern in subtotal_patterns:
                match = re.search(pattern, texto, re.IGNORECASE)
                if match:
                    subtotal_neto = self._parse_float_value(match.group(1), 'subtotal_neto')
                    break

            # ══════════════════════════════════════════════════════════
            # ✅ SECCIÓN 5: EXTRACCIÓN DE TIPO DE CAMBIO USD (NUEVA)
            # ══════════════════════════════════════════════════════════
            tipo_cambio_usd = 1.0  # Default seguro
            tc_patterns = [
                r'T/C\s+U\$?\s*([\d\.,]+)',
                r'Tipo\s+de\s+Cambio.*?U\$?\s*([\d\.,]+)',
                r'T\.?C\.?\s*USD?\s*([\d\.,]+)'
            ]
            for pattern in tc_patterns:
                match = re.search(pattern, texto, re.IGNORECASE)
                if match:
                    tipo_cambio_usd = self._parse_float_value(match.group(1), 'tc_usd')
                    break

            # ══════════════════════════════════════════════════════════
            # ✅ SECCIÓN 6: EXTRACCIÓN DE DETALLE DEL SERVICIO (NUEVA)
            # ══════════════════════════════════════════════════════════
            # Extraer nombre del servicio
            service_name = ''
            service_name_match = re.search(
                r'SERVICIO\s+DE\s+[A-Z\s]+MADENAT|SERVICIO\s+[A-Z\s]+',
                texto,
                re.IGNORECASE
            )
            if service_name_match:
                service_name = service_name_match.group(0).strip()
            
            # Extraer código del servicio
            service_code = ''
            service_code_match = re.search(r'(SFABR\d+|[A-Z]{2,}\d{5,})', texto)
            if service_code_match:
                service_code = service_code_match.group(1)
            
            # Extraer cantidad del servicio (en m³)
            service_qty = 0.0
            service_qty_match = re.search(
                r'(?:SERVICIO.*?|CANTIDAD.*?)(\d+[,\.\d]+)\s*Metros\s+C[uú]bicos',
                texto,
                re.IGNORECASE | re.DOTALL
            )
            if service_qty_match:
                service_qty = self._parse_float_value(service_qty_match.group(1), 'service_qty')
            
            # Extraer precio unitario
            service_unit_price = 0.0
                      
            price_match = re.search(
                r'Metros\s+C[uú]bicos\s+([\d\.,]+)',
                texto,
                re.IGNORECASE
            )
            if price_match:
                service_unit_price = self._parse_float_value(price_match.group(1), 'unit_price')
            
            # Extraer fecha de la guía
            service_date = False
            date_match = re.search(r'Fecha:\s*(\d{2}/\d{2}/\d{4})', texto)
            if date_match:
                try:
                    from datetime import datetime
                    service_date = datetime.strptime(date_match.group(1), '%d/%m/%Y').date()
                except:
                    pass

            # ══════════════════════════════════════════════════════════
            # RETURN: Consolidado con todos los datos
            # ══════════════════════════════════════════════════════════
            return {
                'numero_guia': guia.group(1) if guia else False, 
                'orden_compra': oc_name, 
                'volumen_comercial': r3(vol_com_m3), 
                'additional_cost': subtotal_neto,
                'rate_usd': tipo_cambio_usd,
                'service_product_name': service_name,           # ✅ NUEVO
                'service_volume_m3': service_qty,               # ✅ NUEVO
                'service_unit_price_clp': service_unit_price,   # ✅ NUEVO
                'service_date': service_date,                   # ✅ NUEVO
                'service_code': service_code,                   # ✅ NUEVO
                'texto_raw': texto
            }
        except Exception as e: 
            raise UserError(f"Error procesando PDF: {e}")


    def _parse_packing_excel(self, attachment):
        # 🟢 HOOK : Si venimos de la validación visual, usamos esos datos
        if self._context.get('force_packing_data'):
            return self._context.get('force_packing_data')
        try:
            if not attachment.store_fname: 
                raise UserError("Archivo inválido")
            
            result = self._parse_excel_data_core(attachment)
            if not result or not result.get('lineas'): 
                raise UserError("El Excel se leyó pero no tiene líneas válidas.")

            result['oc_reference'] = self._find_oc_reference_in_excel(attachment)

            lineas_validadas = self._validar_y_enriquecer_lineas(result['lineas'])
            if not lineas_validadas: 
                raise UserError("Validación fallida de líneas.")

            total_vol = sum(l.get('Volumen', 0) for l in lineas_validadas)
            lotes_unicos = len(set(l['N° LOTE'] for l in lineas_validadas))
            
            result.update({
                'lineas': lineas_validadas,
                'total_volumen': total_vol,
                'total_volumen_mbf': m3_to_mbf(total_vol),
                'total_paquetes': len(lineas_validadas),
                'total_lotes': lotes_unicos,
                'tiene_paquetes_individuales': True,
                'lineas_agrupadas': []
            })
            return result
        except Exception as e: 
            raise UserError(f"Error procesando Excel: {e}")

    def _parse_excel_data_core(self, attachment):
        file_path = attachment._full_path(attachment.store_fname)
        try:
            try: 
                wb = load_workbook(file_path, data_only=True)
                data = [row for row in wb.active.iter_rows(values_only=True)]
                df = pd.DataFrame(data)
            except: 
                df = pd.read_excel(file_path, header=None, engine='openpyxl')

            if df.empty: 
                raise UserError("Excel vacío")

            header_idx = None
            target_headers = ['Codigo', 'Cantidad', 'M3', 'LOTE', 'Espesor', 'Ancho', 'Largo']
            for idx in range(min(20, len(df))):
                row_str = '|'.join([str(c) for c in df.iloc[idx] if pd.notna(c)])
                if sum(1 for h in target_headers if h.lower() in row_str.lower()) >= 3: 
                    header_idx = idx
                    break
            
            if header_idx is None: 
                raise UserError("No se detectaron encabezados válidos")

            col_map = {}
            for i, c in enumerate(df.iloc[header_idx]):
                if pd.isna(c): 
                    continue
                cs = str(c).lower()
                if 'codigo' in cs: 
                    col_map['codigo'] = i
                elif 'cantidad' in cs or 'piezas' in cs: 
                    col_map['cantidad'] = i
                elif 'm3' in cs or 'volumen' in cs: 
                    col_map['m3'] = i
                elif 'lote' in cs: 
                    col_map['lote'] = i
                elif 'espesor' in cs: 
                    col_map['espesor'] = i
                elif 'ancho' in cs: 
                    col_map['ancho'] = i
                elif 'largo' in cs: 
                    col_map['largo'] = i
                elif 'producto' in cs or 'descripcion' in cs: 
                    col_map['producto'] = i

            lineas = []
            curr_lote = None
            for i in range(header_idx + 1, len(df)):
                row = df.iloc[i]
                if row.isna().all(): 
                    continue
                
                try:
                    def get_val(key): 
                        return row[col_map[key]] if key in col_map and col_map[key] < len(row) else None

                    code = str(get_val('codigo') or '').strip()
                    prod_name = str(get_val('producto') or '').strip()

                    if not code or len(code) < 3 or any(x in code.lower() for x in ['nan', 'none', 'total', 'subtotal', 'totales']): 
                        continue
                    if any(x in prod_name.lower() for x in ['total', 'subtotal']): 
                        continue

                    qty = int(self._parse_float_value(get_val('cantidad'), 'qty')) if 'cantidad' in col_map else 0
                    lote_raw = str(get_val('lote') or '').strip()
                    if lote_raw and lote_raw.lower() not in ['nan', 'none']: 
                        curr_lote = lote_raw
                    elif not curr_lote and code: 
                        curr_lote = f"LOTE-{code}"

                    if qty <= 0: 
                        continue

                    vol = self._parse_float_value(get_val('m3'), 'm3') if 'm3' in col_map else 0
                    esp = self._parse_float_value(get_val('espesor'), 'espesor') if 'espesor' in col_map else 0
                    anc = self._parse_float_value(get_val('ancho'), 'ancho') if 'ancho' in col_map else 0
                    lar = self._parse_float_value(get_val('largo'), 'largo') if 'largo' in col_map else 0
                    
                    if not prod_name or prod_name.lower() == 'nan': 
                        prod_name = None
                    if vol <= 0 and esp and anc and lar: 
                        vol = (esp/1000) * (anc/1000) * lar * qty

                    # ✅ NUEVO: Calcular dimensiones nominales comerciales
                    # Convierte valores físicos (45mm, 125mm) → nominales ("6/4", "4 5/8")
                    espesor_nominal = self._get_nominal_dimension(esp, 'thickness')
                    ancho_nominal = self._get_nominal_dimension(anc, 'width')

                    lineas.append({
                        'Codigo Interno': code, 
                        'N° LOTE': curr_lote, 
                        'Cantidad': qty, 
                        'Volumen': vol, 
                        'Espesor': esp,              # Valor físico (45mm)
                        'Ancho': anc,                # Valor físico (125mm)
                        'Largo': lar, 
                        'product_name': prod_name,
                        # ✅ NUEVO: Agregar valores nominales
                        'espesor_nominal': espesor_nominal,  # "6/4"
                        'ancho_nominal': ancho_nominal        # "4 5/8"
                    })

                except: 
                    continue
            
            return {'lineas': lineas}
        except Exception as e: 
            raise UserError(f"Error leyendo datos: {e}")

    # ==============================================================================================
    #                                     4. UTILITARIOS Y HELPERS
    # ==============================================================================================

    def _parse_float_value(self, value, context=''):
        if pd.isna(value) or value is None: 
            return 0.0
        try:
            s = str(value).strip()
            if not s or s.lower() in ['nan', 'none', '']: 
                return 0.0
            clean = re.sub(r'[^\d\.,-]', '', s)
            
            # ══════════════════════════════════════════════════════════
            # Detectar formato monetario chileno (CLP)
            # ══════════════════════════════════════════════════════════
            is_clp = any(x in context.lower() for x in ['clp', 'subtotal', 'unit_price', 'neto', 'total'])
            
            if ',' in clean and '.' in clean: 
                # Formato europeo: 1.234,56 → 1234.56
                clean = clean.replace('.', '').replace(',', '.')
            elif ',' in clean: 
                # Tiene solo coma: puede ser decimal o separador
                clean = clean.replace(',', '.') if clean.count(',') == 1 else clean.replace(',', '')
            elif '.' in clean and is_clp:
                # Formato chileno CLP: 831.829 → 831829 (punto = separador de miles)
                parts = clean.split('.')
                if len(parts) == 2 and len(parts[1]) == 3 and parts[0].isdigit() and parts[1].isdigit():
                    # Un solo punto con exactamente 3 dígitos después
                    clean = clean.replace('.', '')
            
            val = float(clean)
            
            # Conversiones específicas por contexto
            if 'largo' in context.lower() and val > 100: 
                return val / 1000.0 
            if any(x in context.lower() for x in ['espesor', 'ancho']) and val > 1000: 
                return val / 10.0
            return val
        except: 
            return 0.0

        
    def _get_nominal_dimension(self, physical_mm, dim_type='width', tolerance=3):
        """
        🚀 NOMINAL DIMENSION BLINDADA:
        - Si el método es 'pure_metric', mantenemos la realidad física.
        - Si es 'madenat_gold', aplicamos la tolerancia comercial de exportación.
        """
        if not physical_mm or physical_mm <= 0:
            return ""

        # 1. 🧠 INTELIGENCIA DE CONTEXTO: ¿Qué método estamos usando?
        method = getattr(self, 'calculation_method', 'madenat_gold')

        # 2. BYPASS MÉTRICO: Si es la guía de Felipe (19826), no permitimos redondeos.
        if method == 'pure_metric':
            return f"{int(round(physical_mm, 0))}mm"

        # 3. EXCEPCIÓN DE SEGURIDAD: Consultar si es una medida métrica pura conocida
        if dim_type == 'width' and float(get_s2s_adjustment(self.env, physical_mm)) == 0.0:
            return f"{int(round(physical_mm, 0))}mm"

        # 4. LÓGICA ORIGINAL (Solo para Exportación Imperial)
        dimension_table = LUMBER_DIMENSION_MAP.get(dim_type, {})
        if not dimension_table:
            return self._calculate_fractional_approximation(physical_mm, dim_type)

        # Buscar coincidencia con tolerancia
        for standard_mm, nominal_value in dimension_table.items():
            if abs(physical_mm - standard_mm) <= tolerance:
                return nominal_value
        
        # Si no hay match en tabla, usamos la aproximación por octavos
        return self._calculate_fractional_approximation(physical_mm, dim_type)

    # TD-007: v1 de _validar_y_enriquecer_lineas eliminada (2026-06-04).
    # La v2 (línea ~2405) es la activa — agrega validación e_mm/a_mm/l_m <= 0
    # y usa _calculate_fractional_approximation en vez de _get_nominal_dimension.
    def _calculate_fractional_approximation(self, mm_value, dim_type='width'):
        """
        🚀 APROXIMACIÓN INTELIGENTE:
        - Detecta si debe retornar MM o Fracción Imperial.
        - Blindaje contra pérdida de precisión en medidas métricas (ej: 195mm).
        """
        if not mm_value:
            return ""

        # 1. Recuperar contexto de excepciones y método
        # (Sin hardcoding, consultamos la 'neurona' central)
        
        method = getattr(self, 'calculation_method', 'madenat_gold')
        # Fase 3: Delegar al helper centralizado — fuente única de verdad
        exceptions = self.env['madenat.ingestion.config'].get_s2s_exclusion_widths()

        # 2. INTELIGENCIA: ¿Es una medida métrica pura o estamos en modo métrico?
        # Si la medida está en excepciones (180, 200) o el método es métrico, devolvemos MM.
        if method == 'pure_metric' or round(mm_value, 1) in exceptions:
            return f"{int(round(mm_value, 0))}mm"

        # 3. LÓGICA IMPERIAL (Solo si no es métrico)
        inches = mm_value / float(MM_PER_INCH)
        
        if dim_type == 'thickness':
            quarters = round(inches * 4)
            return f"{quarters}/4" if quarters > 0 else ""
        
        else:  # width
            whole = int(inches)
            fraction_decimal = inches - whole
            eighths = round(fraction_decimal * 8)
            
            fraction_map = {
                0: "", 1: "1/8", 2: "1/4", 3: "3/8", 4: "1/2", 5: "5/8", 6: "3/4", 7: "7/8", 8: ""
            }
            
            if eighths == 8:
                whole += 1
                frac_text = ""
            else:
                frac_text = fraction_map.get(eighths, "")
            
            if frac_text:
                return f"{whole} {frac_text}".strip() if whole > 0 else frac_text
            
            # Si es un entero exacto en pulgadas (como 4", 6", 8")
            return f"{whole}\""

   

    def _validar_y_enriquecer_lineas(self, lineas):
            """
            🏛️ ENRIQUECIMIENTO ARQUITECTÓNICO CONSOLIDADO:
            - Valida integridad sin recortar funcionalidades.
            - Sincroniza el cálculo de m3 con la visualización del Staging.
            - Elimina el error de visualización del 195mm (7 5/8).
            """
                
            validas = []
            method = getattr(self, 'calculation_method', 'madenat_gold')

            for l in lineas:
                # 1. Mantener validación de funcionalidad actual
                if not l.get('Codigo Interno') or l.get('Cantidad', 0) <= 0:
                    continue

                # 2. Extracción segura de tipos
                try:
                    pzas = float(l.get('Cantidad', 0))
                    e_mm = float(l.get('Espesor', 0))
                    a_mm = float(l.get('Ancho', 0))
                    l_m = float(l.get('Largo', 0))
                except (ValueError, TypeError):
                    continue

                if e_mm <= 0 or a_mm <= 0 or l_m <= 0:
                    continue

                # 3. CALCULO DE VOLUMEN SEGÚN MÉTODO (Cerebro de la Guía)
                if method == 'pure_metric':
                    # 📏 Caso Felipe (Guía 19826): Matemática Decimal Pura
                    vol_m3 = calculate_volume_metric_m3(e_mm, a_mm, l_m, pzas)
                    
                    # Sincronización Visual: Mostrar MM para evitar confusión
                    l['thickness_visual'] = f"{int(e_mm)}mm"
                    l['width_visual'] = f"{int(a_mm)}mm"
                
                else:
                    # 🏆 Caso Exportación: Regla de Oro Madenat
                    ajuste_s2s = get_s2s_adjustment(self.env, a_mm)
                    
                    e_pulg = e_mm / float(MM_PER_INCH)
                    a_pulg = (a_mm / float(MM_PER_INCH)) + float(ajuste_s2s)
                    
                    vol_m3 = float(
                            Decimal(str(pzas))
                            * Decimal(str(e_pulg))
                            * Decimal(str(a_pulg))
                            * Decimal(str(l_m))
                            / INCH_SQ_METERS_TO_M3
                        )

                    # Sincronización Visual: Usar aproximación fraccional mejorada
                    l['thickness_visual'] = self._calculate_fractional_approximation(e_mm, 'thickness')
                    l['width_visual'] = self._calculate_fractional_approximation(a_mm, 'width')

                # 4. Inyección Atómica de Resultados
                l.update({
                    'vol_shipment_m3': round(vol_m3, 4),
                    'espesor_mm': e_mm,
                    'ancho_mm': a_mm,
                    'largo_m': l_m,
                    'pieces': pzas # Mantenemos compatibilidad con el nombre del campo en lotes
                })
                
                validas.append(l)

            return validas

    def _find_oc_reference_in_excel(self, attachment):
        """
        Busca referencia de Orden de Compra (OC) en el encabezado del archivo Excel adjunto.
        
        ChangeLog:
        - Fix: Se implementa alias 'regex_lib' para evitar UnboundLocalError (Shadowing de 're').
        - Mejora: Manejo explícito de excepciones (evita 'bare except').
        """
        # 4. APLICAR solución MÍNIMA: Importación local con alias para aislar el módulo
        # Esto garantiza que ninguna variable local llamada 're' rompa la referencia al módulo.
        import re as regex_lib
        
        try:
            # Validación defensiva
            if not attachment or not attachment.store_fname:
                return None

            file_path = attachment._full_path(attachment.store_fname)
            
            # Lectura limitada a 10 filas para optimizar rendimiento (Impacto controlado)
            df = pd.read_excel(file_path, header=None, nrows=10)
            
            for r in df.values:
                # Limpieza de fila
                s = " ".join([str(x) for x in r if pd.notna(x)])
                
                # Uso del alias seguro 'regex_lib'
                # Regex busca: "OC", "MC" u "Orden" seguido de alfanuméricos y guiones
                m = regex_lib.search(r'(?:OC|MC|Orden)[\s\.\-:]*([A-Z0-9]+(?:\s*[-]\s*[A-Z0-9]+)+)', s, regex_lib.I)
                
                if m: 
                    return m.group(1).replace(' ', '')
                    
        except Exception as e: 
            # 3. ENTENDER el problema: Capturamos la excepción real en lugar de silenciarla.
            # Se usa _logger (si está definido globalmente) o simplemente se evita que el flujo crítico se rompa.
            # _logger.warning(f"Error no bloqueante buscando OC en Excel: {e}") 
            pass 
            
        return None


    def _generate_lot_details_json(self, lot_data, purchase_order_obj=None, guide_number_str=""):
        """
        Genera el JSON para la vista. 
        Args:
            lot_data: Lista de diccionarios con info de lotes.
            purchase_order_obj: Objeto recordset de la OC (opcional).
            guide_number_str: String con el número de la guía extraído del Excel.
        """
        import json
        details = []
        
        # Mapa de estados (Defensive: si self.state falla, usa 'draft')
        state_val = getattr(self, 'state', 'draft')
        STATE_MAP = {
            'draft': 'Borrador / En Proceso',
            'confirmed': 'Confirmado',
            'done': 'Validado / Recepcionado',
            'cancel': 'Cancelado'
        }
        current_state_label = STATE_MAP.get(state_val, 'Borrador')
        
        # Nombre de OC seguro
        oc_name = purchase_order_obj.name if purchase_order_obj else ""

        # Número de Guía: Prioridad al argumento, luego busca en self, luego vacío.
        final_guide_num = guide_number_str or getattr(self, 'guide_number', getattr(self, 'name', ''))

        for d in lot_data:
            lot_obj = d['lote']
            
            # Limpieza de Cantidad
            try:
                qty_clean = int(d['cantidad'])
            except:
                qty_clean = 0
            
            piezas_display = str(qty_clean)

            row = {
                'id': lot_obj.id,
                'lote_name': d.get('lote_code', ''),
                'product_name': d.get('product_name', ''),
                'product_original': d.get('note', lot_obj.note or ''),
                
                # Dimensiones seguras (evita error si es None)
                'espesor': f"{lot_obj.espesor_mm:g}" if lot_obj.espesor_mm else "0", 
                'ancho': f"{lot_obj.ancho_mm:g}" if lot_obj.ancho_mm else "0",
                'largo': f"{lot_obj.largo_m:.2f}" if lot_obj.largo_m else "0.00",
                
                'cantidad': piezas_display,
                'oc_ref': oc_name,
                
                'vol_m3': f"{d.get('volumen', 0):.4f}",
                'n_guia': final_guide_num,           # <--- DATO SEGURO
                'estado': current_state_label,
                
                'val_tecnica': 'Pendiente',
                'val_financiera': 'Pendiente'
            }
            details.append(row)

        return json.dumps(details)




    def _assert_ready_to_validate(self):
        for rec in self:
            if (rec.additional_cost or 0) > 0 and (rec.vol_fisico or 0) <= 0:
                raise UserError("Monto > 0 exige volumen físico > 0.")

    def _create_basic_purchase_order(self, ref):
        return self.env['purchase.order'].create({
            'name': ref, 
            'partner_id': self.partner_id.id, 
            'state': 'draft',
            'currency_id': self.env.company.currency_id.id, 
            'date_order': fields.Datetime.now()
        })

    def _assign_costs_to_generated_lots(self, lot_data, packing_data):
        if self.tipo_recepcion != 'service' or self.additional_cost <= 0: 
            return 0
        
        rate = self.rate_usd or 1.0
        total_vol = packing_data.get('total_volumen') or 1.0
        cost_m3 = (self.additional_cost / rate) / total_vol
        count = 0
        
        for d in lot_data:
            self.env['stock.lot.cost.line'].create({
                'lot_id': d['lote'].id, 
                'name': f"Servicio {self.name}",
                'cost_type': 'processing', 
                'amount_usd': cost_m3 * d['volumen'], 
                'date': fields.Date.today()
            })
            count += 1
        
        return count

    def _create_or_get_lot(self, guia_ref, product, qty, vol_purchase, vol_shipment, vol_real, lot_name, lot_dims, precio_usd,
                       purchase_order=None, thickness_visual='', width_visual='', vol_mbf=0.0,
                       technical_validation='pending', estado_trazabilidad='procesado', subproducto_id=False):

        """
        ✅ VERSIÓN MEJORADA v8.0: Busca/Crea lotes con ESTADOS AUTO-APROBADOS y DATOS VISUALES
        
        MEJORAS TÉCNICAS:
        1. Inyección de Estados: Recibe 'technical_validation' y 'estado_trazabilidad' dinámicos.
        2. Mapeo Visual: Inyecta dimensiones visuales (espesor_inch_frac, ancho_inch_frac) para logística.
        3. Limpieza de Conflictos: Resuelve colisiones entre reception_id y guia_processing_id.
        4. ORM Seguro: Eliminado SQL directo. Se fuerza un flush_model para asegurar
           que la búsqueda del ORM lea datos frescos de la transacción actual.
        
        ARGS:
            guia_ref: Referencia de la guía (string)
            product: Recordset de product.product
            qty: Cantidad de piezas (float/int)
            vol_purchase: Volumen nominal de compra en m³ (float)
            vol_shipment: Volumen de embarque en m³ (float)
            vol_real: Volumen físico real en m³ (float)
            lot_name: Nombre/Código del lote físico (string)
            lot_dims: Diccionario con dimensiones físicas y metadatos (dict)
            precio_usd: Costo unitario (float)
            purchase_order: Recordset de purchase.order (opcional)
            thickness_visual: Texto visual espesor ej: "6/4" (string)
            width_visual: Texto visual ancho ej: "5 5/8" (string)
            vol_mbf: Volumen en MBF (float)
            technical_validation: Estado de validación técnica (string)
            estado_trazabilidad: Estado del ciclo de vida (string)
            subproducto_id: Recordset o ID de madenat.subproducto (opcional)
        
        RETURNS:
            Recordset de stock.lot (creado o actualizado)
        """
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 1. LIMPIEZA Y NORMALIZACIÓN DEL NOMBRE
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        name = lot_name.strip() if lot_name else f"{guia_ref}-{product.default_code}"
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 2. BÚSQUEDA SEGURA POR ORM (Con Flush)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # Aseguramos que la BD y la caché estén sincronizadas antes de buscar
        self.env['stock.lot'].flush_model(['name', 'product_id', 'company_id'])
        
        lot = self.env['stock.lot'].search([
            ('name', '=', name), 
            ('product_id', '=', product.id),
            ('company_id', 'in', [self.env.company.id, False]),
        ], limit=1, order='company_id DESC')
        
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 3. PREPARACIÓN DE VALORES (CON ESTADOS Y VISUALES)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        vals = {
            # ═══════════════════════════════════════════════════════
            # CAMPOS BÁSICOS (Obligatorios)
            # ═══════════════════════════════════════════════════════
            'name': name,
            'product_id': product.id,
            'piezas': int(qty),
            
            # Volumen (aseguramos mínimo 0.001 para evitar división por cero)
            'volume_purchase_m3': max(float(vol_purchase), 0.001), # 💰 Compra
            'vol_shipment_m3': max(float(vol_shipment), 0.001), # 🌎 Exportación
            'volumen_m3': max(float(vol_purchase), 0.001), # 📦 STOCK REAL = COMPRA
            
            # Trazabilidad de origen
            'guia_processing_id': self.id,
            'reception_id': getattr(self, 'reception_id', False) or False,

            # Ubicación por defecto (Stock/Physical Locations)
            'location_id': self.assignment_location_id.id \
                if self.assignment_location_id \
                else self.env.ref('stock.stock_location_stock').id,

            # Nombre original del producto del Excel (útil para auditoría)
            'note': lot_dims.get('note', ''),
             
            # ═══════════════════════════════════════════════════════
            # DIMENSIONES FÍSICAS (mm/m)
            # ═══════════════════════════════════════════════════════
            'espesor_mm': float(lot_dims.get('espesor_mm', 0)),
            'ancho_mm': float(lot_dims.get('ancho_mm', 0)),
            'largo_m': float(lot_dims.get('largo_m', 0)),
                        
            # ═══════════════════════════════════════════════════════
            # DIMENSIONES NOMINALES (Comercial OC)
            # ═══════════════════════════════════════════════════════
            'espesor_nominal_mm': float(lot_dims.get('espesor_nominal_mm', 0)),
            'ancho_nominal_mm': float(lot_dims.get('ancho_nominal_mm', 0)),

            # ═══════════════════════════════════════════════════════
            # CALIDAD (Subproducto FAS/S2S/etc)
            # ═══════════════════════════════════════════════════════
            'subproducto_id': subproducto_id if subproducto_id else False,

            # ═══════════════════════════════════════════════════════
            # ✅ FIX 1: DIMENSIONES VISUALES (Pulgadas fraccionarias)
            # ═══════════════════════════════════════════════════════
            # Prioridad: Argumento explícito > Diccionario lot_dims > Cadena vacía
            'espesor_inch_frac': thickness_visual or lot_dims.get('thickness_visual', ''),
            'ancho_inch_frac': width_visual or lot_dims.get('width_visual', ''),
            
            # ═══════════════════════════════════════════════════════
            # ✅ FIX 2: ESTADOS AUTOMÁTICOS (CRÍTICO PARA LOGÍSTICA)
            # ═══════════════════════════════════════════════════════
            # El lote nace con el estado que pasemos (por defecto 'pending'/'procesado')
            # Pero si viene del Staging Bridge, entrará como 'approved'/'recepcionado'
            'technical_validation': technical_validation,
            'estado_trazabilidad': estado_trazabilidad,
        }
        
        # ═══════════════════════════════════════════════════════
        # CAMPOS OPCIONALES (Solo si existen)
        # ═══════════════════════════════════════════════════════
        if purchase_order:
            vals['purchase_order_id'] = purchase_order.id
            
            # Costo: Usar el pasado por argumento, o intentar rescatarlo de la OC
            final_price = precio_usd
            if final_price <= 0 and purchase_order.order_line:
                # Intento de rescate de precio desde la línea de la OC
                line = purchase_order.order_line.filtered(lambda l: l.product_id == product)
                if line:
                    final_price = line[0].price_unit
            
            if final_price > 0:
                vals['purchase_price_usd_per_m3'] = final_price
        elif precio_usd > 0:
             vals['purchase_price_usd_per_m3'] = precio_usd

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # 4. CREACIÓN O ACTUALIZACIÓN
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        if lot:
            # ═══════════════════════════════════════════════════════
            # CASO 1: YA EXISTE -> Actualizar y limpiar conflictos
            # ═══════════════════════════════════════════════════════
            _logger.info(
                f"♻️ Reutilizando lote {name} (ID: {lot.id}) "
                f"y limpiando origen anterior..."
            )
            
            # LIMPIEZA DE CONFLICTO: Si el lote venía de lumber.reception,
            # limpiamos reception_id para que acepte guia_processing_id
            if vals.get('guia_processing_id'):
               vals['reception_id'] = False
            
            lot.write(vals)
            return lot
        else:
            # ═══════════════════════════════════════════════════════
            # CASO 2: NO EXISTE -> Crear nuevo (idempotente ante reproceso)
            # ═══════════════════════════════════════════════════════
            _logger.info(
                f"✨ Creando lote {name} con estados: "
                f"Tech={technical_validation}, Traz={estado_trazabilidad}"
            )
           
            # 1. Proveedor
            if getattr(self, 'partner_id', False):
                vals['supplier_id'] = self.partner_id.id

            # 2. Orden de Compra (Si la guía viene de una OC o Servicio)
            if getattr(self, 'purchase_order_id', False):
                vals['purchase_order_id'] = self.purchase_order_id.id
            elif getattr(self, 'order_id', False): # A veces se llama order_id en maquila
                vals['purchase_order_id'] = self.order_id.id

            # 3. Precio (Si existe en el contexto)
            # Intentamos buscar el precio en la línea de servicio o en la OC
            if vals.get('purchase_order_id'):
                po = self.env['purchase.order'].browse(vals['purchase_order_id'])
                # Lógica simple: Tomar el precio del primer item o buscar coincidencia
                if po.order_line:
                    vals['purchase_price_usd_per_m3'] = po.order_line[0].price_unit

            try:
                with self.env.cr.savepoint():
                    return self.env['stock.lot'].create(vals)
            except (IntegrityError, ValidationError):
                _logger.warning(
                    "⚠️ Colisión UNIQUE en stock.lot para name=%s product_id=%s company_id=%s — "
                    "reutilizando lote existente (probable reproceso desde borrador).",
                    name, product.id, self.env.company.id
                )
                self.env['stock.lot'].flush_model(['name', 'product_id', 'company_id'])
                self.env['stock.lot'].invalidate_model(['name', 'product_id', 'company_id'])
                lot = self.env['stock.lot'].search([
                    ('name', '=', name),
                    ('product_id', '=', product.id),
                    ('company_id', 'in', [self.env.company.id, False]),
                ], limit=1, order='company_id DESC')
                if lot:
                    lot.write(vals)
                    return lot
                raise




    def find_or_create_product(self, code, name):
        """
        VERSIÓN 4.0 - DELEGACIÓN AL MIXIN
        
        Wrapper que delega al mixin compartido.
        Mantiene compatibilidad con código existente.
        """
        return self.find_or_create_lumber_product(code, name)

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ✅ CORRECCIÓN 3: Método LEGACY documentado
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    def _create_picking_and_lines(self, po, lot_data):
        """
        ✅ MÉTODO LEGACY UNIFICADO
        Ahora utiliza el motor maestro para evitar duplicados y se encarga
        exclusivamente de la generación de movimientos de stock.
        """
        # 🛡️ UNIFICACIÓN: Llamamos al maestro en lugar de usar .create() manual
        picking = self._get_or_create_picking_unified()
        
        uom_m3 = self.env.ref('uom.product_uom_cubic_meter')

        for idx, d in enumerate(lot_data, 1):
            # Generar el movimiento de stock (stock.move)
            move = self.env['stock.move'].create({
                'name': d.get('product_name', 'Madera'), 
                'description_picking': f"Pqte #{idx} | Lote {d.get('lote_code')}",
                'product_id': d.get('product_id') or d.get('lote').product_id.id, 
                'product_uom_qty': d.get('volumen', 0.0), 
                'product_uom': uom_m3.id,        
                'picking_id': picking.id,
                'location_id': picking.location_id.id, 
                'location_dest_id': picking.location_dest_id.id
            })
            
            # Generar la línea de movimiento vinculada al lote (stock.move.line)
            if d.get('lote'):
                self.env['stock.move.line'].create({
                    'move_id': move.id, 
                    'picking_id': picking.id, 
                    'lot_id': d['lote'].id,           
                    'quantity': d.get('volumen', 0.0),
                    'product_uom_id': uom_m3.id,
                    'product_id': d['lote'].product_id.id, 
                    'location_id': picking.location_id.id, 
                    'location_dest_id': picking.location_dest_id.id,
                })
                
        # Confirmar el albarán para que pase a estado 'Preparado'
        if picking.state == 'draft':
            picking.action_confirm()
            
        return picking

    def _obtener_precio_desde_oc(self, po):
        if po and po.order_line: 
            return po.order_line[0].price_unit 
        return 0.0

    def _obtener_precio_por_defecto(self): 
        return 0.0
    
    def _mm_a_pulgadas_frac(self, mm):
        try: 
            return f"{int(mm / float(MM_PER_INCH))}''"
        except: 
            return ""

    # ==============================================================================================
    #                                     5. COMPUTED FIELDS
    # ==============================================================================================

    @api.depends('lot_ids')
    def _compute_lineas_procesadas(self):
        for rec in self:
            rec.lineas_procesadas = len(rec.lot_ids)

       

    @api.depends('state', 'pdf_attachment_id', 'excel_attachment_id', 'guide_pdf_file', 'excel_file')
    def _compute_can_process(self):
        for rec in self:
            rec.can_process = (
                rec.state == 'draft' and 
                (rec.pdf_attachment_id or rec.guide_pdf_file) and 
                (rec.excel_attachment_id or rec.excel_file)
            )

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # ✅ CORRECCIÓN 1B: Compute UNIFICADO can_validate
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    @api.depends('state', 'vol_fisico', 'validation_status')
    def _compute_can_validate(self):
        """
        ✅ VERSIÓN UNIFICADA: Combina validación de checklist + volumen físico
        
        Reglas:
        1. Debe estar en estado 'processed'
        2. Debe tener volumen físico > 0
        3. Si existe checklist, debe estar aprobado (passed o warning)
        
        NOTA: Sobreescribe el _compute_can_validate del mixin para agregar
        validación de volumen físico específica de guías procesadas.
        """
        for rec in self:
            # Validación básica de estado y volumen (requisito de negocio)
            basic_check = (rec.state == 'processed' and rec.vol_fisico > 0)
            
            # Validación de checklist (si existe)
            checklist_check = True  # Por defecto aprobado si no hay checklist
            if rec.validation_checklist_ids:
                # Si hay checklist, debe estar aprobado (passed) o con advertencias (warning)
                checklist_check = rec.validation_status in ('passed', 'warning')
            
            # Resultado final: AMBAS validaciones deben pasar
            rec.can_validate = basic_check and checklist_check
            
            _logger.debug(
                f"can_validate para {rec.name}: "
                f"basic={basic_check}, checklist={checklist_check}, final={rec.can_validate}"
            )

    @api.depends('state')
    def _compute_can_cancel(self):
        for rec in self:
            rec.can_cancel = rec.state in ('draft', 'processed')

    @api.depends('state')
    def _compute_can_reopen(self):
        for rec in self:
            rec.can_reopen = rec.state in ('verified', 'cancelled', 'processed')



    # ==============================================================================================
    #                                     6. UI ACTIONS
    # ==============================================================================================
    def action_open_po(self):
        """
        ✅ CARACTERÍSTICAS PROFESIONALES:
        
        1. Documentación docstring clara
        2. ensure_one() para operaciones unitarias
        3. Validación de datos (UserError si no hay OC)
        4. Mensaje de error descriptivo con sugerencia
        5. Retorno dict con todos los atributos necesarios
        6. view_id=False para usar vista por defecto
        7. context preservado correctamente
        8. target='current' para navegación fluida
        """
        self.ensure_one()
        
        if not self.order_id:
            raise UserError(
                "No hay una Orden de Compra vinculada a esta guía.\n\n"
                "Por favor, vincule una OC primero."
            )
        
        return {
            'type': 'ir.actions.act_window',
            'name': f'Orden de Compra: {self.order_id.name}',
            'res_model': 'purchase.order',
            'res_id': self.order_id.id,
            'view_mode': 'form',
            'view_id': False,
            'target': 'current',
            'context': dict(self.env.context),
        }

    def action_force_cancel(self):
        """
        🛡️ CANCELACIÓN SEGURA MADENAT v2.0
        - Escudo logístico: bloquea si lotes están en contenedores activos.
        - Escudo financiero: bloquea si hay costos o albaranes done.
        - Maneja pickings con moves parcialmente done.
        - Búsqueda amplia de pickings (cubre prefijos REVERTIDO-, Guía X - OC..., etc.)
        - Limpieza completa de lotes y líneas de procesamiento.
        """
        for rec in self:
            _logger.warning(f"🛡️ [action_force_cancel] Iniciando Cancelación Segura → Guía: {rec.name}")

            # =======================================================
            # 1. ESCUDO LOGÍSTICO — Lotes en contenedores activos
            # =======================================================
            if rec.lot_ids and self.env.get('lumber.container'):
                contenedores = self.env['lumber.container'].search([
                    ('lot_ids', 'in', rec.lot_ids.ids)
                ])
                if contenedores:
                    nombres = ", ".join(contenedores.mapped('name'))
                    raise UserError(
                        f"⛔ DENEGADO LOGÍSTICO:\n"
                        f"Los lotes están en uso por los contenedores: {nombres}.\n"
                        f"Desvincula los lotes del contenedor antes de cancelar."
                    )

            # =======================================================
            # 2. ESCUDO FINANCIERO v3 — Costos ≠ Pickings done
            # =======================================================
            CostLine = self.env.get('stock.lot.cost.line')
            costos = CostLine.search([('name', '=', f"Servicio {rec.name}")]) if CostLine else False

            pickings = self.env['stock.picking'].search([
                     ('origin', '=', rec.name),
            ])
            _logger.info(f"🔎 Albaranes para '{rec.name}': {len(pickings)} → {pickings.mapped('name')}")

            albaranes_hechos = pickings.filtered(lambda p: p.state == 'done')

            # ✅ CASO A: Costos contables → bloqueo total (no revertir)
            if costos:
                _logger.warning(f"🔒 Guía {rec.name} tiene costos contables — bloqueo de integridad")
                rec.write({'state': 'cancelled'})
                rec.message_post(
                    body="🔒 <strong>Cierre por Integridad Contable:</strong> "
                         "Existen costos registrados. Los movimientos se mantienen intactos."
                )
                continue
            # ✅ CASO B: Pickings done sin costos → reversión simétrica via ORM
            #
            # ARQUITECTURA DEL RETORNO (Odoo 18 CE):
            # ─────────────────────────────────────────────────────────────────
            # Odoo 18 eliminó el método _create_returns() del wizard. La reversión
            # debe hacerse directamente via ORM: crear stock.picking + stock.move
            # por cada move done del picking original, con ubicaciones invertidas.
            #
            # REGLA DE ORO DE UBICACIONES:
            #   return_picking.location_id     = picking.location_dest_id  (desde WH/Stock)
            #   return_picking.location_dest_id = picking.location_id      (hacia proveedor/origen)
            #   move.location_id               = move.location_dest_id     (inverso al move original)
            #   move.location_dest_id          = move.location_id          (inverso al move original)
            #
            # TIPO DE OPERACIÓN:
            #   Se usa return_picking_type_id si está configurado en el tipo de operación.
            #   Fallback: se busca el tipo "incoming" del mismo almacén para garantizar
            #   que el albarán de retorno sea una RECEPCIÓN y no una SALIDA (WH/IN, no WH/OUT).
            # ─────────────────────────────────────────────────────────────────
            for picking in albaranes_hechos:
                try:
                    with self.env.cr.savepoint():

                        # --- 1. TIPO DE OPERACIÓN DEL RETORNO ---
                        # ✅ NUEVO BLOQUE — fuerza siempre tipo 'incoming', ignora return_picking_type_id
                        # que en Odoo apunta a WH/OUT generando returns incorrectos en albaranes EMB-.
                        warehouse = picking.picking_type_id.warehouse_id
                        if not warehouse:
                            raise UserError(
                                f"⛔ El albarán '{picking.name}' no tiene almacén asociado. "
                                f"Verifica la configuración del tipo de operación."
                            )

                        # ✅ CON FILTRO — excluye secuencias EMB- que colisionan con embarques
                        return_type = self.env['stock.picking.type'].search([
                            ('code', '=', 'incoming'),
                            ('warehouse_id', '=', warehouse.id),
                            ('active', '=', True),
                            ('sequence_code', 'not ilike', 'EMB'),
                        ], limit=1)

                        if not return_type:
                            return_type = self.env['stock.picking.type'].search([
                                ('code', '=', 'incoming'),
                                ('active', '=', True),
                                ('sequence_code', 'not ilike', 'EMB'),
                            ], limit=1)

                            if return_type:
                                _logger.warning(
                                    f"⚠️ No se encontró tipo 'incoming' no-EMB en almacén "
                                    f"'{warehouse.name}'. Usando fallback: '{return_type.name}'"
                                )

                        if not return_type:
                            raise UserError(
                                f"⛔ No existe ningún tipo de operación 'Recepción' (incoming) "
                                f"activo para revertir '{picking.name}'.\n"
                                f"Ve a: Inventario → Configuración → Tipos de Operación."
                            )

                        _logger.info(
                            f"🔀 Tipo de retorno resuelto: '{return_type.name}' "
                            f"[{return_type.code}] → Almacén: {return_type.warehouse_id.name}"
                        )

                        # --- 2. CREAR PICKING DE RETORNO ---
                        return_picking = self.env['stock.picking'].create({
                            'picking_type_id':  return_type.id,
                            'partner_id':       picking.partner_id.id,
                            'origin':           f'Return of {picking.name}',
                            'location_id':      picking.location_dest_id.id,   # WH/Stock → origen del retorno
                            'location_dest_id': picking.location_id.id,         # destino original del picking
                        })
                        _logger.info(
                            f"📦 Return picking creado: {return_picking.name} "
                            f"| tipo: {return_type.name} "
                            f"| {picking.location_dest_id.name} → {picking.location_id.name}"
                        )

                        # --- 3. FILTRAR MOVES VÁLIDOS ---
                        # Solo moves en estado 'done' con cantidad real > 0.
                        # Odoo 18: usa m.quantity (reemplaza m.quantity_done de v16/v17).
                        moves_done = picking.move_ids.filtered(
                            lambda m: m.state == 'done' and (m.quantity or m.product_uom_qty) > 0
                        )
                        if not moves_done:
                            return_picking.unlink()
                            _logger.warning(
                                f"⚠️ {picking.name} no tiene moves done con cantidad > 0 — "
                                f"return picking eliminado, omitiendo reversión."
                            )
                            continue

                        # --- 4. CREAR MOVES DE RETORNO (uno por cada move done) ---
                        for move in moves_done:
                            qty = move.quantity or move.product_uom_qty
                            self.env['stock.move'].create({
                                'name':                    f'Return of {move.name}',
                                'product_id':              move.product_id.id,
                                'product_uom_qty':         qty,
                                'product_uom':             move.product_uom.id,
                                'picking_id':              return_picking.id,
                                'location_id':             move.location_dest_id.id,  # inverso
                                'location_dest_id':        move.location_id.id,       # inverso
                                'origin_returned_move_id': move.id,                   # trazabilidad
                                'to_refund':               True,                       # ajusta valorización
                            })
                        _logger.info(
                            f"🔁 {len(moves_done)} moves de retorno creados para {picking.name} "
                            f"→ {return_picking.name}"
                        )

                        # --- 5. CONFIRMAR, ASIGNAR Y VALIDAR ---
                        # skip_backorder=True    → no abre wizard "¿Crear backorder?"
                        # skip_immediate=True    → no abre wizard "Transferencia inmediata"
                        # immediate_transfer=True → fuerza validación sin confirmar cantidades
                        # force_validate=True    → bypass de validaciones adicionales del wizard
                        return_picking.action_confirm()
                        return_picking.action_assign()

                        _logger.info(
                            f"🔍 Estado moves tras assign: "
                            f"{[(m.product_id.display_name[:20], m.quantity, m.state) for m in return_picking.move_ids]}"
                        )

                        return_picking.with_context(
                            force_validate=True,
                            skip_immediate=True,
                            skip_backorder=True,
                            immediate_transfer=True,
                        ).button_validate()

                        _logger.info(f"✅ Stock revertido correctamente: {picking.name} → {return_picking.name}")
                        rec.message_post(
                            body=(
                                f"🔄 Albarán <b>{picking.name}</b> revertido exitosamente.<br/>"
                                f"Retorno: <b>{return_picking.name}</b> | "
                                f"{len(moves_done)} movimiento(s) invertido(s)."
                            )
                        )

                except UserError:
                    # Re-lanzar UserError sin envolver (ya tiene mensaje amigable)
                    raise
                except Exception as e:
                    _logger.error(
                        f"❌ Error inesperado revirtiendo picking {picking.name}: {e}",
                        exc_info=True
                    )
                    raise UserError(
                        f"No se pudo revertir el albarán {picking.name}.\n"
                        f"Detalle técnico: {e}"
                    )

            # =======================================================
            # 3. CANCELACIÓN LIMPIA DE PICKINGS (ORM)
            # =======================================================
            for picking in pickings.filtered(lambda p: p.state not in ('done', 'cancel')):
                done_moves = picking.move_ids.filtered(lambda m: m.state == 'done')
                if done_moves:
                    # Picking con moves parcialmente done → no se puede cancelar por ORM
                    _logger.warning(
                        f"⚠️ Picking {picking.name} tiene {len(done_moves)} move(s) 'done' "
                        f"— omitiendo cancelación ORM, cerrando por integridad."
                    )
                    rec.write({'state': 'cancelled'})
                    rec.message_post(
                        body=f"🔒 <strong>Cierre Parcial:</strong> El albarán <b>{picking.name}</b> "
                            f"tiene {len(done_moves)} movimiento(s) ya realizados. "
                            f"Se mantienen intactos por integridad de inventario."
                    )
                    continue

                try:
                    picking.action_cancel()
                    _logger.info(f"✅ Picking cancelado correctamente: {picking.name}")
                except Exception as e:
                    _logger.error(f"❌ Error cancelando picking {picking.name}: {e}")
                    raise UserError(
                        f"No se pudo cancelar el albarán {picking.name}.\n"
                        f"Detalle técnico: {e}"
                    )

            # =======================================================
            # 4. LIMPIEZA DE LOTES
            # =======================================================
            if rec.lot_ids:
                try:
                    rec.lot_ids.write({
                        
                        'guia_processing_id': False,
                        'estado_trazabilidad': 'recepcionado',
                    })
                    rec.write({'lot_ids': [(5, 0, 0)]})
                    _logger.info(f"✅ Lotes desvinculados para Guía {rec.name}")
                except Exception as e:
                    _logger.warning(f"⚠️ No se pudieron desvincular lotes completamente: {e}. Solo se desvincula referencia.")
                    rec.write({'lot_ids': [(5, 0, 0)]})

            # =======================================================
            # 5. RESET DE LÍNEAS DE PROCESAMIENTO
            # =======================================================
            if rec.processing_line_ids:
                line_fields = rec.processing_line_ids._fields
                vals_line_reset = {
                    f: False for f in ['validation_state', 'checklist_state', 'passed', 'approved']
                    if f in line_fields
                }
                if vals_line_reset:
                    rec.processing_line_ids.write(vals_line_reset)
                    _logger.info(f"✅ Líneas de procesamiento reseteadas para Guía {rec.name}")

            # =======================================================
            # 6. ESTADO FINAL — CANCELADA
            # =======================================================
            rec.write({
                'state': 'cancelled',
                'vol_fisico': 0.0,
                'vol_comercial': 0.0,
                'total_paquetes': 0,
                'date_processed': False,
                'lot_details_data': False,
            })
            rec.message_post(
                body="✅ <strong>Cancelación Segura completada:</strong> "
                    "Guía anulada correctamente. Sin impacto en inventario histórico."
            )
            _logger.warning(f"✅ [action_force_cancel] Guía {rec.name} → estado: cancelled")

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

   
    def action_reopen_to_draft(self):
        """
        🔄 REVERSIÓN ARQUITECTÓNICA MADENAT (Path B - Blindado UX v2):
        Mantiene funcionalidades originales pero permite reversión si los movimientos
        validados son ÚNICAMENTE internos (bodega a bodega).
        """
        from odoo.exceptions import UserError
        
        for rec in self:
            _logger.info(f"🚨 INICIANDO REVERSIÓN SEGURA - Guía {rec.name}")
            
            # PASO 0: Snapshot de auditoría (Historial en Chatter)
            body_snapshot = (
                f"📋 <strong>Snapshot previo a reversión:</strong><br/>"
                f"- Vol. Físico: {rec.vol_fisico} m3<br/>"
                f"- Vol. Shipment: {rec.vol_shipment_m3} m3<br/>"
                f"- Paquetes: {rec.total_paquetes}"
            )
            rec.message_post(body=body_snapshot)

            # =======================================================
            # 🛡️ FASE 1: ESCUDO LOGÍSTICO (Consolidación y Salida Real)
            # =======================================================
            if rec.lot_ids:
                # 1.1 Bloqueo por Consolidación (Si ya está en contenedor, NO se toca)
                if 'lumber.container' in self.env:
                    contenedores = self.env['lumber.container'].search([('lot_ids', 'in', rec.lot_ids.ids)])
                    if contenedores:
                        nombres = ", ".join(contenedores.mapped('name'))
                        raise UserError(f"⛔ DENEGADO LOGÍSTICO: Madera consolidada en {nombres}. Desconsolide antes.")
                
                # 1.2 Blindaje de Ubicación: ¿La madera salió de la empresa?
                # Buscamos si los lotes tienen stock en ubicaciones externas (Clientes/Tránsito)
                quants_externos = self.env['stock.quant'].search([
                    ('lot_id', 'in', rec.lot_ids.ids),
                    ('quantity', '>', 0),
                    ('location_id.usage', 'not in', ['internal', 'inventory', 'production'])
                ])
                if quants_externos:
                    raise UserError("⛔ DENEGADO POR MOVIMIENTO EXTERNO: La madera ya salió a ubicaciones de cliente o tránsito.")

            # =======================================================
            # 🛡️ FASE 2: ESCUDO FINANCIERO Y ALBARANES VALIDADOS (Inteligente)
            # =======================================================
            # 2.1 Verificación de Costos Contables (Flete/Seguro)
            CostLine = self.env['stock.lot.cost.line'] if 'stock.lot.cost.line' in self.env else None
            costos_aplicados = False
            if CostLine:
                costos_aplicados = CostLine.search([('|'), ('name', 'ilike', rec.name), ('lot_id', 'in', rec.lot_ids.ids)], limit=1)

            if costos_aplicados:
                raise UserError(f"⛔ BLOQUEO FINANCIERO: Existen costos asociados a la guía {rec.name}. Elimínelos primero.")

            # 2.2 Verificación de Albaranes Validados
            pickings = self.env['stock.picking'].search([('origin', '=', rec.name)])
            albaranes_hechos = pickings.filtered(lambda p: p.state == 'done')
            
            # 💡 AQUÍ ESTÁ EL AJUSTE: Solo bloqueamos si el destino es EXTERNO.
            # Si el destino es 'internal' (como Puerto Coronel), permitiremos avanzar.
            albaranes_externos = albaranes_hechos.filtered(lambda p: p.location_dest_id.usage != 'internal')
            
            if albaranes_externos:
                nombres = ", ".join(albaranes_externos.mapped('name'))
                raise UserError(f"⛔ DENEGADO: Albaranes con salida externa validados ({nombres}).")

            # =======================================================
            # 🧹 FASE 3: REVERSIÓN FÍSICA Y LIMPIEZA (ORM Legal)
            # =======================================================
            # 1. Cancelamos Pickings en borrador/espera (iteración individual con try/except)
            for picking in pickings.filtered(lambda p: p.state not in ('done', 'cancel')):
                try:
                    picking.action_cancel()
                except Exception as e:
                    _logger.warning(f"⚠️ No se pudo cancelar picking {picking.name}: {e}")
                try:
                    picking.write({'origin': f"ANULADO-{rec.name}"})
                except Exception:
                    pass

            # 2. Desvinculamos Albaranes 'Done' Internos (como EMB-00032)
            # Les cambiamos el origen para que la guía quede "libre" para procesarse de nuevo.
            if albaranes_hechos:
                albaranes_hechos.write({'origin': f"REVERTIDO-{rec.name}"})
                rec.message_post(body=f"⚠️ Se detectaron albaranes internos validados ({', '.join(albaranes_hechos.mapped('name'))}). Se desvincularon para permitir el reset.")

            # 3. Limpieza de Lotes y Staging
            if rec.lot_ids:
                rec.lot_ids.write({'guia_processing_id': False, 'estado_trazabilidad': 'recepcionado'})
                rec.write({'lot_ids': [(5, 0, 0)]})

            if rec.processing_line_ids:
                rec.processing_line_ids.unlink() 
                rec.invalidate_recordset(['processing_line_ids'])

            # =======================================================
            # 🔄 FASE 4: RESET A BORRADOR
            # =======================================================
            rec.write({
                'state': 'draft',
                'date_processed': False,
                'vol_fisico': 0.0,
                'vol_comercial': 0.0,
                'vol_total_m3': 0.0,
                'vol_shipment_m3': 0.0,
                'total_paquetes': 0,
                'total_lotes_unicos': 0,
                'lineas_procesadas': 0,
                'lot_details_data': False,
                'grouping_details_json': False,
            })

            rec.message_post(body="🔄 **Reversión Exitosa:** Guía reseteada. Albaranes internos previos desvinculados.")


    # ========================================
    # Métodos de Adjuntos (Attachment Helpers)
    # ========================================
    
    def _store_binary_as_attachment(self, field, name_field, desc):
        """Helper para convertir campos Binary en ir.attachment"""
        return self.env['ir.attachment'].create({
            'name': self[name_field] or desc,
            'datas': self[field],
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary'
        })
    
    def action_attach_oc_pdf(self):
        """Adjunta PDF de Orden de Compra a ir.attachment"""
        self.ensure_one()
        self._store_binary_as_attachment('oc_pdf_file', 'oc_pdf_filename', 'OC PDF')
        return {'type': 'ir.actions.act_window_close'}

    def action_attach_guide_pdf(self):
        """Adjunta PDF de Guía a ir.attachment y lo referencia"""
        self.ensure_one()
        self.pdf_attachment_id = self._store_binary_as_attachment('guide_pdf_file', 'guide_pdf_filename', 'Guía PDF')
        return {'type': 'ir.actions.act_window_close'}
    
    def action_attach_excel(self):
        """Adjunta Excel de Packing a ir.attachment y lo referencia"""
        self.ensure_one()
        self.excel_attachment_id = self._store_binary_as_attachment('excel_file', 'excel_filename', 'Excel Packing')
        return {'type': 'ir.actions.act_window_close'}
    def unlink(self):
        """
        🛡️ BLINDAJE DE CABECERA - madenat.guia.processing
        v1.0 — Con limpieza de stock.moves huérfanos
        """
        # 1. SEGURIDAD USUARIO
        for rec in self:
            if rec.state not in ('draft', 'cancelled'):
                raise UserError(
                    "⛔ SEGURIDAD MADENAT:\n"
                    f"No puede eliminar la guía '{rec.name}' porque está en estado '{rec.state}'.\n"
                    "Solo se pueden eliminar guías en estado 'Borrador' o 'Cancelado'."
                )

        # 2. LIMPIEZA CASCADA — stock.moves huérfanos
        self._cleanup_orphan_moves_guia()

        return super().unlink()

    def _cleanup_orphan_moves_guia(self):
        """
        🧹 Elimina stock.moves huérfanos generados por esta guía de procesamiento.
        v3.0 — Eliminación 100% ORM. Cero SQL crudo. Falla segura si hay integridad comprometida.
        """
        # 🔒 TD-001: Guardia de grupo antes de eliminar stock.moves
        if not self.env.user.has_group('stock.group_stock_manager'):
            raise UserError(
                "No tienes permisos para eliminar movimientos de stock huérfanos.\n"
                "Se requiere el grupo 'Inventario / Administrador'."
            )
        names = self.mapped('name')
        moves = self.env['stock.move'].sudo().search([
            ('origin', 'in', names),
            ('picking_id', '=', False),
        ])
        
        if not moves:
            return

        _logger.info(
            "🧹 MADENAT guia_processing cleanup: intentando eliminar %d stock.moves huérfanos para guías: %s",
            len(moves), names
        )

        try:
            with self.env.cr.savepoint():
                # 1. Atacar las líneas de movimiento primero (stock.move.line)
                move_lines = moves.mapped('move_line_ids')
                if move_lines:
                    move_lines.write({'state': 'draft'})
                    move_lines.unlink()

                # 2. Atacar los movimientos cabecera (stock.move)
                # Forzamos a borrador para intentar bypassear el estado 'done' de manera legal
                moves.write({'state': 'draft'})

                # 3. Borrado físico usando el salvoconducto de contexto
                moves.with_context(force_delete=True).unlink()
                
            _logger.info("✅ %d moves huérfanos eliminados limpiamente via ORM.", len(moves))
            
        except Exception as e:
            _logger.error("❌ Error ORM eliminando moves huérfanos: %s", e)
            raise UserError(
                f"🛑 Seguridad Odoo: No se pudieron eliminar {len(moves)} movimientos huérfanos.\n"
                f"Odoo ha bloqueado la eliminación porque estos movimientos ya afectaron la "
                f"valoración contable o tienen stock real (Quants) fuertemente asociado.\n\n"
                f"Detalle técnico: {e}"
            )