# -*- coding: utf-8 -*-
"""
📦 RECEPCIÓN DE MADERA BRUTA DESDE PROVEEDOR

FLUJO: Proveedor externo -> Guía despacho -> Lotes brutos
COSTEO: wood_cost_usd (precio compra)
GENEALOGÍA: parent_lot_id = None (es el origen)
"""
# -*- coding: utf-8 -*-
"""
Módulo Orquestador para la Recepción de Guías de Despacho Nacionales
VERSIÓN 4.0.0 - LÓGICA INTELIGENTE DE OC + CAMPOS MARÍA VICTORIA + VALORIZACIÓN USD
"""
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from .utils_uom import (
    INCH_SQ_METERS_TO_M3,
    S2S_WIDTH_LOOKUP,
    calculate_volume_imperial_to_m3,
    calculate_volume_metric_m3,
    get_s2s_adjustment,
    r3,
    r4,
)
import base64
import io
import math
import logging
import re
from decimal import Decimal, ROUND_HALF_UP
from odoo.tools import float_round
from datetime import datetime
from .ingestion_gate import Gate0PreUpload, Gate1DocumentReconciliation
from openpyxl import load_workbook
from .reception_service import LumberReceptionService
from .width_mapping import WidthMappingTable

_logger = logging.getLogger(__name__)

# ============================================================================
# ✅ NUEVO: LÍNEA DE VERIFICACIÓN (STAGING)
# ============================================================================
class LumberReceptionLine(models.Model):
    """
    Tabla temporal para validar volúmenes antes de crear lotes definitivos.
    Hereda del Mixin para estandarizar el cálculo Real vs Nominal.
    """
    _name = 'lumber.reception.line'
    _inherit = ['madenat.lumber.ingest.line.mixin']
    _description = 'Línea de Verificación de Recepción'
  
    # ==================== GATE 3: AUDITORÍA CRIPTOGRÁFICA ====================
    audit_snapshot = fields.Text(
        string='Snapshot JSON', 
        readonly=True, 
        copy=False,
        help="Fotografía inmutable de los datos comerciales al momento de confirmar la recepción."
    )
    audit_hash = fields.Char(
        string='Firma SHA-256', 
        readonly=True, 
        copy=False,
        help="Huella digital criptográfica del snapshot para garantizar que nadie alteró los volúmenes por base de datos."
    )

    # ==========================================================================
    # 🔗 RELACIONES Y CAMPOS DE IDENTIDAD (Evitan el Owl/KeyError)
    # ==========================================================================
    reception_id = fields.Many2one('lumber.reception', string='Recepción', ondelete='cascade', required=True)
    lot_name = fields.Char(string="N° Lote", required=True)
    subproduct_id = fields.Many2one('madenat.subproducto', string="Subproducto/Grado")
    product_name_clean = fields.Char(
        string="Nombre Producto",
        compute='_compute_product_name_clean',
        store=True,
        readonly=True,
    )

    @api.depends('product_id', 'product_id.name')
    def _compute_product_name_clean(self):
        for rec in self:
            rec.product_name_clean = rec.product_id.name or ''
    excel_product_name = fields.Char(
        string="Producto (Excel)", 
        help="Nombre exacto capturado desde el archivo original"
    )
    wood_species_id = fields.Char("Especie")
    quality = fields.Selection([
        ('col_a', 'Col A'), ('col_b', 'Col B'), 
        ('industrial', 'Industrial'), ('premium', 'Premium')
    ], string='Calidad')
    export_calculation_rule = fields.Selection([
        ('metric', 'Métrico (Físico)'),
        ('f1550', 'Factor 1550 (Metros)'),
        ('f5085', 'Factor 5085 (Pies)')
    ], string="Regla Cálculo", default='metric')

    # ==========================================================================
    # DIMENSIONES NOMINALES Y VISUALES (COMERCIALES - OC)
    # ==========================================================================
    thickness_nominal = fields.Float(
        "Espesor Nom. (mm)",
        help="Espesor teórico/comercial pactado en la OC (ej: 45mm)"
    )
    width_nominal = fields.Float("Ancho Nom. (mm)")
    length_nominal = fields.Float("Largo Nom. (m)")

    thickness_visual = fields.Char(
        "Espesor (Nom)", 
        compute="_compute_visual_defaults", 
        store=True, 
        readonly=False, 
        help="Ej: 6/4. Se sugiere desde el Análisis Comercial pero es 100% editable."
    )

    width_visual = fields.Char(
        string="Ancho (Nom)", 
        compute="_compute_visual_defaults", 
        store=True, 
        readonly=False, 
        help="Ej: 5 5/8. Se sugiere desde el Análisis Comercial pero es 100% editable."
    )

    thickness_nominal_frac = fields.Char(
        string="Esp. Nom. (Pulg)", 
        compute="_compute_visual_defaults",  # 🚀 EL ENLACE MÁGICO AL MOTOR
        store=True,
        readonly=False, # Vital: Permite que el Wizard lo siga sobreescribiendo
        help="Valor fraccionario visual (Ej. 1 9/16)"
    )
    
    width_nominal_frac = fields.Char(
        string="Ancho Nom. (Pulg)", 
        compute="_compute_visual_defaults",  # 🚀 EL ENLACE MÁGICO AL MOTOR
        store=True,
        readonly=False, # Vital: Permite que el Wizard lo siga sobreescribiendo
        help="Valor fraccionario visual (Ej. 3 5/8)"
    )

    # ==========================================================================
    # 🛡️ GUARDIÁN DE IDENTIDAD: EAN-13 (CONSOLIDADO)
    # ==========================================================================
    
    def _sanitize_lot_name(self, val):
        if not val: return False
        val = str(val).strip().split('.')[0]
        return val.zfill(13) if val.isdigit() else val

    @api.onchange('lot_name')
    def _onchange_lot_name(self):
        if self.lot_name:
            self.lot_name = self._sanitize_lot_name(self.lot_name)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('lot_name'): # Normalizar lot_name
                vals['lot_name'] = self._sanitize_lot_name(vals['lot_name'])

            if any(k in vals for k in ('thickness', 'width', 'length')):
                thickness = vals.get('thickness', 0.0)
                width = vals.get('width', 0.0)
                length = vals.get('length', 0.0)
                if thickness or width or length:
                    self._validate_lot_dimensions(thickness, width, length)

            if not vals.get('excel_product_name') and vals.get('product_id'):
                product = self.env['product.product'].browse(vals['product_id'])
                vals['excel_product_name'] = product.name
        return super().create(vals_list)

    def write(self, vals):
        if 'lot_name' in vals:
            vals['lot_name'] = self._sanitize_lot_name(vals['lot_name'])

        if any(k in vals for k in ('thickness', 'width', 'length')):
            for line in self:
                thickness = vals.get('thickness', line.thickness)
                width = vals.get('width', line.width)
                length = vals.get('length', line.length)
                if thickness or width or length:
                    line._validate_lot_dimensions(thickness, width, length)

        return super().write(vals)

   
   # === DIMENSIONES REALES (FÍSICAS - BODEGA) ===
    thickness = fields.Float("Espesor (mm)")
    width = fields.Float("Ancho (mm)")
    length = fields.Float("Largo (m)")

    # === CAMPOS PARA INGRESO DE LARGO CON UNIDAD ===
    lengthuom = fields.Selection(
        selection=[('m', 'Metros'), ('mm', 'Milímetros'), ('ft', 'Pies')],
        string='Unidad Largo',
        default='m',
        required=True,
        help='Unidad en la que se ingresa el largo. Se convierte automáticamente a metros.'
    )

    length_input_raw = fields.Float(
        string='Largo ingreso',
        digits=(10, 4),
        help='Largo en la unidad seleccionada. Se convierte a metros al guardar.'
    )

    lengthm = fields.Float(
        string='Largo convertido m',
        compute='_compute_lengthm',
        store=True,
        digits=(10, 4),
        help='Largo normalizado a metros para cálculos y validaciones.'
    )

    # === DIMENSIONES NOMINALES (COMERCIALES - OC) ===
    # (Los campos nominales ya se definen arriba)

    # 1. Volumen Físico REAL (Solo Informativo/Staging)
    # Este calcula mm * mm * m. Sirve para ver "cuánto perdemos/ganamos" vs el nominal.
    vol_physical_real_m3 = fields.Float(
        string="Vol. Físico Real (Staging)",
        digits=(16, 3),
        compute="_compute_volume_physical_real",
        store=True,
        help="Cálculo milimétrico puro. NO afecta inventario contable."
    )

    # 2. Volumen Compra/Inventario (LA VERDAD DE BODEGA)
    # Este es el nominal. Y según tu regla, ESTE es el que manda en bodega.
    vol_physical_m3 = fields.Float(
        string="Vol. Físico (m³)",
        compute="_compute_vol_physical_strict",
        store=True,
        readonly=True,
        digits=(16, 3)
    )

    vol_purchase_m3 = fields.Float(
        string="Vol. Compra (m³)",
        compute="_compute_volume_purchase",
        inverse="_inverse_vol_purchase_m3",
        store=True,
        readonly=False,  # Este sí se puede editar
        digits=(16, 3)
    )

    # === DATOS FINANCIEROS ===
    estimated_cost_usd = fields.Float("Costo Est. (USD)")

    # NUEVO: El costo prorrateado que usaremos para el inventario real
    cost_clp_unit = fields.Float(
        string="Costo Unitario (CLP)",
        compute="_compute_line_cost",
        store=True,
        help="Costo neto de la madera prorrateado por m3 comercial"
    )

    # 🚀 Agrega este campo para que el widget monetary pueda funcionar y sumar
    currency_id = fields.Many2one(
        'res.currency',
        string='Moneda',
        default=lambda self: self.env.company.currency_id.id,
        readonly=True
    )

  # Este campo lo mantenemos con su compute específico abajo

    @api.depends('length_input_raw', 'lengthuom', 'length')
    def _compute_lengthm(self):
        for rec in self:
            raw = rec.length_input_raw if rec.length_input_raw not in (False, None) else 0.0
            if raw:
                if rec.lengthuom == 'ft':
                    rec.lengthm = round(raw * 0.3048, 6)
                elif rec.lengthuom == 'mm':
                    rec.lengthm = round(raw * 0.001, 6)
                else:
                    rec.lengthm = raw
            else:
                rec.lengthm = rec.length or 0.0

    @api.onchange('length_input_raw', 'lengthuom')
    def _onchange_length_input_to_length(self):
        for rec in self:
            if rec.length_input_raw:
                if rec.lengthuom == 'ft':
                    rec.length = round(rec.length_input_raw * 0.3048, 6)
                elif rec.lengthuom == 'mm':
                    rec.length = round(rec.length_input_raw * 0.001, 6)
                else:
                    rec.length = rec.length_input_raw
            else:
                rec.length = 0.0

    board_feet = fields.Float(
        string="Pie Tabla (PT)",
        digits=(12, 2),
        compute="_compute_export_values",  # 🔗 Conectado
        store=True
    )

    # 4. RESULTADOS (OUTPUTS CALCULADOS POR LA REGLA DE ORO)
    vol_mbf = fields.Float(
        string="Volumen MBF",
        digits=(16, 3),
        compute="_compute_export_values",  # 🔗 Conectado
        store=True
    )




    # ==========================================================================
    # 🌎 TERCERA REALIDAD: VOLUMEN DE EMBARQUE (Exportación)
    # ==========================================================================
    
    # Redefinimos el campo del Mixin para que deje de ser "Físico" y sea "Exportación"
    vol_shipment_m3 = fields.Float(
        string="Vol. Exp (m³)", 
        digits=(12, 3),  # ✅ Precisión 3 decimales
        compute="_compute_export_values", # 🔗 Conectado
        store=True,
        help="Calculado con Regla de Oro (Factor 5085/1550 + 1/8)"
    )

    # ==========================================================================
    # 🏭 1. CÁLCULO FÍSICO (HISTÓRICO / TRAZABILIDAD)
    # ==========================================================================
    @api.depends('thickness', 'width', 'length', 'pieces')
    def _compute_vol_physical_strict(self):
        """
        Calcula el volumen de la Recepción (Bodega).
        BLINDAJE: Solo mira dimensiones físicas. Si cambias el nominal, esto NO se mueve.
        Protege la integridad de lo importado.
        """
        for line in self:
            # Fórmula Física: (mm * mm * m * pzas) / 1.000.000
            vol = calculate_volume_metric_m3(
                line.thickness, line.width, line.length, line.pieces
            )
            if not self._is_valid_volume(vol):
                _logger.warning(
                    "MADENAT: vol_physical_m3 inválido para línea %s: %s",
                    line.id or 'new', vol
                )
                vol = 0.0
            line.vol_physical_m3 = vol

    
   
    @api.onchange('subproduct_id')
    def _onchange_subproduct_id(self):
        """
        🎯 Regla de Oro: Auto-llenado de medidas nominales desde el producto.
        BLINDAJE: El subproducto NO tiene autoridad para pisar un nominal que
        el usuario o el Wizard ya hayan fijado. Solo actúa si la celda está en cero.
        """
    
        if self.subproduct_id:
            sub_t = self.subproduct_id.thickness_nominal or 0.0
            sub_w = self.subproduct_id.width_nominal or 0.0
            sub_l = self.subproduct_id.length_nominal or 0.0

            # Solo inyectar si nuestro nominal actual está vacío/cero
            if not self.thickness_nominal and sub_t > 0:
                self.thickness_nominal = sub_t
            if not self.length_nominal and sub_l > 0:
                self.length_nominal = sub_l
            if not self.width_nominal and sub_w > 0:
                self.width_nominal = sub_w

    @api.onchange('length_input_raw', 'lengthuom')
    def _onchange_length_input(self):
        """
        Convierte el largo ingresado en la unidad seleccionada a metros.
        Factores de conversión:
        - mm → m: × 0.001
        - ft → m: × 0.3048
        - m → m: × 1.0
        """
        factors = {'m': 1.0, 'mm': 0.001, 'ft': 0.3048}
        uom = self.lengthuom or 'm'
        raw = self.length_input_raw or 0.0
        self.length = round(raw * factors[uom], 4)

    @api.model
    def _get_subproduct_domain(self):
        """
        🛡️ Blindaje dinámico: Filtra productos según el perfil de la cabecera.
        """
        parent_profile = self.reception_id.ingestion_profile
        
        # Mapeo de Perfil -> Categoría de Producto (Ajusta los nombres a tus categorías reales)
        category_map = {
            'f5085': 'Blanks',
            'f1550': 'S2S / Ripped',
            'metric': 'Madera Bruta' # Asumiendo que 'metric' es para madera bruta
        }
        
        target_cat = category_map.get(parent_profile)
        if target_cat:
            return [('categ_id.name', 'ilike', target_cat)]
        return []
    # ==========================================================================
    # 🧠 LÓGICA DE NEGOCIO DESACOPLADA
    # ==========================================================================
    @api.depends('thickness', 'width', 'length', 'pieces', 'reception_id.ingestion_profile')
    def _compute_volume_physical_real(self):
        """
        📐 Cálculo de Volumen Físico Real (Audit Excel)
        MEJORA: Implementa un 'Switch' de seguridad por perfil.
        - Blanks Clear (f5085): Convierte Pies a Metros.
        - S2S/Métrico: Mantiene Metros originales.
        """
        for line in self:
            # 1. Dimensiones base extraídas por el Parser
            t = line.thickness or 0.0
            w = line.width or 0.0
            l_raw = line.length or 0.0
            p = line.pieces or 0

            # 2. 🛡️ FILTRO DE SEGURIDAD POR PERFIL
            # Sustituimos el "if l >= 7.0" por una validación de perfil.
            # Esto evita que una madera nacional de 8 metros sea convertida por error.
            if line.reception_id.ingestion_profile == 'f5085':
                # Es Blanks: Los pies del Excel pasan a metros
                l_final = l_raw * 0.3048
            else:
                # Es S2S o Métrico: El largo ya está en metros
                l_final = l_raw

            # 3. Cálculo volumétrico métrico puro
            # Usamos tu helper de confianza calculate_volume_metric_m3
            raw_vol = calculate_volume_metric_m3(t, w, l_final, p)
            
            # 4. Blindaje de 3 decimales (Regla de Oro Madenat)
            line.vol_physical_real_m3 = float_round(
                raw_vol, 
                precision_digits=3, 
                rounding_method='HALF-UP'
            )
    # B. Sincronización UNIDIRECCIONAL (Nominal -> Visual)
    # FIX v5.5: Agregamos dependencias físicas ('thickness', 'width') para que el ORM 
    # autocalcule el espejo justo después de crear la línea en el staging.
    @api.depends('thickness_nominal', 'width_nominal', 'thickness', 'width', 'reception_id.ingestion_profile')
    def _compute_visual_defaults(self):
        """
        🎨 TRADUCTOR DE IDENTIDAD VISUAL (V6.1 - CONSOLIDADO)
        Separa la Verdad Física (Pestaña 1) de la Sugerencia Comercial (Pestaña 2).
        """
        for line in self:
            profile = line.reception_id.ingestion_profile
            
            # 🛡️ MUNDO 1: S2S / ROUGH (INTACTO - NO SE TOCA)
            if profile != 'f5085':
                t_nom = line.thickness_nominal or line.thickness
                if 37 <= t_nom <= 46: 
                    res_t = "6/4"
                elif 22 <= t_nom <= 29: 
                    res_t = "4/4"
                else:
                    quarters = int(round((t_nom / 25.4) * 4))
                    res_t = f"{quarters}/4" if quarters > 0 else ""
                
                line.thickness_visual = res_t
                line.width_visual = self._get_trader_width_text(line.width_nominal or line.width)
                
                # Sincronizamos los nominales por integridad (aunque el XML los oculte en S2S)
                line.thickness_nominal_frac = line.thickness_visual
                line.width_nominal_frac = line.width_visual
            
            # 🇺🇸 MUNDO 2: BLANKS CLEAR (EL DESACOPLE)
            else:
                # 1. ESPEJO FÍSICO (Pestaña 1): Siempre usa la medida real del Excel
                t_phys_in = line.thickness if line.thickness < 10.0 else line.thickness / 25.4
                w_phys_in = line.width if line.width < 10.0 else line.width / 25.4
                
                line.thickness_visual = self._get_fraction_text(t_phys_in)
                line.width_visual = self._get_fraction_text(w_phys_in)
                
                # 2. SUGERENCIA COMERCIAL (Pestaña 2): Sugiere físico, pero escucha al Nominal
                # Si el nominal es 0 (antes del wizard), t_val será igual al físico.
                t_val = line.thickness_nominal if line.thickness_nominal > 0 else line.thickness
                w_val = line.width_nominal if line.width_nominal > 0 else line.width
                
                t_nom_in = t_val if t_val < 10.0 else t_val / 25.4
                w_nom_in = w_val if w_val < 10.0 else w_val / 25.4
                
                # AQUÍ ESTABA EL FALLO: Ahora asignamos los campos de la Pestaña 2
                line.thickness_nominal_frac = self._get_fraction_text(t_nom_in)
                line.width_nominal_frac = self._get_fraction_text(w_nom_in)
   

    def action_suggest_nominal_defaults(self):
        """
        Botón explícito para sugerir nominales desde físico.
        Al ser un método accionado por el usuario, no rompe la caché del navegador.
        """
        for line in self:
            if not line.thickness_nominal and line.thickness > 0:
                line.thickness_nominal = line.thickness
            if not line.width_nominal and line.width > 0:
                line.width_nominal = line.width  
            if not line.length_nominal and line.length > 0:
                line.length_nominal = line.length
                
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': '✅ Nominales sugeridos',
                'message': 'Se copiaron los valores físicos a nominales donde estaban vacíos.',
                'type': 'success'
            }
        }
   

    # -------------------------------------------------------------------------
    # 🧩 TRADUCTORES (El cerebro que evita el RPC_ERROR)
    # -------------------------------------------------------------------------
    def _parse_smart_dimension(self, text):
        if not text: return 0.0
        try:
            raw = str(text).replace(',', '.').strip()
            if ' ' in raw and '/' in raw:
                p = raw.split()
                return float(p[0]) + (float(p[1].split('/')[0]) / float(p[1].split('/')[1]))
            elif '/' in raw:
                return float(raw.split('/')[0]) / float(raw.split('/')[1])
            val = float(raw)
            return val / 25.4 if val > 10.0 else val
        except: return 0.0

    def _get_fraction_text(self, value):
            """
            📐 CONVERSOR PROFESIONAL DE PULGADAS DECIMALES A FRACCIÓN
            Soporta hasta 1/16 para Blanks y simplifica para S2S (1/2, 1/4, etc.)
            """
            if not value or value <= 0: 
                return ""
                
            whole = int(value)
            frac = value - whole
            
            # 🎯 Cambiamos a base 16 para capturar la precisión de los Blanks (9/16)
            sixteenths = int(round(frac * 16))
            
            # Casos borde: redondeo al entero superior o sin fracción
            if sixteenths == 16: 
                return str(whole + 1)
            if sixteenths == 0: 
                return str(whole) if whole > 0 else ""
                
            # --- Lógica de Simplificación Automática ---
            # (Ejemplo: 8/16 -> 1/2 | 4/16 -> 1/4 | 9/16 -> 9/16)
            from math import gcd
            common = gcd(sixteenths, 16)
            num = sixteenths // common
            den = 16 // common
            
            f_text = f"{num}/{den}"
            
            return f"{whole} {f_text}".strip() if whole > 0 else f_text
    # ==========================================================================
    # 🧮 LA REGLA DE ORO (MODIFICACIÓN QUIRÚRGICA)
    # Reemplaza SOLO el método _compute_export_golden_rule en lumber_reception.py
    # ==========================================================================
       # ==========================================================================
    # 🚀 MOTOR DE CÁLCULO DE EXPORTACIÓN (NUEVO Y UNIFICADO)
    # Reemplaza a las funciones antiguas. Calcula Vol. Exp + MBF + PT.
    # ==========================================================================
    
    
    @api.depends(
        'thickness_visual', 'width_visual', 'length', 'pieces',
        'export_calculation_rule', 'thickness', 'width',
        'thickness_nominal', 'width_nominal', 'length_nominal',
        'length_input_raw'
    )
    def _compute_export_values(self):
        """
        🚀 MOTOR DE CÁLCULO DE EXPORTACIÓN (CONSOLIDADO Y BLINDADO)
        Aplica Reglas de Oro: Factor 1550.003, Factor 5085.312 y Recargo S2S (+0.125").
        """
        # --- CONSTANTES DE INGENIERÍA ---
        FACTOR_5085 = 5085.312
        METRO_A_PIE = 0.3048

        # 1. PARSEO AISLADO (Funcionalidad intacta)
        def parse_to_inches(val_visual, val_real, current_line):
            if val_visual:
                try:
                    # Intento de parseo inteligente
                    if hasattr(current_line, '_parse_smart_dimension'):
                        return current_line._parse_smart_dimension(val_visual)
                    
                    # Respaldo: Procesamiento de fracciones (Ej. "1 9/16" o "6/4")
                    raw = str(val_visual).replace(',', '.').strip()
                    if '/' in raw:
                        if ' ' in raw:
                            parts = raw.split()
                            return float(parts[0]) + (float(parts[1].split('/')[0]) / float(parts[1].split('/')[1]))
                        return float(raw.split('/')[0]) / float(raw.split('/')[1])
                    
                    # Si es mayor a 10, asumimos milímetros y convertimos. Si no, son pulgadas.
                    val_float = float(raw)
                    return val_float / 25.4 if val_float > 10.0 else val_float

                except Exception as e:
                    _logger.warning(f"MADENAT: Error parseando dimensión visual '{val_visual}': {e}")
                    pass # Falla silenciosa y segura hacia el valor real
            
            # Fallback Seguro: Usar el valor real en mm pasado como parámetro
            return (val_real / 25.4) if val_real else 0.0

        # 2. PROCESAMIENTO DEL RECORDSET
        for line in self:
            # Inicialización por defecto para evitar errores en UI
            vol_exp = 0.0
            val_mbf = 0.0
            
            try:
                # 🎯 REGLA CANÓNICA: Nominal manda, Físico respalda.
                t_mm = line.thickness_nominal if line.thickness_nominal > 0 else line.thickness
                w_mm = line.width_nominal if line.width_nominal > 0 else line.width
                l_m = line.length_nominal if line.length_nominal > 0 else line.length
                qty = line.pieces or 0.0

                # Obtención segura en Pulgadas, inyectando los mm CANÓNICOS (t_mm, w_mm) como val_real
                t_in = parse_to_inches(line.thickness_visual, t_mm, line)
                w_in = parse_to_inches(line.width_visual, w_mm, line)
                
                # Autodetección SEGURA
                rule = line.export_calculation_rule
                if not rule:
                    rule = 'f5085' if t_in > 0 else 'metric'

                # Ajuste S2S usando el ancho canónico
                _s2s = float(get_s2s_adjustment(self.env, w_mm))
                w_calc = w_in + _s2s if w_in > 0 else 0.0
                l_feet = l_m / METRO_A_PIE if l_m else 0.0

                # 3. MATRIZ DE INGENIERÍA
                if t_in > 0 and w_in > 0 and l_m > 0 and qty > 0:
                    if rule == 'f5085':
                        # BLANK CLEAR: Fórmula con largo en PIES directos del Excel
                        # t_in = (thickness_mm / 25.4) - 0.0625  # -1/16" deducción cara
                        # w_in = (width_mm / 25.4) + 0.125       # +1/8" S2S
                        # l_ft = length_input_raw (pies, sin convertir)
                        t_in_bc = (t_mm / 25.4) - 0.0625
                        w_in_bc = (w_mm / 25.4) + 0.125
                        l_ft = line.length_input_raw if line.length_input_raw else (l_m / METRO_A_PIE)
                        vol_exp = (t_in_bc * w_in_bc * l_ft * qty) / FACTOR_5085
                        val_mbf = (t_in_bc * w_in_bc * l_ft * qty) / 12000.0

                    elif rule == 'f1550':
                        vol_exp = float(
                            Decimal(str(t_in))
                            * Decimal(str(w_calc))
                            * Decimal(str(l_m))
                            * Decimal(str(qty))
                            / INCH_SQ_METERS_TO_M3
                        )
                        val_mbf = (t_in * w_in * l_feet * qty) / 12000.0
                        
                    elif rule == 'metric':
                        # Regla Métrica Pura: Usa los valores canónicos en MM
                        vol_exp = calculate_volume_metric_m3(t_mm, w_mm, l_m, qty)
                        val_mbf = 0.0

            except Exception as e:
                _logger.error(f"MADENAT: Colapso en cálculo de exportación para línea {line.id}: {e}")

            # 4. ASIGNACIÓN FINAL (ORM Standard)
            if not self._is_valid_volume(vol_exp):
                _logger.warning(
                    "MADENAT: vol_shipment_m3 inválido para línea %s: %s",
                    line.id or 'new', vol_exp
                )
                vol_exp = 0.0

            line.vol_shipment_m3 = float_round(vol_exp, precision_digits=3, rounding_method='HALF-UP')
            line.vol_mbf = float_round(val_mbf, precision_digits=3, rounding_method='HALF-UP')
            line.board_feet = float_round(val_mbf * 1000.0, precision_digits=2, rounding_method='HALF-UP')
  
    # ==========================================================================
    # 🧠 CEREBRO DE DATOS (TABLA DE VERDAD EXCEL - ANCHOS ROUGH vs S2S)
    # ==========================================================================
    
    def _get_excel_mapping(self, mm_value):
        """
        Retorna el valor DECIMAL S2S exacto según tu Excel 'ANCHOS COMPRA COL ROUGH A S2S'.
        """
        if not mm_value:
            return 0.0
        # Fallback para valores grandes (>160mm) que no están en tabla
        if mm_value >= 160.0:
            return mm_value
        return WidthMappingTable.get_value(mm_value, format_type='decimal') or 0.0

    def _get_trader_width_text(self, value_mm):
        """
        Lógica Trader ANCHO (TABLA DE VERDAD EXCEL):
        Mapeo exacto basado en 'ANCHOS COMPRA COL ROUGH A S2S.xlsx'.
        """
        if not value_mm: return ""

        mapped_text = WidthMappingTable.get_value(value_mm, format_type='text')
        if mapped_text:
            return mapped_text

        # 3. FALLBACK (Para medidas > 160mm o fuera de tabla)
        # Según el Excel, mayores a 160mm se muestran como enteros (170 -> 170)
        val_in = value_mm / 25.4
        if value_mm >= 160.0:
            return str(int(round(value_mm))) 
        else:
            return self._get_fraction_text(val_in)


    # ==========================================================================
    # 💲 2. CÁLCULO COMERCIAL (NOMINAL / FINANCIERO)
    # ==========================================================================
    # FIX V5.8: Agregadas dependencias 'reception_id.ingestion_profile' y 'vol_shipment_m3'
    @api.depends(
        'thickness_nominal', 'width_nominal', 'length_nominal',
        'thickness', 'width', 'length', 'pieces',
        'reception_id.ingestion_profile', 'vol_shipment_m3'
    )
    def _compute_volume_purchase(self):
        """
        📐 MOTOR DE PRECISIÓN V6.6 (Redondeo Suizo)
        """
        for line in self:
            profile = line.reception_id.ingestion_profile
            
            if profile != 'f5085':
                # S2S: Matemática métrica pura redondeada a 3 decimales
                t = line.thickness_nominal or line.thickness
                w = line.width_nominal or line.width
                l = line.length_nominal or line.length
                line.vol_purchase_m3 = round((t * w * l * line.pieces) / 1000000.0, 3)
                # Usar r3() para consistencia con Regla de Oro
                line.vol_purchase_m3 = r3((t * w * l * line.pieces) / 1000000.0)
            
            else:
                # BLANKS: Escalado Proporcional con redondeo a 3 decimales
                if line.thickness_nominal == 0 and line.width_nominal == 0:
                    # Si no hay Wizard, es exactamente el del Excel
                    line.vol_purchase_m3 = line.vol_shipment_m3
                else:
                    t_nom = line.thickness_nominal if line.thickness_nominal > 0 else line.thickness
                    w_nom = line.width_nominal if line.width_nominal > 0 else line.width
                    
                    factor_t = t_nom / line.thickness if line.thickness > 0 else 1.0
                    factor_w = w_nom / line.width if line.width > 0 else 1.0
                    
                    # 🎯 REDONDEO CRÍTICO: Elimina la milésima de discrepancia
                    line.vol_purchase_m3 = round(line.vol_shipment_m3 * factor_t * factor_w, 3)
                    line.vol_purchase_m3 = r3(line.vol_shipment_m3 * factor_t * factor_w)

            if not self._is_valid_volume(line.vol_purchase_m3):
                _logger.warning(
                    "MADENAT: vol_purchase_m3 inválido para línea %s: %s",
                    line.id or 'new', line.vol_purchase_m3
                )
                line.vol_purchase_m3 = 0.0

    def _inverse_vol_purchase_m3(self):
        """
        Permite al operador sobreescribir vol_purchase_m3 manualmente
        durante el staging. El valor se persiste (store=True) y no
        es sobreescrito por el compute si el usuario lo editó.
        Arquitectura: STAGING EDITABLE — operador es responsable.
        Ver: WIKI/02_TECNICO/arquitectura_ingesta_recepciones.md
        """
        pass  # store=True persiste el valor automáticamente

    @api.depends('reception_id.total_amount_clp', 'reception_id.commercial_volume_m3', 'vol_purchase_m3')
    def _compute_line_cost(self):
        """
        Cálculo Quirúrgico de Prorrateo usando campos de María Victoria.
        """
        for line in self:
            # total_amount_clp es el campo de la línea 689
            if line.reception_id.commercial_volume_m3 > 0:
                price_per_m3 = line.reception_id.total_amount_clp / line.reception_id.commercial_volume_m3
                line.cost_clp_unit = price_per_m3 * line.vol_purchase_m3
            else:
                line.cost_clp_unit = 0.0
    def unlink(self):
            """
            🛡️ BLINDAJE CON SALVOCONDUCTO TÉCNICO:
            Permite borrar líneas viejas al re-procesar, incluso si la guía se quedó pegada.
            """
            # 1. EL SALVOCONDUCTO (La clave para desbloquear su situación)
            # Si el código viene con la orden 'force_delete', ignoramos el estado y borramos.
            if self.env.context.get('force_delete'):
                return super(LumberReceptionLine, self).unlink()

            # 2. SEGURIDAD MANUAL (Sigue protegiendo contra errores humanos)
            for line in self:
                if line.reception_id and line.reception_id.state not in ('draft', 'error', 'cancel'):
                    raise UserError(
                        f"⛔ SEGURIDAD MADENAT:\n"
                        f"No puede eliminar la etiqueta '{line.lot_name}' manualmente.\n"
                        f"La Guía '{line.reception_id.name}' está en estado '{line.reception_id.state}'.\n"
                        "Use el botón 'Resetear a Borrador' si necesita corregir."
                    )
            
            return super(LumberReceptionLine, self).unlink()
    
class LumberReception(models.Model):
    _name = 'lumber.reception'
    _description = '''📦 RECEPCIÓN DE MADERA BRUTA DESDE PROVEEDOR
        FLUJO: Proveedor externo → Guía despacho → Lotes brutos
        COSTEO: wood_cost_usd (precio compra)
        GENEALOGÍA: parent_lot_id = None (es el origen)'''
    
    _inherit = ['mail.thread', 'mail.activity.mixin', 'madenat.lumber.ingest.mixin']
    _order = 'reception_date desc'

    audit_snapshot = fields.Text('Snapshot JSON', readonly=True)
    audit_hash = fields.Char('Firma SHA-256', readonly=True)
    # ==================== CAMPOS PRINCIPALES ====================

     # ✅ NUEVO: Relación con Staging
    reception_line_ids = fields.One2many(
        'lumber.reception.line', 
        'reception_id', 
        string='Análisis de Validación'
    )
   
    
    state = fields.Selection([
        ('draft', '📝 Borrador'),
        ('processing', '⚙️ Procesando'),
        ('verified', '👀 Verificado'),  # Cambiado de 'processed' a 'verified'
        ('done', '✅ Recibido'),
        ('cancel', '🚫 Cancelado'),
        ('error', 'Error'),             # Mantenemos por seguridad
        ('pending_link', '⏳ Pendiente OC') # Mantenemos por seguridad
    ], string='Estado', default='draft', tracking=True, index=True)
    

    ingestion_profile = fields.Selection([
            ('f5085', '🪵 Blanks Clear (Factor 5085 - Pies)'),
            ('f1550', '📐 S2S / Rough (Factor 1550 - Métrico)'),
            ('metric', '📏 Madera Bruta (Milimétrico Directo)'),
        
        ], string='Perfil de Ingesta', required=True, default='f5085', tracking=True, 
        help="Define explícitamente la regla matemática que aplicará el parser al Excel.")


    # ================ FLAGS DE ESTADO (COPIA DE GUÍAS) ================
    can_process_reception = fields.Boolean(
        compute="_compute_can_process_reception", 
        string="Puede procesar"
    )
    can_reopen_reception = fields.Boolean(
        compute="_compute_can_reopen_reception"
    )
    can_cancel_reception = fields.Boolean(
        compute="_compute_can_cancel_reception"
    )
    # ================================================================

    name = fields.Char('Número Guía', tracking=True, copy=False)

    manual_entry = fields.Boolean(
        string="Entrada Manual",
        help="Activado si falla el reconocimiento automático del PDF",
        default=False
    )
    # ==================== CAMPOS DE AUDITORÍA ====================
    omitted_count = fields.Integer(
        string='Omisiones en esta recepción',
        compute='_compute_omitted_count',
        store=True
    )

    audit_log_ids = fields.One2many(
        'madenat.audit.log', 'reception_id',
        string='Logs de Auditoría'
    )

    # ==================== CAMPOS MARÍA VICTORIA (CRÍTICOS) ====================
    guia_numero = fields.Char(
        'Nº Guía Despacho',
        tracking=True,
        help="Número oficial de la guía de despacho del proveedor"
    )
    
    guia_fecha = fields.Date(
        'Fecha Guía',
        tracking=True,
        help="Fecha de emisión de la guía de despacho"
    )
      # --- NUEVO CAMPO DE ENLACE ---
    origin_processing_id = fields.Many2one(
        'madenat.guia.processing',
        string='Origen Procesamiento',
        readonly=True,
        copy=False,
        help="Referencia a la guía de procesamiento que originó esta recepción."
    )
    # ==================== RELACIONES ====================
    reception_date = fields.Datetime('Fecha Recepción', default=fields.Datetime.now, tracking=True)
    
    supplier_id = fields.Many2one(
        'res.partner', 
        'Proveedor', 
        domain=['|', ('is_company', '=', True), ('parent_id', '!=', False)],
        tracking=True
    )
    
    purchase_id = fields.Many2one(
        'purchase.order', 
        'Orden de Compra',
        domain="[('partner_id', '=', supplier_id)]",
        index=True,
        tracking=True
    )
    
    # ✅ NUEVO: Campo para entrada manual editable
    manual_po_name = fields.Char(
        string='N° Orden (Manual)',
        tracking=True,
        help="Si no existe OC en Odoo, ingrese el número manualmente aquí."
    )
    
    # MODIFICADO: Este campo ahora debe calcularse para mostrar la OC Real o la Manual
    purchase_order = fields.Char(
        string='Orden Compra (Visual)', 
        compute='_compute_purchase_order_display',
        store=True,
        readonly=True
    )


    # ===================== FLAGS PARA LAS VALIDACIONES =====================
    @api.depends('state', 'files_ready')
    def _compute_can_process_reception(self):
        for rec in self:
            rec.can_process_reception = rec.state == 'draft' and rec.files_ready

    @api.depends('state', 'lot_ids')
    def _compute_can_reopen_reception(self):
        for rec in self:
            rec.can_reopen_reception = rec.state in ('validated', 'done') and not rec.lot_ids

    @api.depends('state')
    def _compute_can_cancel_reception(self):
        for rec in self:
            rec.can_cancel_reception = rec.state in ('draft', 'processing')

    # ==========================================================


    @api.depends('purchase_id', 'manual_po_name')
    def _compute_purchase_order_display(self):
        """Prioriza la OC oficial de Odoo, si no, usa la manual."""
        for rec in self:
            if rec.purchase_id:
                rec.purchase_order = rec.purchase_id.name
            else:
                rec.purchase_order = rec.manual_po_name or 'SIN ORDEN'
   # ==================== VOLÚMENES DUALES (m³ + MBF) ====================
    commercial_volume_m3 = fields.Float(
        digits=(16, 3), 
        string='Volumen Comercial Guía (m³)',
        required=False,  # <-- CAMBIO AQUÍ: Ahora es opcional al crear
        default=0.0,     # <-- Agregamos un valor por defecto seguro
        tracking=True,   # (Opcional) Bueno para auditoría
        help="Volumen DECLARADO en la guía de despacho (ingresar manualmente del PDF). "
            "Este es el volumen facturado/comercial, NO el medido físicamente."
    )

    commercial_volume_mbf = fields.Float(
        digits=(16, 3), # ⬅️ CORRECCIÓN: Estrictamente 3 decimales (Punto 1.5)
        string='Volumen Comercial MBF',
        compute='_compute_commercial_mbf_from_m3',
        store=True,
        readonly=False, # ⬅️ CORRECCIÓN: El Origen permite escritura manual (Punto 1.4)
        help="Volumen comercial convertido a MBF. Calculado automáticamente pero modificable manualmente."
    )

    physical_volume_m3 = fields.Float(
        digits=(16, 3), 
        string='Volumen Físico m³',
        readonly=False,
        help="Volumen según packing list recibido (medición física)"
    )

    physical_volume_mbf = fields.Float(
        digits=(16, 3), 
        string='Volumen Físico MBF',
        compute='_compute_totals',
        store=True,
        readonly=False,
        help="Volumen físico en MBF desde lotes"
    )

    volume_variance_percent = fields.Float(
        digits=(16, 2), 
        string='Diferencia %',
        compute='_compute_volume_variance',
        store=True
    )

    volume_variance_m3 = fields.Float(
        digits=(16, 3), 
        string='Diferencia (m³)',
        compute='_compute_volume_variance',
        store=True,
        help="Diferencia absoluta: Físico - Comercial"
    )

    
    tolerance_status = fields.Selection([
        ('ok', '✅ Dentro Tolerancia'),
        ('warning', '⚠️ Atención Requerida'), 
        ('critical', '🚨 Revisión Crítica')
    ], compute='_compute_tolerance_status', store=True)
    
    # ==================== VALORIZACIÓN (USD + CLP) ====================
    currency_id = fields.Many2one('res.currency', 'Moneda', default=lambda self: self.env.ref('base.CLP').id)
    usd_currency_id = fields.Many2one('res.currency', 'Moneda USD', default=lambda self: self.env.ref('base.USD').id)
    
    exchange_rate_date = fields.Date('Fecha Tipo Cambio', default=fields.Date.context_today)
    exchange_rate = fields.Float('Tipo de Cambio USD', digits=(12, 4), tracking=True)
    
    total_amount_clp = fields.Float('Monto Total CLP', digits=(16, 2), tracking=True)
    total_amount_usd = fields.Float('Monto Total USD', compute='_compute_usd_amount', store=True, digits=(16, 2))
    
    price_per_m3_usd = fields.Float(
        'Precio USD/m³',
        compute='_compute_unit_prices',
        store=True,
        digits=(16, 2),
        help="Precio unitario calculado: total_amount_usd / physical_volume_m3"
    )

    # === INSERCIÓN PARA TRADING (CLP) ===
    average_price_m3 = fields.Float(
        string="Precio Promedio (CLP/m³)",
        compute="_compute_average_price_clp",
        store=True,
        digits=(12, 0), # Sin decimales para pesos chilenos visualmente limpios
        help="Costo unitario real calculado: Neto CLP / Vol. Comercial"
    )

    @api.depends('total_amount_clp', 'commercial_volume_m3')
    def _compute_average_price_clp(self):
        for rec in self:
            rec.average_price_m3 = rec.total_amount_clp / rec.commercial_volume_m3 if rec.commercial_volume_m3 > 0 else 0.0
    
    price_per_mbf_usd = fields.Float(
        'Precio USD/MBF',
        compute='_compute_unit_prices',
        store=True,
        digits=(16, 2),
        help="Precio unitario calculado: total_amount_usd / physical_volume_mbf"
    )
    
    # ==================== ARCHIVOS - FLUJO SECUENCIAL ====================
    oc_pdf_file = fields.Binary('📄 1. PDF Orden de Compra', help="Obligatorio solo si la OC no existe en Odoo")
    oc_pdf_filename = fields.Char('Nombre PDF OC')
    
    pdf_file = fields.Binary('📄 2. PDF Guía Despacho', required=True, help="Obligatorio: contiene datos de la guía")
    pdf_filename = fields.Char('Nombre PDF Guía')
    
    excel_file = fields.Binary('📊 3. Excel Packing List', required=True, help="Obligatorio: contiene detalle de lotes")
    excel_filename = fields.Char('Nombre Excel')
    
    files_ready = fields.Boolean('Archivos Listos', compute='_compute_files_ready', store=True)
    
    po_missing_alert = fields.Html(
        '⚠️ Alerta OC',
        compute='_compute_po_missing_alert',
        sanitize=False
    )
    
    # ==================== RELACIONES GENERADAS ====================
    picking_id = fields.Many2one('stock.picking', 'Albarán de Recepción', readonly=True, copy=False)
    invoice_id = fields.Many2one('account.move', 'Factura de Proveedor', readonly=True, copy=False)
    
    lot_ids = fields.One2many('stock.lot', 'reception_id', 'Lotes Importados', readonly=True)
    
    # ==================== ASIGNACIÓN DE PATIO (LOCATION) ====================
   # ASÍ DEBE QUEDAR (Blindado)
    location_id = fields.Many2one('stock.location', string='Patio de Asignación') # Sin required=True # Sin required=True
    
    # ==================== TOTALES ====================
    total_packages = fields.Integer('Total Paquetes', compute='_compute_totals', store=True)
    total_volume_m3 = fields.Float('Volumen Total m³', compute='_compute_totals', store=True, digits=(16, 3))
    total_volume_mbf = fields.Float('Volumen Total MBF', compute='_compute_totals', store=True, digits=(16, 3))
    
    log_notes = fields.Text('Registro de Procesamiento', readonly=True)

    # ==================== MÉTODOS COMPUTADOS ====================
    @api.depends('audit_log_ids')
    def _compute_omitted_count(self):
        """Calcular número de omisiones para mostrar en vista"""
        for rec in self:
            rec.omitted_count = len(rec.audit_log_ids.filtered(lambda l: l.action_type == 'omission'))

    def action_view_audit_logs(self):
        """Acción para ver logs de auditoría filtrados por esta recepción"""
        self.ensure_one()
        return {
            'name': f'Auditoría - {self.name}',
            'type': 'ir.actions.act_window',
            'res_model': 'madenat.audit.log',
            'view_mode': 'list,form',
            'domain': [('reception_id', '=', self.id)],
            'context': {'search_default_group_by_month': 1, 'search_default_group_by_lot': 1}
        }

    @api.depends('pdf_file', 'excel_file', 'purchase_id', 'oc_pdf_file')
    def _compute_files_ready(self):
        """Validar que los archivos necesarios estén cargados"""
        for rec in self:
            if rec.purchase_id or rec.manual_po_name:
                # OC existe en Odoo o está referenciada manualmente → solo necesita PDF + Excel
                rec.files_ready = bool(rec.pdf_file and rec.excel_file)
            else:
                # Sin OC referenciada → exige también el PDF de la OC
                rec.files_ready = bool(rec.oc_pdf_file and rec.pdf_file and rec.excel_file)

    def action_reset_to_draft(self):
        """
        🔄 RESETEO NIVEL DIOS (Arquitectura Madenat V3 - Odoo 18)
        -------------------------------------------------------
        - INTEGRIDAD: Bloquea si la madera ya avanzó (Contenedores/Consolidación).
        - ATOMICIDAD: Limpia Quants -> Lotes -> Pickings en una sola transacción.
        - SEGURIDAD: Usa sudo() y context 'force_delete' para bypass de bloqueos.
        - CERO RESIDUOS: Borrado físico (unlink) en lugar de desvincular (5,0,0).
        """
        self.ensure_one()
        _logger.info(f"🔄 INICIANDO RESETEO TOTAL - Guía: {self.name}")
        self._add_log("=" * 60)
        self._add_log("🔄 Iniciando saneamiento profundo de base de datos...")

        # ======================================================================
        # 1. VALIDACIÓN DE INTEGRIDAD GLOBAL (BLOQUEO DE SEGURIDAD)
        # ======================================================================
        if self.lot_ids:
            # A. Bloqueo por Contenedor
            lotes_en_contenedor = self.lot_ids.filtered(lambda l: getattr(l, 'container_id', False))
            if lotes_en_contenedor:
                raise UserError("⛔ INTEGRIDAD: Lotes ya están en CONTENEDOR. Saque la madera del contenedor primero.")

            # B. Bloqueo por Consolidación
            if any(getattr(l, 'is_consolidated', False) for l in self.lot_ids):
                raise UserError("⛔ INTEGRIDAD: La mercadería ya está CONSOLIDADA. Desconsolide antes de resetear.")

        try:
            with self.env.cr.savepoint():
                # ==============================================================
                # 2. SANEAMIENTO DE EXISTENCIAS (QUANTS) - ¡CRÍTICO!
                # ==============================================================
                # Borramos los registros de stock real para evitar el error de Foreign Key
                if self.lot_ids:
                    quants = self.env['stock.quant'].sudo().search([('lot_id', 'in', self.lot_ids.ids)])
                    if quants:
                        _logger.info(f"🧹 Eliminando {len(quants)} registros de stock físico (Quants)...")
                        quants.sudo().unlink()

                # ==============================================================
                # 3. ELIMINACIÓN DE ALBARANES (PICKINGS) - VERSIÓN ANIQUILADORA
                # ==============================================================
                # Buscamos IDs para evitar problemas de caché del objeto
                p_ids = self.env['stock.picking'].sudo().search([('origin', '=', self.name)]).ids
                if self.picking_id:
                    p_ids.append(self.picking_id.id)
                
                # Procesamos cada ID de forma individual y aislada
                for p_id in list(set(p_ids)):
                    # Buscamos el registro fresco de la DB en cada iteración
                    picking = self.env['stock.picking'].sudo().browse(p_id)
                    
                    if not picking.exists():
                        continue
                        
                    _logger.info(f"🔥 Aniquilando Albarán {picking.name} (ID: {p_id})")
                    
                    try:
                        with self.env.cr.savepoint():
                            # A. Forzar estado para saltar bloqueos de Odoo 18
                            picking.write({'state': 'draft'})
                            
                            # B. Limpiar movimientos (Moves) de forma atómica
                            if picking.move_ids:
                                moves = picking.move_ids
                                # Cambiamos estado de moves a draft para poder borrarlos
                                moves.write({'state': 'draft'})
                                picking.write({'move_ids': [(5, 0, 0)]}) # Desvincular
                                moves.with_context(force_delete=True).unlink()
                            
                            # C. Borrado físico final
                            picking.with_context(force_delete=True).unlink()
                            _logger.info(f"✅ Registro {p_id} eliminado exitosamente.")
                            
                    except Exception as e:
                        _logger.error(f"❌ Falló borrado ORM del albarán {p_id}: {e}")
                        # Rompemos la transacción intencionalmente. El raw SQL dejaba basura relacional.
                        raise UserError(
                            f"🛑 No se pudo eliminar el albarán {picking.name} por restricciones de base de datos.\n"
                            f"Desvincule o elimine sus movimientos manualmente antes de resetear.\n"
                            f"Detalle técnico: {e}"
                        )
                # ==============================================================
                # 4. ELIMINACIÓN DE LOTES Y STAGING (BORRADO FÍSICO)
                # ==============================================================
                # Primero el Staging (Tabla de pre-visualización)
                if self.reception_line_ids:
                    _logger.info("🧹 Eliminando líneas de Staging...")
                    self.reception_line_ids.with_context(force_delete=True).unlink()

                # Finalmente los Lotes (Ya sin bloqueos de Quants ni Pickings)
                if self.lot_ids:
                    lots_to_delete = self.lot_ids
                    # IMPORTANTE: Desvincular de la cabecera antes de borrar el objeto
                    self.write({'lot_ids': [(5, 0, 0)]})
                    lots_to_delete.sudo().with_context(force_delete=True).unlink()
                    _logger.info("🗑️ Lotes eliminados físicamente.")

                # ==============================================================
                # 5. RESETEO DE CABECERA
                # ==============================================================
                self.write({
                    'state': 'draft',
                    'picking_id': False,
                    'invoice_id': False,
                    'log_notes': False,
                    'omitted_count': 0,
                    'volume_variance_percent': 0.0,
                    'physical_volume_m3': 0.0,
                    'commercial_volume_m3': 0.0,
                })

        except Exception as e:
            _logger.error(f"❌ FALLO CRÍTICO EN RESETEO: {str(e)}")
            raise UserError(f"🛑 Error de Integridad: No se pudo limpiar la base de datos.\nDetalle: {e}")

        self._add_log("✅ <strong>SANEAMIENTO COMPLETADO:</strong> BD limpia, lista para nueva carga.")
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'lumber.reception',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_cancel(self):
        """Cancela la recepción sin borrar datos, solo cambia estado."""
        for rec in self:
            rec.write({'state': 'cancel'})
            rec._add_log("🚫 Recepción CANCELADA por el usuario.")

    # Campo para la tabla visual de resumen
    reception_summary_html = fields.Html(
        string="Resumen de Carga", 
        compute="_compute_reception_summary"
    )

    @api.depends('lot_ids')
    def _compute_reception_summary(self):
        """
        Genera el resumen HTML agrupando por DIMENSIÓN COMERCIAL (Visual).
        Arregla el problema de ver 7/4 o decimales en lugar de 6/4.
        """
        for rec in self:
            if not rec.lot_ids:
                rec.reception_summary_html = "<p class='text-muted'>No hay datos procesados aún.</p>"
                continue

            # Diccionario para agrupar
            summary = {}
            total_pieces = 0
            total_vol = 0.0

            for lot in rec.lot_ids:
                # -----------------------------------------------------------
                # 1. OBTENER DATOS (Ajusta los nombres si usas espesor_mm)
                # -----------------------------------------------------------
                # Intenta obtener 'thickness' o 'espesor_mm'
                t_mm = getattr(lot, 'thickness', getattr(lot, 'espesor_mm', 0.0))
                w_mm = getattr(lot, 'width', getattr(lot, 'ancho_mm', 0.0))
                l_m  = getattr(lot, 'length', getattr(lot, 'largo_m', 0.0))
                
                # -----------------------------------------------------------
                # 2. CALCULAR TEXTO VISUAL (LA REGLA DE ORO)
                # -----------------------------------------------------------
                
                # A. Espesor Visual (Fix 45mm -> 6/4)
                if 23.5 <= t_mm <= 24.5:
                    t_str = "24mm"
                else:
                    # Mapeo Comercial Manual
                    t_val = 0.0
                    if 37 <= t_mm <= 46: t_val = 1.5      # 6/4 (Tu regla)
                    elif 22 <= t_mm <= 29: t_val = 1.0    # 4/4
                    elif 30 <= t_mm <= 36: t_val = 1.25   # 5/4
                    elif 47 <= t_mm <= 56: t_val = 2.0    # 8/4
                    else: t_val = r4((t_mm / 25.4) * 4) / 4.0
                    
                    quarters = int(t_val * 4)
                    t_str = f"{quarters}/4"

                # B. Ancho Visual (Octavos)
                w_in = w_mm / 25.4
                eighths = int(round(w_in * 8))
                whole = eighths // 8
                rem = eighths % 8
                
                if rem == 0: w_str = f"{whole}"
                elif rem == 4: w_str = f"{whole} 1/2" if whole else "1/2"
                elif rem == 2: w_str = f"{whole} 1/4" if whole else "1/4"
                elif rem == 6: w_str = f"{whole} 3/4" if whole else "3/4"
                else: w_str = f"{whole} {rem}/8" if whole else f"{rem}/8"

                # -----------------------------------------------------------
                # 3. AGRUPAR POR TEXTO (NO POR NÚMERO)
                # -----------------------------------------------------------
                # Clave: (Producto, "6/4", "4 5/8", Largo)
                key = (lot.product_id.name, t_str, w_str, l_m)
                
                if key not in summary:
                    summary[key] = {'count': 0, 'pieces': 0, 'volume': 0.0}
                
                # Obtener piezas y volumen
                pz = getattr(lot, 'pieces', getattr(lot, 'piezas', 1))
                vol = getattr(lot, 'vol_physical_m3', getattr(lot, 'volume_purchase_m3', 0.0))

                summary[key]['count'] += 1
                summary[key]['pieces'] += pz
                summary[key]['volume'] += vol
                
                total_pieces += pz
                total_vol += vol

            # -----------------------------------------------------------
            # 4. GENERAR HTML
            # -----------------------------------------------------------
            html = """
            <table class="table table-bordered table-sm table-hover">
                <thead class="thead-light">
                    <tr>
                        <th>Producto</th>
                        <th class="text-center">Espesor</th>
                        <th class="text-center">Ancho</th>
                        <th class="text-center">Largo</th>
                        <th class="text-center">Paq.</th>
                        <th class="text-right">Piezas</th>
                        <th class="text-right">Volumen (m³)</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for key in sorted(summary.keys()):
                data = summary[key]
                prod, esp_vis, anc_vis, larg = key # Desempaquetamos los STRINGS
                
                html += f"""
                    <tr>
                        <td><b>{prod}</b></td>
                        <td class="text-center">{esp_vis}</td>  <td class="text-center">{anc_vis}</td>  <td class="text-center">{larg:.2f}</td>
                        <td class="text-center">{data['count']}</td>
                        <td class="text-right">{data['pieces']}</td>
                        <td class="text-right"><b>{data['volume']:.3f}</b></td>
                    </tr>
                """
            
            # Pie de tabla con totales
            html += f"""
                <tr class="table-active font-weight-bold">
                    <td colspan="4" class="text-right">TOTALES:</td>
                    <td class="text-center">{len(rec.lot_ids)}</td>
                    <td class="text-right">{total_pieces}</td>
                    <td class="text-right">{total_vol:.3f}</td>
                </tr>
                </tbody>
            </table>
            """
            rec.reception_summary_html = html
    
    # 2. AGREGAR este método nuevo
    
    def action_link_to_existing_po(self):
        """Acción llamada desde el botón de lista 'Vincular OC'"""
        for rec in self:
            # Lógica de vinculación - mantener placeholder
            pass

          
    @api.depends('purchase_id', 'manual_po_name', 'state')
    def _compute_po_missing_alert(self):
        for rec in self:
            if rec.state in ['draft', 'processing', 'verified'] and not rec.purchase_id and rec.manual_po_name:
                rec.po_missing_alert = f'''
                    <div class="alert alert-warning" role="alert" style="margin: 10px 0;">
                        <i class="fa fa-exclamation-triangle"></i>
                        <strong>⚠️ Atención:</strong>
                        La recepción está operando con una <strong>referencia manual de OC</strong>:
                        <strong>{rec.manual_po_name}</strong>.
                        <br/>
                        La operación <strong>no se detiene</strong>, pero debe completar o crear
                        la Orden de Compra correspondiente para mantener la trazabilidad,
                        facilitar el costeo y preparar el cierre financiero.
                    </div>
                '''
            elif rec.state in ['draft', 'processing'] and not rec.purchase_id and not rec.manual_po_name:
                rec.po_missing_alert = '''
                    <div class="alert alert-warning" role="alert" style="margin: 10px 0;">
                        <i class="fa fa-exclamation-triangle"></i>
                        <strong>⚠️ Atención:</strong>
                        No se detectó una Orden de Compra ni una referencia manual.
                        <br/>
                        Puede continuar la ingesta, pero se recomienda ingresar una referencia
                        de OC para asegurar la trazabilidad de esta operación.
                    </div>
                '''
            else:
                rec.po_missing_alert = False
           
    @api.depends('commercial_volume_m3')
    def _compute_commercial_mbf_from_m3(self):
        """Convertir m³ comercial (manual) a MBF automáticamente"""
        for rec in self:
            if rec.commercial_volume_m3 and rec.commercial_volume_m3 > 0:
                rec.commercial_volume_mbf = rec.commercial_volume_m3 / 2.36
            else:
                rec.commercial_volume_mbf = 0.0

    @api.depends('commercial_volume_m3', 'physical_volume_m3')
    def _compute_volume_variance(self):
        """Calcular diferencia porcentual y absoluta entre volúmenes"""
        for rec in self:
            if rec.commercial_volume_m3 and rec.commercial_volume_m3 > 0:
                # 1. Calculamos la fracción pura (SIN multiplicar por 100)
                variance = (rec.physical_volume_m3 - rec.commercial_volume_m3) / rec.commercial_volume_m3
                
                # 2. Guardamos con 4 decimales para que el porcentaje tenga 2 decimales (0.1533 = 15.33%)
                rec.volume_variance_percent = r4(variance)
                
                rec.volume_variance_m3 = r3(
                    rec.physical_volume_m3 - rec.commercial_volume_m3
                )
            else:
                rec.volume_variance_percent = 0.0
                rec.volume_variance_m3 = 0.0

    @api.depends('volume_variance_percent')
    def _compute_tolerance_status(self):
        """Determinar estado de tolerancia automático"""
        for rec in self:
            abs_variance = abs(rec.volume_variance_percent)
            if abs_variance <= 2.0:
                rec.tolerance_status = 'ok'
            elif abs_variance <= 10.0:
                rec.tolerance_status = 'warning'
            else:
                rec.tolerance_status = 'critical'
    
    @api.depends('reception_line_ids', 
                 'reception_line_ids.vol_purchase_m3', 
                 'reception_line_ids.vol_physical_m3',
                 'reception_line_ids.vol_mbf') # Agregamos dependencia de MBF
    def _compute_totals(self):
        """
        Cálculo BLINDADO (Consolidado):
        Suma siempre desde las líneas para reflejar cambios en tiempo real.
        Mantiene lógica de MBF y Paquetes original.
        """
        for rec in self:
            lines = rec.reception_line_ids
            
            # 1. Volúmenes (Físico y Comercial)
            rec.physical_volume_m3 = sum(lines.mapped('vol_physical_m3'))
            vol_comercial = sum(lines.mapped('vol_purchase_m3'))
            
            rec.commercial_volume_m3 = vol_comercial
            rec.total_volume_m3 = vol_comercial # Regla de Oro: Total = Comercial
            
            # 2. Conteo de Paquetes (Funcionalidad original mantenida)
            rec.total_packages = len(lines)
            
            # 3. MBF (Funcionalidad original mantenida)
            rec.total_volume_mbf = sum(lines.mapped('vol_mbf'))
    @api.depends('total_amount_clp', 'exchange_rate')
    def _compute_usd_amount(self):
        """Calcular monto total en USD basado en tipo de cambio"""
        for rec in self:
            if rec.exchange_rate and rec.exchange_rate > 0:
                rec.total_amount_usd = rec.total_amount_clp / rec.exchange_rate
            else:
                rec.total_amount_usd = 0.0
    @api.depends('total_amount_usd', 'physical_volume_m3', 'physical_volume_mbf')
    def _compute_unit_prices(self):
        """Calcular precios unitarios"""
        for rec in self:
            if rec.physical_volume_m3 > 0:
                rec.price_per_m3_usd = rec.total_amount_usd / rec.physical_volume_m3
            else:
                rec.price_per_m3_usd = 0.0
            
            if rec.physical_volume_mbf > 0:
                rec.price_per_mbf_usd = rec.total_amount_usd / rec.physical_volume_mbf
            else:
                rec.price_per_mbf_usd = 0.0
    # ==================== UTILIDADES ====================
    def _add_log(self, message):
        """Agregar mensaje timestamp al registro de procesamiento"""
        self.ensure_one()
        timestamp = datetime.now().strftime("%H:%M:%S")
        current_log = self.log_notes or ""
        self.log_notes = f"{current_log}\n[{timestamp}] {message}"

    def _get_current_exchange_rate(self):
        """
        Obtener tipo de cambio auxiliar (USD -> CLP).
        Busca el Dólar Oficial en Odoo para la fecha de la Guía.
        
        POLÍTICA FINANCIERA: No se permiten fallbacks silenciosos.
        Si no hay tipo de cambio disponible, el sistema BLOQUEA el procesamiento.
        """
        self.ensure_one()
        usd = self.env.ref('base.USD', raise_if_not_found=False)
        if not usd:
            raise UserError(
                "⛔ POLÍTICA FINANCIERA VIOLADA\n\n"
                "La moneda USD no está configurada en el sistema.\n"
                "Configure la moneda USD antes de procesar recepciones."
            )

        # Priorizamos la fecha de la guía, si no, la fecha de recepción
        target_date = self.guia_fecha or self.exchange_rate_date or fields.Date.context_today(self)

        try:
            # Obtiene la tasa de conversión nativa de Odoo (Cuántos CLP es 1 USD)
            rate = usd._get_conversion_rate(
                from_currency=usd,
                to_currency=self.env.company.currency_id, # CLP
                company=self.env.company,
                date=target_date
            )
            if rate <= 0:
                raise UserError(
                    "⛔ POLÍTICA FINANCIERA VIOLADA\n\n"
                    f"No hay tipo de cambio USD→CLP válido para la fecha {target_date}.\n"
                    "Configure el tipo de cambio en Contabilidad → Monedas antes de continuar."
                )
            return rate
        except Exception as e:
            raise UserError(
                "⛔ POLÍTICA FINANCIERA VIOLADA\n\n"
                f"Error obteniendo tipo de cambio: {str(e)}\n"
                "El sistema requiere un tipo de cambio válido para procesar recepciones."
            )
            

     #### ==================== MÉTODO PRINCIPAL CONSOLIDADO ====================


    def _create_lots_from_packing(self, pl_data):
        """
        ✅ VERSIÓN v8.0 (REFACTORIZADA A SERVICIO)
        Delega la creación pesada de lotes al servicio LumberReceptionService.
        Mantiene compatibilidad con el flujo actual.
        """
        self.ensure_one()
        service = LumberReceptionService(self.env)
        stats = service.create_lots_from_staging(self)
        
        self._add_log(
            f"✅ Proceso Finalizado (Servicio):\n"
            f"   - Lotes: {stats['created'] + stats['updated']}\n"
            f"   - Omitidos: {stats['omitted']}"
        )

        return True

    # ========================================================================
    # ✅ MÉTODO: LLENAR TABLA DE VALIDACIÓN (STAGING)
    # ========================================================================
    def _fill_staging_table(self, pl_data):
        """
        ✅ MÉTODO v5.5: LLENAR TABLA DE VALIDACIÓN (STAGING)
        - Inyecta metadatos de Triple Capa validados en shell.
        - Mantiene trazabilidad original (lot_name = p_code).
        - DELEGA el cálculo visual al ORM para garantizar consistencia.
        """
        # 1. Limpieza preventiva (Sin cambios)
        self.reception_line_ids.unlink()
        
        lines_to_create = []
        raw_lines = pl_data.get('lines', [])
        
        # Invocamos la inteligencia del Mixin
        IngestMixin = self.env['madenat.lumber.ingest.mixin']
        
        for item in raw_lines:
            # A. Obtención de Producto
            p_code = item.get('product_code', '') 
            p_name = item.get('product_name', 'Madera Genérica')
            product = IngestMixin.find_or_create_lumber_product(p_code, p_name)
            
            vol_detectado = item.get('volume_m3', 0.0)
            
            # B. Construcción de la Línea (Blindada)
            lines_to_create.append({
                'reception_id': self.id,
                
                # 🔥 IDENTIDAD REAL: Se mantiene tu lógica original
                'lot_name': p_code if p_code else f"S/N-{item.get('package_no', 0)}", 
                'product_id': product.id,
                'pieces': item.get('pieces', 0),
                
                # 💎 REALIDADES VOLUMÉTRICAS
                'vol_shipment_m3': vol_detectado, 
                'vol_purchase_m3': vol_detectado, 
                
                # 🪞 CAPA ESPEJO (UI)
                # FIX v5.5: Comentamos la inyección directa del Parser. 
                # Delegamos al @api.depends _compute_visual_defaults para que aplique 
                # la matemática correcta (fracciones vs cuartos) según el perfil.
                # 'thickness_visual': item.get('thickness_visual', ''),
                # 'width_visual': item.get('width_visual', ''),
                'export_calculation_rule': item.get('export_rule', 'metric'),
                
                # 🎯 CAPA NOMINAL (Dato validado en Odoo Shell)
                'thickness_nominal': item.get('thickness_nominal', 0.0),
                'width_nominal': item.get('width_nominal', 0.0),
                
                # 📏 CAPA FÍSICA (Base para stock milimétrico)
                'thickness': item.get('thickness_mm', 0.0),
                'width': item.get('width_mm', 0.0),
                'length': item.get('length_m', 0.0),

                # 📐 LARGO INPUT (Blanks: pies directos del Excel)
                'length_input_raw': item.get('length_input_raw', item.get('length_m', 0.0)),
                'lengthuom': item.get('length_uom', 'm'),
                
                # COSTEO
                'estimated_cost_usd': vol_detectado * item.get('unit_price_usd', 0.0)
            })
        
        # 2. Creación en bloque (Optimizado)
        if lines_to_create:
            self.env['lumber.reception.line'].create(lines_to_create)
            _logger.info(f"✅ Staging listo: {len(lines_to_create)} líneas inyectadas.")
            self._add_log(f"📋 Pestaña de Análisis Comercial preparada con {len(lines_to_create)} paquetes.")

    def _validate_product_uom_for_lumber(self, product):
        """
        ✅ VALIDACIÓN CRÍTICA: Verifica que el producto tenga UoM de categoría Volume
        
        Args:
            product: product.product o product.template
            
        Raises:
            UserError: Si la UoM no es de categoría Volume
        """
        if not product:
            return
        
        # Obtener product.template si recibimos product.product
        if hasattr(product, 'product_tmpl_id'):
            product_tmpl = product.product_tmpl_id
        else:
            product_tmpl = product
        
        # Referencia a categoría Volume
        volume_category = self.env.ref('uom.product_uom_categ_vol', raise_if_not_found=False)
        
        if not volume_category:
            _logger.warning("⚠️ No se encontró categoría UoM 'Volume'. Saltando validación.")
            return
        
        # VALIDACIÓN
        if product_tmpl.uom_id.category_id != volume_category:
            raise UserError(_(
                "❌ ERROR CRÍTICO DE UNIDAD DE MEDIDA\n\n"
                "No se puede recepcionar el producto:\n"
                "%(product)s (Código: %(code)s)\n\n"
                "🔍 PROBLEMA DETECTADO:\n"
                "   • UoM del producto: %(uom)s\n"
                "   • Categoría actual: %(category)s ❌\n"
                "   • Categoría requerida: Volume (m³, MBF) ✅\n\n"
                "⚠️ CAUSA:\n"
                "Los productos de madera DEBEN usar unidades de volumen (m³ o MBF) "
                "para garantizar cálculos correctos de inventario y costeo.\n\n"
                "🔧 SOLUCIÓN:\n"
                "1. Vaya a: Inventario → Productos → %(product)s\n"
                "2. En 'Información General', cambie 'Unidad de Medida' a 'm³'\n"
                "3. Guarde el producto\n"
                "4. Vuelva a procesar esta recepción\n\n"
                "💡 IMPORTANTE:\n"
                "Si cambia la UoM, verifique que no haya inventario existente "
                "del producto que pueda generar inconsistencias."
            ) % {
                'product': product_tmpl.display_name,
                'code': product_tmpl.default_code or 'Sin código',
                'uom': product_tmpl.uom_id.name,
                'category': product_tmpl.uom_id.category_id.name,
            })
    
    def _create_stock_picking(self):
            """
            📦 MOTOR DE STOCK V5.4 (Odoo 18 Optimized + Blindaje 3 Decimales)
            Asegura que 'Demanda' y 'Hecho' sean idénticos para evitar duplicados.
            """
            self.ensure_one()
            uom_cubic = self.env.ref('uom.product_uom_cubic_meter')
            
            # 🛡️ Importamos la herramienta nativa de redondeo de Odoo
            from odoo.tools.float_utils import float_round
            
            # 1. Búsqueda de operación
            picking_type = self.env['stock.picking.type'].search([
                ('code', '=', 'incoming'),
                ('company_id', '=', self.env.company.id)
            ], limit=1)

            if not picking_type:
                raise UserError("No se encontró tipo de operación 'Recepción'.")

            # 2. Cabecera del Albarán
            picking = self.env['stock.picking'].create({
                'partner_id': self.supplier_id.id,
                'picking_type_id': picking_type.id,
                'location_id': picking_type.default_location_src_id.id,
                'location_dest_id': self.location_id.id or picking_type.default_location_dest_id.id,
                'origin': self.name,  # Mantenemos el nombre puro para su consulta SQL
                'lumber_reception_id': self.id,
                'company_id': self.env.company.id,
            })

            # 3. Generación de Movimientos (Atómicos)
            for lot in self.lot_ids:
                
                # 🎯 BLINDAJE ESTRICTO: Forzamos matemáticamente a 3 decimales.
                # Corta de raíz el problema de los "Flotantes Fantasma" de Python.
                safe_qty = float_round(lot.volume_purchase_m3, precision_digits=3)

                # Creamos el movimiento (Demanda)
                move = self.env['stock.move'].create({
                    'name': f"Lote: {lot.name}",
                    'product_id': lot.product_id.id,
                    'product_uom_qty': safe_qty,
                    'product_uom': uom_cubic.id,
                    'picking_id': picking.id,
                    'location_id': picking.location_id.id,
                    'location_dest_id': picking.location_dest_id.id,
                    'company_id': self.env.company.id,
                    'picked': True, # Odoo 18: Marca como listo para validar
                })
                
                # Creamos la línea de movimiento (El 'Hecho')
                # Vincular el lote aquí es lo que crea el Quant correcto
                self.env['stock.move.line'].create({
                    'move_id': move.id,
                    'picking_id': picking.id,
                    'product_id': lot.product_id.id,
                    'lot_id': lot.id,
                    'quantity': safe_qty,
                    'product_uom_id': uom_cubic.id,
                    'location_id': picking.location_id.id,
                    'location_dest_id': picking.location_dest_id.id,
                    'picked': True,
                })

            # 4. VALIDACIÓN FINAL
            if picking.move_ids:
                picking.action_confirm()
                picking.action_assign()
                # button_validate dispara la creación de Quants (12 in / 12 out)
                picking.button_validate()
                
            return picking
    
    # =================================================================================
    # 1. MOTOR INTELIGENTE DE ÓRDENES DE COMPRA
    # =================================================================================
    def _find_or_create_po_intelligent(self, dg_data, oc_data):
            """
            🎯 LÓGICA INTELIGENTE V4.1 — CONSOLIDADA, AUDITADA Y BLINDADA
            
            Propósito:
            1. Gestión de Proveedor: Identifica por RUT. Si el nombre es genérico o contiene el RUT, 
            aplica limpieza quirúrgica y actualiza los datos maestros.
            2. Vínculo de OC: Búsqueda tolerante por clave normalizada (Display vs Key).
            3. Creación Automática: Genera la OC si hay datos de compra disponibles.
            4. Modo Manual: Fallback no bloqueante con trazabilidad total en Chatter.
            
            Rigor Técnico aplicado: Separación estricta de Identidad (VAT) y Nombre (Legal).
            """
            self.ensure_one()
            # 🎯 INSTANCIAR PARSER EXTERNO
            parser = self.env['madenat.reception.parser']
            
            # 📥 1. EXTRACCIÓN Y PREPARACIÓN DE IDENTIFICADORES
            supplier_rut = (dg_data.get('supplier_rut') or '').strip()
            raw_detected_name = dg_data.get('supplier_name_detected') or dg_data.get('supplier_name') or ""
            po_ref_raw = dg_data.get('po_ref') or ''

            # Fallback desde PDF de OC si la guía no detectó referencia
            if not po_ref_raw and oc_data:
                po_ref_raw = oc_data.get('po_ref_fallback') or ''

            # Normalización usando el nuevo Parser Externo
            po_ref = parser.normalize_po_display(po_ref_raw)
            po_key = parser.normalize_po_key(po_ref_raw)

            # 🛑 2. VALIDACIONES DE SEGURIDAD CRÍTICAS (No permitir avanzar sin datos base)
            if not po_key:
                raise UserError(
                    "⚠️ No se detectó referencia de OC en el PDF.\n"
                    "Por favor, complete el campo 'OC Manual' para continuar el procesamiento."
                )
            if not supplier_rut:
                raise UserError(
                    "⚠️ No se detectó RUT del proveedor en el PDF.\n"
                    "Debe seleccionar o crear el Proveedor manualmente antes de procesar."
                )

            # ✂️ 3. RIGOR TÉCNICO: LIMPIEZA DEL NOMBRE DEL PROVEEDOR
            # Amputamos el RUT del nombre si vienen pegados (Caso Forestal Andes)
            import re
            # Cortamos por la palabra RUT o por el patrón numérico del RUT chileno
            clean_name_parts = re.split(r'R\.?U\.?T\.?|[\d]{1,2}\.[\d]{3}\.', raw_detected_name, flags=re.IGNORECASE)
            detected_name = clean_name_parts[0].strip()
            # Limpieza de caracteres residuales al final (puntos, guiones, comas)
            detected_name = re.sub(r'[:\-\s\.,]+$', '', detected_name).strip()

            # 🛡️ 4. GESTIÓN DE PROVEEDOR (PARTNER)
            supplier = self.env['res.partner'].search([('vat', '=', supplier_rut)], limit=1)
            
            if supplier:
                # Auditoría de Calidad: Si el nombre en BD es genérico o tiene el RUT pegado, lo curamos.
                is_generic = any(x in supplier.name for x in ["Proveedor", supplier_rut])
                if is_generic and detected_name and len(detected_name) > 2:
                    supplier.write({'name': detected_name[:128]})
                    self._add_log(f"🔄 Datos Maestros: Nombre actualizado a '{detected_name}' (VAT preservado)")
            else:
                # Creación profesional: RUT limpio en VAT y Nombre limpio en Name.
                name_to_use = detected_name or f"Proveedor {supplier_rut}"
                supplier = self.env['res.partner'].create({
                    'name': name_to_use[:128],
                    'vat': supplier_rut,
                    'is_company': True,
                    'supplier_rank': 1,
                })
                self._add_log(f"📝 Nuevo proveedor registrado en BD: {supplier.name}")

            # 🔎 5. BÚSQUEDA DE OC (CASO 1: EXISTENTE EN SISTEMA)
            candidate_pos = self.env['purchase.order'].search([
                ('partner_id', '=', supplier.id),
                ('state', 'in', ['draft', 'sent', 'to approve', 'purchase', 'done'])
            ])

            # Match tolerante por clave (compara contra nombre Odoo y referencia proveedor)
            po = candidate_pos.filtered(
                lambda p: self._po_key(p.partner_ref or '') == po_key
                    or self._po_key(p.name or '') == po_key
            )[:1]

            if po:
                self._add_log(f"✅ OC ENCONTRADA: {po.name} vinculada exitosamente.")

                # Análisis de Volumen Comparativo (Auditoría de Saldo)
                total_ordered = po.order_line[0].product_qty if po.order_line else 0
                previous_receptions = self.env['lumber.reception'].search([
                    ('purchase_id', '=', po.id),
                    ('state', '=', 'done'),
                    ('id', '!=', self.id)
                ])
                total_received = sum(previous_receptions.mapped('physical_volume_m3'))

                self._add_log(f"   📊 Análisis de Saldo: OC {total_ordered}m³ | Recibido {total_received}m³ | Esta Guía {self.physical_volume_m3}m³")

                # Escritura de vínculo exitoso
                self.write({
                    'purchase_id': po.id,
                    'purchase_order': po.name,
                    'manual_po_name': False,
                    'supplier_id': supplier.id,
                })
                return po, supplier

            # 🆕 6. GESTIÓN DE OC (CASO 2: CREACIÓN DESDE DATOS DE PDF COMPRA)
            if oc_data:
                self._add_log(f"🆕 OC '{po_ref}' no existe → Iniciando creación desde PDF de Compra.")
                po = self.create_po_from_oc_data(po_ref, supplier, oc_data)
                # Aseguramos que la referencia del proveedor en la PO sea la normalizada
                po.write({'partner_ref': po_ref})
                
                self.write({
                    'purchase_id': po.id,
                    'purchase_order': po.name,
                    'manual_po_name': False,
                    'supplier_id': supplier.id,
                })
                return po, supplier

            # ⚠️ 7. GESTIÓN DE OC (CASO 3: FALLBACK MANUAL / NO BLOQUEANTE)
            self._add_log(f"⚠️ OC '{po_ref}' sin coincidencia → Modo Referencia Manual habilitado.")
            
            # Guardamos la intención del PDF en manual_po_name para guiar al usuario
            self.write({
                'state': 'processing',
                'purchase_id': False,
                'purchase_order': False,
                'manual_po_name': po_ref,
                'supplier_id': supplier.id,
            })
            
            # Notificación en Chatter para auditoría futura
            self.message_post(
                body=(
                    f"🔎 <b>Información de Ingesta OCR:</b><br/>"
                    f"Referencia detectada: <b>{po_ref}</b><br/>"
                    f"Proveedor: {supplier.name} (RUT: {supplier_rut})<br/><br/>"
                    f"<i>Aviso: La Orden de Compra no existe en Odoo. El sistema ha activado el flujo manual "
                    f"para que el usuario realice la vinculación definitiva antes de validar.</i>"
                )
            )
            return None, supplier


    # ==================== MÉTODOS DE PROCESAMIENTO DE OC ====================
    def create_po_from_oc_data(self, po_ref, supplier, oc_data):
            """
            ✅ VERSIÓN 3.0 - CONSOLIDADA ODOO 18
            - Usa 'type'='consu' (Goods)
            - Tracking 'lot'
            - Lógica de preservación de UoM (variante _NEW)
            """
            self.ensure_one()

            product_code = oc_data.get('product_code', 'MADERA_GENERICA')
            product_name = oc_data.get('product_name', product_code)
            uom_cubic = self.env.ref('uom.product_uom_cubic_meter')

            # ====================================================================
            # 1️⃣ BUSCAR O CREAR PRODUCTO (ODOO 18)
            # ====================================================================
            lumber_product = self.env['product.product'].search([
                ('default_code', '=', product_code)
            ], limit=1)

            if not lumber_product:
                # ✅ CREAR PRODUCTO ALMACENABLE (Odoo 18 - Correcto)
                _logger.info(f"🆕 Creando producto {product_code} para OC...")
                
                lumber_product = self.env['product.product'].create({
                    'name': product_name,
                    'default_code': product_code,
                    'type': 'consu',  # ✅ Almacenable (Goods con stock)
                    'tracking': 'lot',  # ✅ Trazabilidad por lote
                    'uom_id': uom_cubic.id,
                    'uom_po_id': uom_cubic.id,
                    'active': True,
                    'purchase_ok': True,
                    'sale_ok': True,
                })
                
                msg = f"🆕 Producto creado: {product_code} (Goods almacenable, m³, tracking por lote)"
                _logger.info(msg)
                self._add_log(msg)
                
            else:
                # ====================================================================
                # 2️⃣ VALIDAR PRODUCTO EXISTENTE (Solo lectura - NO mutar)
                # ====================================================================
                tmpl = lumber_product.product_tmpl_id
                
                # ✅ Validación de tipo (Odoo 18 compatible)
                if tmpl.type != 'consu':
                    msg = (f"⚠️ Producto {product_code} tiene tipo '{tmpl.type}' "
                        f"(esperado 'consu' para almacenable). Revisar maestro.")
                    _logger.warning(msg)
                    self._add_log(msg)
                
                # ✅ Validación de UoM con control de movimientos
                if tmpl.uom_id != uom_cubic:
                    # Verificar si hay movimientos antes de cambiar UoM
                    moves_exists = self.env['stock.move'].search([
                        ('product_id.product_tmpl_id', '=', tmpl.id)
                    ], limit=1)
                    
                    if moves_exists:
                        # ❌ NO SE PUEDE CAMBIAR - Crear producto nuevo
                        msg = (f"❌ Producto {product_code} tiene movimientos de stock. "
                            f"No se puede cambiar UoM. Creando variante nueva...")
                        _logger.warning(msg)
                        self._add_log(msg)
                        
                        # Archivar producto antiguo y crear uno nuevo
                        # tmpl.active = False (Mejor no archivar automáticamente, solo crear nuevo)
                        
                        lumber_product = self.env['product.product'].create({
                            'name': tmpl.name,
                            'default_code': f"{tmpl.default_code}_NEW",
                            'type': 'consu',  # ✅ Almacenable
                            'tracking': 'lot',  # ✅ Trazabilidad
                            'uom_id': uom_cubic.id,
                            'uom_po_id': uom_cubic.id,
                            'active': True,
                            'purchase_ok': True,
                            'sale_ok': True,
                        })
                        
                        msg = f"✅ Producto nuevo creado: {lumber_product.default_code}"
                        _logger.info(msg)
                        self._add_log(msg)
                        
                    else:
                        # ✅ SEGURO cambiar UoM (no hay movimientos)
                        tmpl.write({
                            'uom_id': uom_cubic.id,
                            'uom_po_id': uom_cubic.id,
                        })
                        msg = f"✅ UoM del producto {product_code} actualizado a m³"
                        _logger.info(msg)
                        self._add_log(msg)

            # ====================================================================
            # 3️⃣ LOG DE VALIDACIÓN
            # ====================================================================
            _logger.info(
                f"✅ Producto para OC: ID={lumber_product.id}, "
                f"UoM={lumber_product.uom_id.name}, "
                f"Tipo={lumber_product.product_tmpl_id.type}"
            )

            # ====================================================================
            # 4️⃣ CREAR ORDEN DE COMPRA
            # ====================================================================
            po_vals = {
                'name': po_ref,
                'partner_id': supplier.id,
                'currency_id': self.env.ref('base.USD').id,
                'date_order': fields.Datetime.now(),
                'order_line': [(0, 0, {
                    'product_id': lumber_product.id,
                    'product_qty': oc_data.get('total_volume_m3', 550),
                    'price_unit': oc_data.get('unit_price_usd', 240),
                    'name': f"{product_name} {oc_data.get('quality', 'COL A').upper()}",
                    'product_uom': lumber_product.uom_id.id,  # ← Usa UoM del producto
                })],
            }
            
            # ====================================================================
            # 5️⃣ CAMPOS OPCIONALES MADENAT (si existen en el modelo)
            # ====================================================================
            if hasattr(self.env['purchase.order'], 'lumber_quality'):
                po_vals['lumber_quality'] = oc_data.get('quality', 'col_a')
            
            if hasattr(self.env['purchase.order'], 'wood_type'):
                po_vals['wood_type'] = 'pine'
            
            if hasattr(self.env['purchase.order'], 'treatment'):
                po_vals['treatment'] = 'kiln_dried'
            
            if hasattr(self.env['purchase.order'], 'thickness_mm'):
                po_vals['thickness_mm'] = oc_data.get('thickness_mm', 45.0)

            # ====================================================================
            # 6️⃣ CREAR Y CONFIRMAR OC
            # ====================================================================
            po = self.env['purchase.order'].with_context(validate_uom=False).create(po_vals)
            po.button_confirm()

            msg = (f"✅ OC creada: {po.name} "
                f"({oc_data.get('total_volume_m3')}m³ @ "
                f"USD {oc_data.get('unit_price_usd')}/m³)")
            self._add_log(msg)
            
            return po


    def _assign_costs_to_lots(self, pl_data):
        """✅ PASO 9: Asignar costos de compra a cada lote"""
        self.ensure_one()
        _logger.info(f"💰 Iniciando asignación de costos para {self.name}")
        self._add_log("💰 Asignando costos a lotes...")
        
        try:
            # Buscar lotes de esta recepción
            lots = self.env['stock.lot'].search([
                ('reception_id', '=', self.id)
            ])
            
            if not lots:
                _logger.warning("⚠️ No hay lotes para asignar costos")
                self._add_log("⚠️ No se encontraron lotes")
                return
            
            # Mapear líneas de packing a lotes
            lines = pl_data.get('lines', [])
            
            for idx, (lot, line) in enumerate(zip(lots, lines), 1):
                try:
                    precio_usd_m3 = float(line.get('unit_price_usd', 0))
                    volumen_m3 = float(line.get('volume_m3', 0))
                    
                    if not precio_usd_m3 or not volumen_m3:
                        _logger.warning(f"⚠️ Lote {lot.name} sin precio/volumen - saltando")
                        continue
                    
                    # Actualizar campos de precio en el lote
                    lot_exchange_rate = self.exchange_rate if self.exchange_rate and self.exchange_rate > 0 else 0.0
                    lot.write({
                        'purchase_price_usd_per_m3': precio_usd_m3,
                        'purchase_exchange_rate': lot_exchange_rate,
                    })
                    
                    # Calcular monto total de madera
                    monto_usd = precio_usd_m3 * volumen_m3
                    
                    # Crear línea de costo tipo "wood"
                    self.env['stock.lot.cost.line'].create({
                        'lot_id': lot.id,
                        'name': f"Costo madera {lot.name}",
                        'cost_type': 'wood',
                        'amount_usd': monto_usd,
                        'date': fields.Date.today(),
                        'notes': f"Origen: {self.name}",
                    })
                    
                    msg = f"✅ Lote {lot.name}: ${monto_usd:.2f} USD (${precio_usd_m3}/m³ × {volumen_m3:.3f} m³)"
                    _logger.info(msg)
                    self._add_log(msg)
                    
                except Exception as e:
                    _logger.error(f"❌ Error asignando costo a lote {idx}: {str(e)}")
                    self._add_log(f"❌ Error lote {idx}: {str(e)}")
                    raise
            
            _logger.info(f"✅ Costos asignados a {len(lots)} lotes")
            self._add_log("✅ Costos asignados correctamente")
            

        except Exception as e:
            _logger.error(f"❌ Error en asignación de costos: {str(e)}")
            self._add_log(f"❌ Error crítico: {str(e)}")
            raise UserError(f"Error asignando costos: {str(e)}")

    def _update_po_reception_stats(self, po, reception_volume_m3):
        """✅ PASO 10: Actualizar estadísticas de recepción en la OC"""
        self.ensure_one()
        
        try:
            # Calcular total recibido de TODAS las recepciones de esta OC
            all_receptions = self.env['lumber.reception'].search([
                ('purchase_id', '=', po.id),
                ('state', '=', 'done')
            ])
            
            total_received = sum(all_receptions.mapped('physical_volume_m3'))
            total_po = sum(po.order_line.mapped('product_qty')) if po.order_line else 0
            pending = total_po - total_received
            percent_complete = (total_received / total_po * 100) if total_po > 0 else 0
            
            self._add_log(f"📊 Estadísticas OC {po.name}:")
            self._add_log(f"   Total ordenado: {total_po:.2f} m³")
            self._add_log(f"   Total recibido: {total_received:.2f} m³ ({percent_complete:.1f}%)")
            self._add_log(f"   Pendiente: {pending:.2f} m³")
            self._add_log(f"   N° guías procesadas: {len(all_receptions)}")
            
            _logger.info(f"📊 OC {po.name}: {total_received:.2f}/{total_po:.2f} m³ ({percent_complete:.1f}%)")
            
            # Si la OC tiene campos personalizados, actualizarlos
            if hasattr(po, 'received_volume_m3'):
                po.write({
                    'received_volume_m3': total_received,
                    'pending_volume_m3': pending,
                    'percent_completed': percent_complete
                })
            
        except Exception as e:
            _logger.error(f"❌ Error actualizando OC: {str(e)}")
            self._add_log(f"⚠️ Estadísticas OC: {str(e)}")

    
    def action_process_documents(self):
        """
        🎯 PUNTO DE ENTRADA SIMPLIFICADO (Fase 4.2)
        Delega el pipeline completo al motor de workflow.
        """
        self.ensure_one()
        from .reception_workflow import LumberReceptionWorkflow
        
        workflow = LumberReceptionWorkflow(self.env)
        return workflow.run_ingestion_pipeline(self)

    def action_verify_data(self):
        """Lee el Excel, calcula volúmenes y muestra datos para revisión"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError("Archivo Excel requerido")
        
        self.reception_line_ids.unlink()
        
        excel_bytes = base64.b64decode(self.excel_file)
        workbook = load_workbook(filename=io.BytesIO(excel_bytes))
        sheet = workbook.active
        
        lines = []
        IngestMixin = self.env['madenat.lumber.ingest.mixin'] # Instanciar el Mixin
        
        # Asumiendo que la primera fila es el encabezado, comenzamos desde la segunda fila (índice 2 en openpyxl)
        for r in range(2, sheet.max_row + 1):
            try:
                # Ajusta los índices de columna según tu estructura de Excel
                lote = str(sheet.cell(r, 1).value or "").strip() # Columna A
                prod = str(sheet.cell(r, 2).value or "").strip() # Columna B
                qty = int(sheet.cell(r, 3).value or 0)           # Columna C
                vol = float(sheet.cell(r, 4).value or 0.0)       # Columna D
                
                if not lote: continue
                
                # Inteligencia del Mixin
                product_obj = IngestMixin.find_or_create_lumber_product(lote, prod)
                calc = IngestMixin.calculate_normalized_volumes(product_obj, vol, 'm3')
                
                lines.append({
                    'reception_id': self.id,
                    'lot_name': lote,
                    'product_id': product_obj.id,
                    'pieces': qty,
                    'vol_shipment_m3': calc['vol_shipment_m3'], # Real
                    'vol_purchase_m3': calc['volume_purchase_m3'], # Nominal
                    'calc_method': calc['method']
                })
            except Exception:
                continue
        
        if lines:
            self.env['lumber.reception.line'].create(lines)
            self.write({'state': 'verified'}) # Pasamos a estado 'Verificado'
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {'title': 'Listo', 'message': f'{len(lines)} líneas verificadas.', 'type': 'success'}
            }
        else:
            raise UserError("No se leyeron datos.")

    def action_confirm_reception(self):

            """
            🚀 CONSOLIDADO MADENAT V4.3 (PRO): Staging -> Stock.Lot
            Alineado con el motor de Fracciones y Protección de Factura Blanks.
            """
            self.ensure_one()
            _logger.info(f"🚀 INICIANDO CONFIRMACIÓN FINAL - Guía: {self.name}")
            
            # 1. VALIDACIONES DE ESTADO
            if self.state != 'verified':
                raise UserError("⚠️ La recepción debe estar en estado 'Verificada' para confirmar.")
            if not self.reception_line_ids:
                raise UserError("⚠️ No existen líneas cargadas en el staging.")

            # =========================================================================
            # 🛡️ GATE 2 & 3: SEGURIDAD Y CRIPTOGRAFÍA
            # =========================================================================
            from .ingestion_gate import Gate2CommercialAnalysis, Gate3PreCommit
            
            # Validación Comercial Final
            gate2 = Gate2CommercialAnalysis(self.env)
            gate2_result = gate2.validate(self.reception_line_ids)
            if not gate2_result.is_valid:
                self._add_log(f"❌ GATE 2 BLOQUEANTE: {gate2_result.audit_summary}")
                raise UserError(gate2_result.user_message)
                
            # Snapshot Inmutable (Notaría Digital)
            gate3 = Gate3PreCommit(self.env)
            snapshot_json, sha256_hash = gate3.generate_signature(self, self.reception_line_ids)
            self.write({
                'audit_snapshot': snapshot_json,
                'audit_hash': sha256_hash
            })
            self._add_log(f"🔐 GATE 3: Firma Digital Creada: {sha256_hash[:15]}...")

            # 2. AUTO-REPARACIÓN DE PRODUCTOS (Odoo 18 Ready)
            for line in self.reception_line_ids:
                prod = line.product_id
                if not prod.is_storable or prod.tracking != 'lot':
                    _logger.info(f"🔧 Reparando configuración técnica de: {prod.name}")
                    prod.sudo().write({
                        'is_storable': True,
                        'type': 'consu', 
                        'tracking': 'lot'
                    })

            try:
                with self.env.cr.savepoint():
                    # 3. CREACIÓN DE STOCK REAL
                    # _create_lots_from_packing usará los volúmenes ya calculados en el staging.
                    self._create_lots_from_packing({}) 

                    created_lots = self.env['stock.lot'].search([('reception_id', '=', self.id)])
                    if not created_lots:
                        raise UserError("⚠️ El motor no generó lotes. Revisa los logs de error.")

                    # 4. GENERACIÓN DEL ALBARÁN (Picking) mediante Servicio
                    service = LumberReceptionService(self.env)
                    picking = service.create_stock_picking(self)

                    # 5. CIERRE Y VINCULACIÓN
                    self.write({
                        'state': 'done',
                        'lot_ids': [(6, 0, created_lots.ids)],
                        'picking_id': picking.id if picking else False
                    })
                    
                    # Aprobación técnica para uso inmediato en producción/venta
                    created_lots.write({'technical_validation': 'approved'})

                    self._add_log("🏁 RECEPCIÓN FINALIZADA: Stock disponible y lotes vinculados.")
                    
                    return {
                        'type': 'ir.actions.client',
                        'tag': 'display_notification',
                        'params': {
                            'title': '¡Recepción Exitosa!',
                            'message': f'Se generaron {len(created_lots)} lotes correctamente.',
                            'type': 'success',
                            'sticky': False,
                            'next': {'type': 'ir.actions.act_window_close'} # Cierra y refresca
                        }
                    }

            except Exception as e:
                _logger.error(f"❌ FALLO CRÍTICO EN CONFIRMACIÓN: {str(e)}", exc_info=True)
                raise UserError(f"🛑 Error en el motor de inventario: {str(e)}")

    def validate_product_configuration(self, product):
        """
        ODOO 18 - Validación estricta de configuración de producto.
        
        Verifica que el producto esté correctamente configurado para:
        - Tipo: "consu" (Goods en Odoo 18)
        - Tracking: "lot" (trazabilidad por lote)
        - UoM: m³
        
        Lanza UserError si la configuración es incorrecta.
        """
        uom_m3 = self.env.ref("uom.product_uom_cubic_meter")
        errors = []
        
        # 1. VALIDAR TIPO (debe ser "consu" en Odoo 18)
        if product.type != "consu":
            errors.append(
                f"• Tipo incorrecto: '{product.type}' (debe ser 'consu' = Goods en Odoo 18)"
            )
        
        # 2. VALIDAR TRACKING (debe ser "lot" para trazabilidad)
        if product.tracking != "lot":
            errors.append(
                f"• Tracking incorrecto: '{product.tracking}' (debe ser 'lot' para trazabilidad por lote/tarja)"
            )
        
        # 3. VALIDAR UoM (debe ser m³)
        if product.uom_id != uom_m3:
            errors.append(
                f"• UoM incorrecta: '{product.uom_id.name}' (debe ser 'm³')"
            )
        
        if product.uom_po_id != uom_m3:
            errors.append(
                f"• UoM de compra incorrecta: '{product.uom_po_id.name}' (debe ser 'm³')"
            )
        
        # 4. LEVANTAR ERROR SI HAY PROBLEMAS
        if errors:
            # Verificar si tiene movimientos de stock
            moves_count = self.env["stock.move"].search_count([
                ("product_id", "=", product.id)
            ])
            
            error_msg = (
                f"❌ CONFIGURACIÓN INCORRECTA DEL PRODUCTO\n\n"
                f"Producto: {product.display_name} (Código: {product.default_code})\n\n"
                f"Problemas detectados:\n" + "\n".join(errors) + "\n\n"
            )
            
            if moves_count > 0:
                error_msg += (
                    f"⚠️ CRÍTICO: El producto tiene {moves_count} movimiento(s) de stock.\n"
                    f"NO se puede cambiar la configuración automáticamente.\n\n"
                    f"SOLUCIONES:\n"
                    f"1. Archivar este producto y crear uno nuevo con código '{product.default_code}_V2'\n"
                    f"2. Contactar con soporte técnico para corrección manual de datos históricos\n"
                    f"3. Usar un código de producto diferente para esta operación"
                )
            else:
                error_msg += (
                    f"✅ El producto NO tiene movimientos de stock.\n\n"
                    f"SOLUCIÓN RÁPIDA:\n"
                    f"1. Ir a Inventario → Productos → {product.display_name}\n"
                    f"2. En 'Tipo de Producto': seleccionar 'Goods'\n"
                    f"3. Activar 'Track Inventory'\n"
                    f"4. En 'Tracking': seleccionar 'By Lots'\n"
                    f"5. Cambiar 'Unidad de Medida' a 'm³'\n"
                    f"6. Guardar y reintentar la operación"
                )
            
            raise UserError(error_msg)
        
        # ✅ Todo correcto
        _logger.info(
            f"✅ Producto validado: {product.display_name} "
            f"(Tipo: {product.type}, Tracking: {product.tracking}, UoM: {product.uom_id.name})"
        )
        return True

    def _log_excel_omissions(self, total_rows, valid_rows):
        """
        Registrar en audit log cuando se omiten líneas del Excel.
        
        Args:
            total_rows: Total de líneas leídas del Excel
            valid_rows: Número de líneas procesadas exitosamente
        
        Returns:
            int: Número de líneas omitidas
        """
        omitted_count = total_rows - valid_rows
        
        if omitted_count > 0:
            # Crear registro en audit log
            try:
                self.env['madenat.audit.log'].create({
                    'reception_id': self.id,
                    'action_type': 'omission',
                    'description': (
                        f"Procesamiento de Excel: {omitted_count} de {total_rows} líneas omitidas.\n"
                        f"Líneas válidas procesadas: {valid_rows}\n"
                        f"Motivo: Datos inválidos, incompletos o con valores cero."
                    ),
                    'user_id': self.env.user.id,
                    'timestamp': fields.Datetime.now(),
                    'batch_id': self.name
                })
                
                _logger.warning(
                    f"⚠️ Omisiones registradas: {omitted_count}/{total_rows} líneas - "
                    f"Recepción: {self.name}"
                )
            except Exception as e:
                _logger.error(f"Error registrando omisiones en audit log: {str(e)}")
            
            # Agregar log visible en la interfaz
            self._add_log(
                f"⚠️ ADVERTENCIA: {omitted_count} de {total_rows} líneas omitidas "
                f"por datos inválidos"
            )
        else:
            _logger.info(f"✅ Sin omisiones: {total_rows} líneas procesadas correctamente")
        
        return omitted_count

    def unlink(self):
        """
        🛡️ BLINDAJE DE CABECERA CON SALVOCONDUCTO TÉCNICO
        v2.0 — Con limpieza de stock.moves huérfanos delegada al servicio
        """
        from .reception_service import LumberReceptionService
        service = LumberReceptionService(self.env)

        # 1. SALVOCONDUCTO
        if self.env.context.get('force_delete'):
            service.cleanup_orphan_moves(self.mapped('name'))
            return super(LumberReception, self).unlink()

        # 2. SEGURIDAD USUARIO
        for rec in self:
            if rec.state not in ('draft', 'cancel'):
                raise UserError(
                    "⛔ SEGURIDAD MADENAT:\n"
                    f"No puede eliminar la guía '{rec.name}' porque está en estado '{rec.state}'.\n"
                    "Solo se pueden eliminar guías en estado 'Borrador' o 'Cancelado'."
                )

            lot_count = self.env['stock.lot'].search_count([
                ('reception_id', '=', rec.id)
            ])
            if lot_count > 0:
                raise UserError(
                    "⛔ SEGURIDAD MADENAT:\n"
                    f"La guía '{rec.name}' ya ha generado {lot_count} lotes en el inventario real.\n"
                    "Por integridad, no se puede eliminar. Archívela o reverse los lotes."
                )

        # 3. LIMPIEZA CASCADA
        service.cleanup_orphan_moves(self.mapped('name'))

        return super(LumberReception, self).unlink()

    def _cleanup_orphan_moves(self):
        """
        🧹 Elimina stock.moves huérfanos (sin picking) generados por esta recepción.
        Método separado para reutilización y testeo independiente.
        """
        names = self.mapped('name')
        moves = self.env['stock.move'].sudo().search([
            ('origin', 'in', names),
            ('picking_id', '=', False),
        ])
        if not moves:
            return

        _logger.info(
            "🧹 MADENAT cleanup: eliminando %d stock.moves huérfanos para guías: %s",
            len(moves), names
        )

        moves.sudo().mapped('move_line_ids').unlink()
        moves.sudo().write({'state': 'draft'})
        moves.sudo().unlink()